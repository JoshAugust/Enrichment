#!/usr/bin/env python3
"""
Merger + XLSX formatter for enriched leads pipeline.
"""

import json
import os
import re
from urllib.parse import urlparse

import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

WS_DIR = "/Users/corgi12/.eragon-joshua_augustine/joshua_augustine_workspace"
OUTPUT_PATH = os.path.join(WS_DIR, "enriched_leads_output.xlsx")

BATCH_FILES = [
    "enriched_batch_1.json",
    "enriched_batch_2.json",
    "enriched_batch_3.json",
    "enriched_batch_4.json",
    "enriched_batch_5.json",
]
PIPELINE_FILE = "pipeline_enriched_results.json"

# ── colours ────────────────────────────────────────────────────────────────
COL_HDR_FILL  = "1F4E79"
COL_HDR_FONT  = "FFFFFF"
COL_ROW_ODD   = "FFFFFF"
COL_ROW_EVEN  = "D6E4F0"

COL_VER_GREEN_FILL = "C6EFCE"; COL_VER_GREEN_FONT = "276221"
COL_VER_YEL_FILL   = "FFEB9C"; COL_VER_YEL_FONT  = "9C5700"
COL_VER_RED_FILL   = "FFC7CE"; COL_VER_RED_FONT  = "9C0006"

COL_SCORE = {5:"C6EFCE", 4:"E2EFDA", 3:"FFEB9C", 2:"FCE4D6", 1:"FFC7CE"}

COL_REC_GREEN  = "C6EFCE"
COL_REC_YEL    = "FFEB9C"
COL_REC_RED    = "FFC7CE"
COL_REC_ORANGE = "FCE4D6"

# ── helpers ────────────────────────────────────────────────────────────────

def load_json(path):
    try:
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return data if isinstance(data, list) else []
    except Exception as e:
        print(f"  WARNING: could not load {path}: {e}")
        return []

def domain_from_url(url):
    if not url:
        return None
    try:
        return urlparse(url).netloc or url
    except Exception:
        return url

def safe(val, default=""):
    if val is None:
        return default
    return val

def normalize_name(name):
    if not name:
        return ""
    return name.strip().lower()

def merge_contacts(existing, incoming):
    """Merge incoming contacts into existing list (dedup by name)."""
    seen = {normalize_name(c.get("name","")) for c in existing if c.get("name")}
    for c in incoming:
        key = normalize_name(c.get("name",""))
        if key and key not in seen:
            existing.append(c)
            seen.add(key)
    return existing

def best_score(a, b):
    try:
        sa = int(a) if a is not None else 0
        sb = int(b) if b is not None else 0
        return sa if sa >= sb else sb
    except Exception:
        return a or b or 0

def make_fill(hex_color):
    return PatternFill(fill_type="solid", fgColor=hex_color)

def make_font(hex_color=None, bold=False, size=10):
    kwargs = {"bold": bold, "size": size}
    if hex_color:
        kwargs["color"] = hex_color
    return Font(**kwargs)

def ver_status_fill_font(status):
    s = (status or "").strip()
    if s == "Verified":
        return make_fill(COL_VER_GREEN_FILL), make_font(COL_VER_GREEN_FONT, size=10)
    elif s == "Discrepancy":
        return make_fill(COL_VER_RED_FILL), make_font(COL_VER_RED_FONT, size=10)
    else:  # Unverified or unknown
        return make_fill(COL_VER_YEL_FILL), make_font(COL_VER_YEL_FONT, size=10)

def score_fill(score):
    try:
        s = int(score)
        return make_fill(COL_SCORE.get(s, "FFFFFF"))
    except Exception:
        return None

def rec_fill(rec_text):
    t = (rec_text or "").lower()
    if t.startswith("pursue"):
        return make_fill(COL_REC_GREEN)
    elif t.startswith("update"):
        return make_fill(COL_REC_YEL)
    elif t.startswith("remove"):
        return make_fill(COL_REC_RED)
    elif t.startswith("fix"):
        return make_fill(COL_REC_ORANGE)
    return None

def make_recommendation(company):
    notes = (safe(company.get("verification_notes","")) + " " +
             safe(company.get("enrichment_notes",""))).lower()
    status = safe(company.get("verification_status","")).strip()
    website_ok = company.get("website_verified", True)

    defunct_words = ["defunct", "acquired", "closed", "out of business", "no longer"]
    if any(w in notes for w in defunct_words):
        return "Remove — company defunct/acquired"

    if not website_ok or status == "Discrepancy":
        # check if website-specific or contact-specific
        if "website" in notes or "url" in notes or "domain" in notes:
            return "Fix website — URL mismatch"
        if "leadership" in notes or "contact" in notes or "title" in notes:
            return "Update contact — leadership changed"
        return "Fix website — URL mismatch"

    if status == "Verified":
        return "Pursue — contact verified"
    elif status == "Unverified":
        return "Update contact — leadership changed"
    return "Pursue — contact verified"

def apply_header_row(ws, headers, col_widths):
    hdr_fill = make_fill(COL_HDR_FILL)
    hdr_font = Font(bold=True, size=11, color=COL_HDR_FONT)
    hdr_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 30

    for col_idx, hdr in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=hdr)
        cell.fill = hdr_fill
        cell.font = hdr_font
        cell.alignment = hdr_align

    # column widths
    for col_idx, width in enumerate(col_widths, start=1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    ws.auto_filter.ref = ws.dimensions
    ws.freeze_panes = ws.cell(row=2, column=1)

def style_data_cell(cell, row_idx, wrap=False, fill_override=None, font_override=None):
    is_even = (row_idx % 2 == 0)
    bg = COL_ROW_EVEN if is_even else COL_ROW_ODD
    cell.fill = fill_override or make_fill(bg)
    cell.font = font_override or make_font(size=10)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)

def write_row(ws, row_idx, values, wrap_cols=None, overrides=None):
    """Write a data row. wrap_cols = set of col indices (1-based). overrides = {col_idx: (fill, font)}"""
    if wrap_cols is None:
        wrap_cols = set()
    if overrides is None:
        overrides = {}
    is_even = (row_idx % 2 == 0)
    bg = COL_ROW_EVEN if is_even else COL_ROW_ODD
    ws.row_dimensions[row_idx].height = 18

    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx, value=val)
        wrap = col_idx in wrap_cols
        fill_ov, font_ov = overrides.get(col_idx, (None, None))
        cell.fill = fill_ov or make_fill(bg)
        cell.font = font_ov or make_font(size=10)
        cell.alignment = Alignment(vertical="top", wrap_text=wrap)


# ══════════════════════════════════════════════════════════════════════════════
# STEP 1  Load all batches
# ══════════════════════════════════════════════════════════════════════════════

print("Loading batch files …")
raw_companies = []
for fname in BATCH_FILES:
    path = os.path.join(WS_DIR, fname)
    batch = load_json(path)
    print(f"  {fname}: {len(batch)} records")
    raw_companies.extend(batch)

print(f"Total before dedup: {len(raw_companies)}")

# Check pipeline file
pipeline_path = os.path.join(WS_DIR, PIPELINE_FILE)
pipeline_data = {}
if os.path.exists(pipeline_path):
    print(f"  Loading {PIPELINE_FILE} …")
    pl_list = load_json(pipeline_path)
    for item in pl_list:
        name = safe(item.get("company_name","")).strip()
        if name:
            pipeline_data[name.lower()] = item
    print(f"  Pipeline records: {len(pipeline_data)}")
else:
    print("  No pipeline_enriched_results.json found — skipping.")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 2  Deduplication & merge
# ══════════════════════════════════════════════════════════════════════════════

print("Deduplicating …")
merged = {}   # key: lower-case name → company dict

for company in raw_companies:
    raw_name = company.get("company_name")
    # Fallback name from website domain
    if not raw_name:
        raw_name = domain_from_url(company.get("website","")) or "Unknown"
    name_key = normalize_name(raw_name)

    if name_key not in merged:
        company["company_name"] = raw_name.strip()
        if not isinstance(company.get("contacts"), list):
            company["contacts"] = []
        merged[name_key] = company
    else:
        existing = merged[name_key]
        # Merge contacts
        incoming_contacts = company.get("contacts") or []
        merge_contacts(existing.setdefault("contacts", []), incoming_contacts)
        # Keep best score
        existing["data_quality_score"] = best_score(
            existing.get("data_quality_score"), company.get("data_quality_score"))
        # Keep better description (longer wins)
        if len(safe(company.get("description",""))) > len(safe(existing.get("description",""))):
            existing["description"] = company.get("description")
            existing["description_source"] = company.get("description_source")
        # Merge notes
        note_tag = f"[Merged duplicate: {company.get('company_name',raw_name)}]"
        existing["enrichment_notes"] = (
            safe(existing.get("enrichment_notes","")) + " " + note_tag).strip()
        # If incoming has better verification
        if existing.get("verification_status") != "Verified" and company.get("verification_status") == "Verified":
            existing["verification_status"] = "Verified"
            existing["verification_notes"] = company.get("verification_notes")

# Merge pipeline data
for name_key, pl_item in pipeline_data.items():
    if name_key in merged:
        existing = merged[name_key]
        # Extra contacts
        merge_contacts(existing.setdefault("contacts", []), pl_item.get("contacts") or [])
        # Extra enrichment fields
        for field in ["recent_news","news_source_url","description","enrichment_notes"]:
            if pl_item.get(field) and not existing.get(field):
                existing[field] = pl_item[field]
        existing["enrichment_notes"] = (
            safe(existing.get("enrichment_notes","")) + " [+pipeline data]").strip()

companies = list(merged.values())
print(f"Total after dedup: {len(companies)}")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 3 + 4  Build XLSX
# ══════════════════════════════════════════════════════════════════════════════

print("Building XLSX …")
wb = openpyxl.Workbook()

# ── Sheet 1: Enriched Leads ──────────────────────────────────────────────────
ws1 = wb.active
ws1.title = "Enriched Leads"
ws1.sheet_properties.tabColor = "2E75B6"

HDR1 = [
    "Company Name","State","Industry","Founded","Employees","Revenue (USD 000s)",
    "Company Description","Contact Name","Contact Title","Is Original Contact",
    "Contact Source URL","LinkedIn URL","Website","Website Verified","Website Notes",
    "Verification Status","Verification Notes","Data Quality Score",
    "Recent News Summary","News Source URL","Enrichment Notes"
]
WIDTHS1 = [35,8,20,10,12,18,50,25,25,12,40,40,40,14,45,18,45,10,45,40,45]
WRAP1   = {7,15,17,19,21}   # 1-based: Description(7), Website Notes(15), VerNotes(17), News(19), EnrNotes(21)

apply_header_row(ws1, HDR1, WIDTHS1)

row1 = 2
total_contacts = 0
for company in companies:
    contacts = company.get("contacts") or []
    # If no contacts at all, write one row with empty contact fields
    if not contacts:
        contacts = [{}]

    for contact in contacts:
        cname = safe(contact.get("name","")).strip()
        if contact and not cname and contact != {}:
            continue  # skip contacts with no name
        total_contacts += 1 if cname else 0

        ver_status = safe(company.get("verification_status","Unverified"))
        score = company.get("data_quality_score")
        score_str = str(score) if score is not None else ""

        values = [
            safe(company.get("company_name","")),
            safe(company.get("state","")),
            safe(company.get("industry","")),
            safe(company.get("founded","")),
            company.get("employees"),
            company.get("revenue_usd_000s"),
            safe(company.get("description","")),
            cname,
            safe(contact.get("title","")),
            "Yes" if contact.get("is_original") else ("No" if contact else ""),
            safe(contact.get("source_url","")),
            safe(contact.get("linkedin_url","")),
            safe(company.get("website","")),
            "Yes" if company.get("website_verified") else "No",
            safe(company.get("website_note","")),
            ver_status,
            safe(company.get("verification_notes","")),
            score_str,
            safe(company.get("recent_news","")),
            safe(company.get("news_source_url","")),
            safe(company.get("enrichment_notes","")),
        ]
        overrides = {}
        vf, vfont = ver_status_fill_font(ver_status)
        overrides[16] = (vf, vfont)  # Verification Status col
        sf = score_fill(score)
        if sf:
            overrides[18] = (sf, make_font(size=10))
        write_row(ws1, row1, values, wrap_cols=WRAP1, overrides=overrides)
        row1 += 1

print(f"  Sheet 1: {row1-2} contact rows written")


# ── Sheet 2: Company Summary ─────────────────────────────────────────────────
ws2 = wb.create_sheet("Company Summary")
ws2.sheet_properties.tabColor = "70AD47"

HDR2 = [
    "Company Name","State","Industry","Founded","Employees","Revenue (USD 000s)",
    "Company Description","Description Source URL","Total Contacts Found",
    "Contact Names & Titles","Website","Website Verified","Verification Status",
    "Verification Notes","Data Quality Score","Recent News","News Source URL",
    "Enrichment Notes"
]
WIDTHS2 = [35,8,20,10,12,18,50,40,14,60,40,14,18,45,12,45,40,45]
WRAP2   = {7,10,14,16,18}  # Description(7), Contacts(10), VerNotes(14), News(16), EnrNotes(18)

apply_header_row(ws2, HDR2, WIDTHS2)

row2 = 2
for company in companies:
    contacts = company.get("contacts") or []
    valid_contacts = [c for c in contacts if c.get("name")]
    contact_str = "; ".join(
        f"{c.get('name','')} — {c.get('title','')}" for c in valid_contacts
    )
    ver_status = safe(company.get("verification_status","Unverified"))
    score = company.get("data_quality_score")
    score_str = f"{score}/5" if score is not None else ""

    values = [
        safe(company.get("company_name","")),
        safe(company.get("state","")),
        safe(company.get("industry","")),
        safe(company.get("founded","")),
        company.get("employees"),
        company.get("revenue_usd_000s"),
        safe(company.get("description","")),
        safe(company.get("description_source","")),
        len(valid_contacts),
        contact_str,
        safe(company.get("website","")),
        "Yes" if company.get("website_verified") else "No",
        ver_status,
        safe(company.get("verification_notes","")),
        score_str,
        safe(company.get("recent_news","")),
        safe(company.get("news_source_url","")),
        safe(company.get("enrichment_notes","")),
    ]
    overrides = {}
    vf, vfont = ver_status_fill_font(ver_status)
    overrides[13] = (vf, vfont)
    sf = score_fill(score)
    if sf:
        overrides[15] = (sf, make_font(size=10))
    write_row(ws2, row2, values, wrap_cols=WRAP2, overrides=overrides)
    row2 += 1

print(f"  Sheet 2: {row2-2} company rows written")


# ── Sheet 3: Verification Report ─────────────────────────────────────────────
ws3 = wb.create_sheet("Verification Report")
ws3.sheet_properties.tabColor = "ED7D31"

HDR3 = [
    "Company Name","State","Original Contact Name","Original Contact Title",
    "Verification Result","Discrepancy Details","Website Issue",
    "Data Quality Score","Recommendation"
]
WIDTHS3 = [35,8,30,30,18,50,40,12,45]
WRAP3   = {6,7,9}

apply_header_row(ws3, HDR3, WIDTHS3)

row3 = 2
for company in companies:
    contacts = company.get("contacts") or []
    orig_contacts = [c for c in contacts if c.get("is_original") and c.get("name")]
    if not orig_contacts:
        orig_contacts = [c for c in contacts if c.get("name")]
    orig = orig_contacts[0] if orig_contacts else {}

    ver_status = safe(company.get("verification_status","Unverified"))
    score = company.get("data_quality_score")
    score_str = str(score) if score is not None else ""

    ver_notes = safe(company.get("verification_notes",""))
    website_issue = ""
    if not company.get("website_verified", True):
        website_issue = safe(company.get("website_note","")) or "Website could not be verified"

    recommendation = make_recommendation(company)

    values = [
        safe(company.get("company_name","")),
        safe(company.get("state","")),
        safe(orig.get("name","")),
        safe(orig.get("title","")),
        ver_status,
        ver_notes,
        website_issue,
        score_str,
        recommendation,
    ]
    overrides = {}
    vf, vfont = ver_status_fill_font(ver_status)
    overrides[5] = (vf, vfont)
    sf = score_fill(score)
    if sf:
        overrides[8] = (sf, make_font(size=10))
    rf = rec_fill(recommendation)
    if rf:
        overrides[9] = (rf, make_font(size=10))
    write_row(ws3, row3, values, wrap_cols=WRAP3, overrides=overrides)
    row3 += 1

print(f"  Sheet 3: {row3-2} company rows written")


# ── Save ─────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
print(f"\nSaved: {OUTPUT_PATH}")
print(f"File size: {os.path.getsize(OUTPUT_PATH):,} bytes")


# ══════════════════════════════════════════════════════════════════════════════
# STEP 5  Summary Stats
# ══════════════════════════════════════════════════════════════════════════════

print("\n" + "="*60)
print("SUMMARY STATS")
print("="*60)

total_cos = len(companies)
total_c = sum(len([c for c in (co.get("contacts") or []) if c.get("name")]) for co in companies)

ver_counts = {"Verified":0, "Unverified":0, "Discrepancy":0}
score_dist = {1:0, 2:0, 3:0, 4:0, 5:0}
defunct_count = 0
website_mismatch = 0

for co in companies:
    vs = safe(co.get("verification_status","Unverified")).strip()
    if vs in ver_counts:
        ver_counts[vs] += 1
    else:
        ver_counts["Unverified"] += 1

    try:
        s = int(co.get("data_quality_score", 0))
        if s in score_dist:
            score_dist[s] += 1
    except Exception:
        pass

    notes_combined = (safe(co.get("verification_notes","")) + " " +
                      safe(co.get("enrichment_notes",""))).lower()
    defunct_words = ["defunct", "acquired", "closed", "out of business", "no longer"]
    if any(w in notes_combined for w in defunct_words):
        defunct_count += 1

    if not co.get("website_verified", True):
        website_mismatch += 1

print(f"Total companies (after dedup): {total_cos}")
print(f"Total contacts across all companies: {total_c}")
print(f"\nVerification breakdown:")
for k, v in ver_counts.items():
    print(f"  {k}: {v}")
print(f"\nData quality distribution:")
for k in sorted(score_dist):
    print(f"  Score {k}: {score_dist[k]}")
print(f"\nCompanies flagged as defunct/acquired/closed: {defunct_count}")
print(f"Companies with website mismatches: {website_mismatch}")
print("="*60)
