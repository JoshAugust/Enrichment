#!/usr/bin/env python3
"""Transform DealScope Excel per Tribal Chief's requirements."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re
from copy import copy

INPUT = '/Users/corgi12/.eragon-joshua_augustine/media/inbound/DealScope_Enriched_1---d14121b1-e382-4ec5-a27b-26362af15009.xlsx'
OUTPUT = '/Users/corgi12/.eragon-joshua_augustine/joshua_augustine_workspace/DealScope_Enriched_v2.xlsx'

wb = openpyxl.load_workbook(INPUT)

# ============================================================
# ENRICHED LEADS SHEET
# ============================================================
ws = wb['Enriched Leads']

# Step 1: Read all data (skip row 1=header, row 2=description row)
headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column + 1)]
print(f"Headers: {headers}")

rows_data = []
for r in range(3, ws.max_row + 1):  # Skip description row (row 2)
    row = {}
    for c in range(1, ws.max_column + 1):
        row[headers[c-1]] = ws.cell(row=r, column=c).value
    rows_data.append(row)

print(f"Loaded {len(rows_data)} data rows from Enriched Leads")

# Step 2: Convert score breakdown to natural language
def natural_language_score(breakdown, score):
    """Convert '.ai domain (+30) | AI/Data keyword...' to readable English."""
    if not breakdown or not isinstance(breakdown, str):
        return "No score breakdown available."
    
    parts = [p.strip() for p in breakdown.split('|')]
    sentences = []
    
    for part in parts:
        # Domain scoring
        m = re.match(r'\.(\w+) domain \(\+(\d+)\)', part)
        if m:
            domain, pts = m.group(1), m.group(2)
            sentences.append(f"Uses a .{domain} domain (+{pts} pts)")
            continue
        
        # Keyword scoring  
        m = re.match(r"([\w/]+) keyword '(\w+)' \(\+(\d+)\)", part)
        if m:
            cat, kw, pts = m.group(1), m.group(2), m.group(3)
            sentences.append(f"industry keyword '{kw}' detected (+{pts} pts)")
            continue
        
        # Deal size band
        m = re.match(r'(Sweet spot|Adjacent band|Outlier) \(\+(\d+)\):\s*~\$([\d,]+)\s*est\.\s*policy\s*\(0\.2%\s*of\s*\$([\d,]+)\s*rev\)', part)
        if m:
            band, pts, policy, rev = m.group(1), m.group(2), m.group(3), m.group(4)
            band_desc = {
                'Sweet spot': 'Revenue falls in the sweet spot',
                'Adjacent band': 'Revenue is in the adjacent band',
                'Outlier': 'Revenue is outside the ideal range'
            }.get(band, band)
            sentences.append(f"{band_desc} (+{pts} pts), estimated policy ~${policy}")
            continue
        
        # No keywords
        m = re.match(r'No high-signal keywords \(\+0\)', part)
        if m:
            sentences.append("no high-signal industry keywords found (+0 pts)")
            continue
        
        # Fallback
        sentences.append(part)
    
    grade = 'A' if score and score >= 85 else 'B' if score and score >= 60 else 'C'
    result = ". ".join(sentences) + f". Overall grade: {grade}."
    return result

# Step 3: Determine startup stage
def get_startup_stage(row):
    """Infer startup stage from available signals."""
    desc = str(row.get('Company Description', '') or '')
    news = str(row.get('Recent News', '') or '')
    notes = str(row.get('Enrichment Notes', '') or '')
    revenue = row.get('Revenue (USD 000s)') or 0
    employees = row.get('Employees') or 0
    founded = row.get('Founded')
    combined = f"{desc} {news} {notes}".lower()
    
    # Check for public company
    if any(kw in combined for kw in ['nasdaq', 'nyse', 'publicly traded', 'ipo', 'public company', 'stock exchange']):
        return 'Public'
    
    # Check for specific funding rounds mentioned
    for stage in ['Series F', 'Series E', 'Series D', 'Series C', 'Series B', 'Series A', 'Seed', 'Pre-Seed']:
        if stage.lower() in combined:
            return stage
    
    # Check for growth/late stage signals
    if any(kw in combined for kw in ['growth financing', 'growth equity', 'late stage', 'pe-backed', 'private equity']):
        return 'Growth/Late'
    
    # Check for bootstrapped
    if any(kw in combined for kw in ['bootstrapped', 'bootstrap', 'self-funded', 'family-owned']):
        return 'Bootstrapped'
    
    # Infer from revenue + employees
    try:
        rev = float(revenue)
    except (ValueError, TypeError):
        rev = 0
    try:
        emp = int(employees)
    except (ValueError, TypeError):
        emp = 0
    
    if rev > 50000 or emp > 500:  # >$50M rev or >500 employees
        return 'Late Stage (est.)'
    elif rev > 10000 or emp > 100:
        return 'Growth (est.)'
    elif rev > 2000 or emp > 20:
        return 'Early Stage (est.)'
    else:
        return 'Seed/Early (est.)'

# Step 4: Extract year from source text for all source columns
def add_year_to_source(source_text):
    """Ensure source text has a year marker in brackets."""
    if not source_text or not isinstance(source_text, str):
        return source_text
    
    # Already has a year in brackets?
    if re.search(r'\(\d{4}\)', source_text):
        return source_text
    
    # Find any year mentioned in text
    years = re.findall(r'20[12]\d', source_text)
    if years:
        latest = max(years)
        return f"{source_text} ({latest})"
    
    # If source is just "DealScope" or "Batch DealScope" with no year
    if source_text.strip().lower() in ['dealscope', 'batch dealscope']:
        return f"{source_text} (2025)"
    
    # Default: mark as undated
    return source_text

# Step 5: Revenue thousands to millions
def rev_to_millions(rev_thousands):
    if rev_thousands is None:
        return None
    try:
        val = float(rev_thousands)
        millions = val / 1000
        if millions >= 1:
            return round(millions, 1)
        else:
            return round(millions, 2)
    except (ValueError, TypeError):
        return rev_thousands

# Step 6: Email fallback - use company email if no personal
def get_company_email(website):
    """Generate a company email from website domain."""
    if not website:
        return None
    domain = str(website).replace('http://', '').replace('https://', '').replace('www.', '').split('/')[0]
    return f"info@{domain}"

# ============================================================
# PROCESS ALL ROWS
# ============================================================

for row in rows_data:
    # Natural language score
    row['Score Breakdown'] = natural_language_score(row.get('Score Breakdown'), row.get('Blueprint Score'))
    
    # Startup stage
    row['Startup Stage'] = get_startup_stage(row)
    
    # Description source (new column)
    desc = str(row.get('Company Description', '') or '')
    if 'nasdaq' in desc.lower() or 'nyse' in desc.lower():
        row['Description Source'] = 'Public filings + web research'
    elif row.get('Website Verified') == 'Yes':
        row['Description Source'] = 'Company website + web research'
    else:
        row['Description Source'] = 'Web research'
    
    # Year-tag all source columns
    for src_col in ['Employee Source', 'Revenue Source', 'Email Source', 'Phone Source', 'News Source']:
        row[src_col] = add_year_to_source(row.get(src_col))
    
    # Revenue to millions
    row['Revenue (USD M)'] = rev_to_millions(row.get('Revenue (USD 000s)'))
    
    # Email fallback
    email = row.get('Email Address')
    if not email or email == 'None' or email == 'Unable to verify':
        company_email = get_company_email(row.get('Website'))
        if company_email:
            row['Email Address'] = company_email
            row['Email Source'] = f"Company general email (no personal email found)"
            row['Email Type'] = 'Company'
        else:
            row['Email Type'] = 'None'
    else:
        row['Email Type'] = 'Personal'

# ============================================================
# SORT: by Blueprint Score desc, then group by company
# ============================================================
# First sort by score descending
rows_data.sort(key=lambda r: -(r.get('Blueprint Score') or 0))

# Within same score, group by company name
from itertools import groupby

# Group by company, keeping order by highest score in group
companies = {}
for row in rows_data:
    name = row.get('Company Name', '')
    if name not in companies:
        companies[name] = []
    companies[name].append(row)

# Sort company groups by their max score
sorted_companies = sorted(companies.items(), key=lambda x: -(max(r.get('Blueprint Score') or 0 for r in x[1])))

final_rows = []
for company_name, company_rows in sorted_companies:
    final_rows.extend(company_rows)

print(f"Processed {len(final_rows)} rows, {len(sorted_companies)} companies")

# ============================================================
# WRITE NEW ENRICHED LEADS SHEET
# ============================================================
# Define new column order
new_headers = [
    'Blueprint Score', 'Startup Stage', 'Score Breakdown',
    'Company Name', 'Company Description', 'Description Source',
    'State', 'Industry / Vertical', 'Website', 'Website Verified',
    'Founded', 'Employees', 'Employee Source',
    'Revenue (USD M)', 'Revenue Confidence', 'Revenue Source',
    'Contact Name', 'Contact Title',
    'Email Address', 'Email Type', 'Email Source',
    'LinkedIn URL', 'Direct Phone', 'Phone Source',
    'Company LinkedIn', 'Verification Status', 'Data Quality Score',
    'Hiring Signals', 'Recent News', 'News Source', 'Enrichment Notes'
]

# Clear and rewrite
wb.remove(ws)
ws = wb.create_sheet('Enriched Leads', 0)

# Color scheme
orange_fill = PatternFill(start_color='FF8C00', end_color='FF8C00', fill_type='solid')
blue_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
purple_fill = PatternFill(start_color='7B68EE', end_color='7B68EE', fill_type='solid')
green_fill = PatternFill(start_color='2E8B57', end_color='2E8B57', fill_type='solid')
teal_fill = PatternFill(start_color='008B8B', end_color='008B8B', fill_type='solid')
gold_fill = PatternFill(start_color='DAA520', end_color='DAA520', fill_type='solid')
white_font = Font(color='FFFFFF', bold=True, size=11)
header_align = Alignment(horizontal='center', wrap_text=True)

# Column color groups
score_cols = {'Blueprint Score', 'Startup Stage', 'Score Breakdown'}
company_cols = {'Company Name', 'Company Description', 'Description Source', 'State', 'Industry / Vertical', 'Website', 'Website Verified', 'Founded'}
employee_rev_cols = {'Employees', 'Employee Source', 'Revenue (USD M)', 'Revenue Confidence', 'Revenue Source'}
contact_cols = {'Contact Name', 'Contact Title', 'Email Address', 'Email Type', 'Email Source', 'LinkedIn URL', 'Direct Phone', 'Phone Source', 'Company LinkedIn'}
quality_cols = {'Verification Status', 'Data Quality Score', 'Hiring Signals', 'Recent News', 'News Source', 'Enrichment Notes'}

# Write headers
for c, header in enumerate(new_headers, 1):
    cell = ws.cell(row=1, column=c, value=header)
    cell.font = white_font
    cell.alignment = header_align
    if header in score_cols:
        cell.fill = gold_fill
    elif header in company_cols:
        cell.fill = orange_fill
    elif header in employee_rev_cols:
        cell.fill = purple_fill
    elif header in contact_cols:
        cell.fill = blue_fill
    elif header in quality_cols:
        cell.fill = teal_fill

# Write data
for r_idx, row in enumerate(final_rows, 2):
    for c_idx, header in enumerate(new_headers, 1):
        val = row.get(header)
        ws.cell(row=r_idx, column=c_idx, value=val)

# Column widths
widths = {
    'Blueprint Score': 14, 'Startup Stage': 16, 'Score Breakdown': 55,
    'Company Name': 30, 'Company Description': 50, 'Description Source': 22,
    'State': 6, 'Industry / Vertical': 28, 'Website': 25, 'Website Verified': 12,
    'Founded': 9, 'Employees': 12, 'Employee Source': 35,
    'Revenue (USD M)': 15, 'Revenue Confidence': 16, 'Revenue Source': 40,
    'Contact Name': 25, 'Contact Title': 30,
    'Email Address': 30, 'Email Type': 12, 'Email Source': 30,
    'LinkedIn URL': 35, 'Direct Phone': 16, 'Phone Source': 20,
    'Company LinkedIn': 35, 'Verification Status': 16, 'Data Quality Score': 14,
    'Hiring Signals': 35, 'Recent News': 45, 'News Source': 35, 'Enrichment Notes': 40
}
for c, header in enumerate(new_headers, 1):
    ws.column_dimensions[get_column_letter(c)].width = widths.get(header, 15)

# Freeze top row
ws.freeze_panes = 'A2'
# Auto-filter
ws.auto_filter.ref = f"A1:{get_column_letter(len(new_headers))}{len(final_rows)+1}"

print(f"Wrote Enriched Leads: {len(final_rows)} rows, {len(new_headers)} columns")

# ============================================================
# COMPANY SUMMARY SHEET - Apply same transforms
# ============================================================
ws2 = wb['Company Summary']
sum_headers = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column + 1)]
sum_rows = []
for r in range(2, ws2.max_row + 1):
    row = {}
    for c in range(1, ws2.max_column + 1):
        row[sum_headers[c-1]] = ws2.cell(row=r, column=c).value
    sum_rows.append(row)

for row in sum_rows:
    row['Score Breakdown'] = natural_language_score(row.get('Score Breakdown'), row.get('Blueprint Score'))
    row['Startup Stage'] = get_startup_stage(row)
    row['Description Source'] = 'Company website + web research' if row.get('Website Verified') == 'Yes' else 'Web research'
    for src_col in ['Employee Source', 'Revenue Source', 'News Source']:
        row[src_col] = add_year_to_source(row.get(src_col))
    row['Revenue (USD M)'] = rev_to_millions(row.get('Revenue (USD 000s)'))

# Sort by score descending
sum_rows.sort(key=lambda r: -(r.get('Blueprint Score') or 0))

# New summary headers
sum_new_headers = [
    'Blueprint Score', 'Startup Stage', 'Score Breakdown',
    'Company Name', 'Company Description', 'Description Source',
    'State', 'Industry / Vertical', 'Website', 'Website Verified',
    'Founded', 'Employees', 'Employee Source',
    'Revenue (USD M)', 'Revenue Confidence', 'Revenue Source',
    'Total Contacts', 'Contacts & Titles', 'Company LinkedIn',
    'Verification Status', 'Data Quality Score',
    'Hiring Signals', 'Recent News', 'News Source',
    'Total Emails', 'All Emails', 'All Phones', 'All LinkedIn URLs'
]

wb.remove(wb['Company Summary'])
ws2 = wb.create_sheet('Company Summary')

for c, header in enumerate(sum_new_headers, 1):
    cell = ws2.cell(row=1, column=c, value=header)
    cell.font = white_font
    cell.alignment = header_align
    if header in score_cols:
        cell.fill = gold_fill
    elif header in company_cols:
        cell.fill = orange_fill
    elif header in employee_rev_cols:
        cell.fill = purple_fill
    elif header in contact_cols:
        cell.fill = blue_fill
    elif header in quality_cols:
        cell.fill = teal_fill

for r_idx, row in enumerate(sum_rows, 2):
    for c_idx, header in enumerate(sum_new_headers, 1):
        val = row.get(header)
        ws2.cell(row=r_idx, column=c_idx, value=val)

for c, header in enumerate(sum_new_headers, 1):
    ws2.column_dimensions[get_column_letter(c)].width = widths.get(header, 15)

ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f"A1:{get_column_letter(len(sum_new_headers))}{len(sum_rows)+1}"

print(f"Wrote Company Summary: {len(sum_rows)} rows, {len(sum_new_headers)} columns")

# ============================================================
# SAVE
# ============================================================
wb.save(OUTPUT)
print(f"\nSaved to {OUTPUT}")

# Stats
scores = [r.get('Blueprint Score') or 0 for r in final_rows]
emails_found = sum(1 for r in final_rows if r.get('Email Type') == 'Personal')
company_emails = sum(1 for r in final_rows if r.get('Email Type') == 'Company')
stages = {}
for r in final_rows:
    s = r.get('Startup Stage', 'Unknown')
    stages[s] = stages.get(s, 0) + 1

print(f"\n--- STATS ---")
print(f"Total contacts: {len(final_rows)}")
print(f"Companies: {len(sorted_companies)}")
print(f"Personal emails: {emails_found}")
print(f"Company fallback emails: {company_emails}")
print(f"No email: {sum(1 for r in final_rows if r.get('Email Type') == 'None')}")
print(f"Score range: {min(scores)}-{max(scores)}")
print(f"\nStartup stages:")
for stage, count in sorted(stages.items(), key=lambda x: -x[1]):
    print(f"  {stage}: {count}")
