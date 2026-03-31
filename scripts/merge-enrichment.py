#!/usr/bin/env python3
"""Merge Orange Slice enrichment results back into DealScope file."""

import json
import openpyxl

INPUT = '/Users/corgi12/.eragon-joshua_augustine/media/inbound/DealScope_Enriched_v2---05dd34c5-7497-4f48-b7de-d013dd65535e.xlsx'
OUTPUT = '/Users/corgi12/.eragon-joshua_augustine/joshua_augustine_workspace/DealScope_Enriched_v3.xlsx'
RESULTS = '/Users/corgi12/.eragon-joshua_augustine/joshua_augustine_workspace/enrich-results.json'

with open(RESULTS) as f:
    data = json.load(f)

phone_results = data['phoneResults']
print(f"Phone results to merge: {len(phone_results)}")

# Build lookup: company name -> phone, address
phone_lookup = {}
for company_name, gmaps_data in phone_results.items():
    phone_lookup[company_name.upper().strip()] = {
        'phone': gmaps_data.get('phone'),
        'address': gmaps_data.get('address'),
        'website': gmaps_data.get('website'),
    }

wb = openpyxl.load_workbook(INPUT)

# === ENRICHED LEADS SHEET ===
ws = wb['Enriched Leads']
headers = [ws.cell(row=1, column=c).value for c in range(1, ws.max_column+1)]

name_col = headers.index('Company Name') + 1
phone_col = headers.index('Direct Phone') + 1
phone_src_col = headers.index('Phone Source') + 1

phones_added = 0
for r in range(2, ws.max_row+1):
    company = str(ws.cell(row=r, column=name_col).value or '').upper().strip()
    current_phone = ws.cell(row=r, column=phone_col).value
    
    if not current_phone and company in phone_lookup:
        new_phone = phone_lookup[company].get('phone')
        if new_phone:
            ws.cell(row=r, column=phone_col, value=new_phone)
            ws.cell(row=r, column=phone_src_col, value='Google Maps via Orange Slice (2026)')
            phones_added += 1

print(f"Enriched Leads: {phones_added} phone numbers added")

# === COMPANY SUMMARY SHEET ===
ws2 = wb['Company Summary']
headers2 = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column+1)]

name_col2 = headers2.index('Company Name') + 1
all_phones_col = headers2.index('All Phones') + 1 if 'All Phones' in headers2 else None

summary_phones = 0
if all_phones_col:
    for r in range(2, ws2.max_row+1):
        company = str(ws2.cell(row=r, column=name_col2).value or '').upper().strip()
        current = ws2.cell(row=r, column=all_phones_col).value
        
        if (not current or current == 'None') and company in phone_lookup:
            new_phone = phone_lookup[company].get('phone')
            if new_phone:
                ws2.cell(row=r, column=all_phones_col, value=new_phone)
                summary_phones += 1

print(f"Company Summary: {summary_phones} phone numbers added")

wb.save(OUTPUT)
print(f"\nSaved to {OUTPUT}")

# Stats
total_contacts = ws.max_row - 1
phones_now = sum(1 for r in range(2, ws.max_row+1) if ws.cell(row=r, column=phone_col).value)
print(f"\nPhone coverage: {phones_now}/{total_contacts} ({phones_now*100//total_contacts}%)")
print(f"Improvement: +{phones_added} phones ({phones_added} Grade A companies)")
