#!/usr/bin/env python3
"""Build GPU Operator leads file with DealScope formatting."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT = '/Users/corgi12/.eragon-joshua_augustine/joshua_augustine_workspace/DealScope_GPU_Operators_v1.xlsx'

# ============================================================
# GPU RVG SCORING MODEL FOR OPERATORS
# ============================================================
# Total: 100 points
# 1. GPU Fleet Scale & Type (0-30) - size of GPU fleet, hardware generation
# 2. Financing Profile (0-25) - debt-financed = gold, VC-funded = silver
# 3. Deal Size Fit (0-20) - mid-market sweet spot for Corgi
# 4. Insurance Readiness (0-15) - operational maturity, revenue visibility
# 5. Contact Quality (0-10) - reachability of decision-makers

def score_operator(company):
    """Score a GPU operator for RVG insurance."""
    score = 0
    reasons = []
    
    # 1. GPU Fleet Scale & Type
    gpu = company.get('gpu_fleet', 0)
    hw = company.get('hardware', '').lower()
    
    if gpu >= 50000:
        score += 25
        reasons.append(f"Massive GPU fleet ({gpu:,}+ GPUs) — high policy value (+25)")
    elif gpu >= 10000:
        score += 30
        reasons.append(f"Large fleet ({gpu:,}+ GPUs) — ideal scale for RVG (+30)")
    elif gpu >= 1000:
        score += 20
        reasons.append(f"Substantial fleet ({gpu:,}+ GPUs) — good fit (+20)")
    elif gpu >= 100:
        score += 12
        reasons.append(f"Growing fleet ({gpu:,}+ GPUs) (+12)")
    else:
        score += 8
        reasons.append(f"Fleet size unconfirmed (+8)")
    
    if any(g in hw for g in ['b200', 'gb200', 'b300', 'blackwell']):
        score += 5  # bonus for latest gen
        reasons.append("Latest-gen Blackwell hardware (+5 bonus)")
    elif any(g in hw for g in ['h100', 'h200']):
        score += 3
        reasons.append("Hopper-gen hardware (+3 bonus)")
    
    # 2. Financing Profile
    fin = company.get('financing', '').lower()
    if 'gpu-backed debt' in fin or 'gpu-collateralized' in fin:
        score += 25
        reasons.append("GPU-backed debt financing — RVG directly reduces cost of capital (+25)")
    elif 'debt' in fin or 'credit facility' in fin:
        score += 22
        reasons.append("Debt-financed — strong RVG value proposition (+22)")
    elif 'venture' in fin or 'series' in fin:
        score += 15
        reasons.append("VC-funded — may seek debt next, RVG enables it (+15)")
    elif 'bootstrapped' in fin or 'self-funded' in fin:
        score += 10
        reasons.append("Self-funded — may not need RVG yet (+10)")
    else:
        score += 12
        reasons.append("Financing profile partially known (+12)")
    
    # 3. Deal Size Fit
    rev = company.get('est_gpu_value_m', 0)
    if 50 <= rev <= 500:
        score += 20
        reasons.append(f"~${rev}M estimated GPU assets — sweet spot for Corgi (+20)")
    elif 500 < rev <= 2000:
        score += 15
        reasons.append(f"~${rev}M GPU assets — large but workable (+15)")
    elif rev > 2000:
        score += 8
        reasons.append(f"~${rev}M+ GPU assets — may exceed current capacity (+8)")
    elif rev > 0:
        score += 12
        reasons.append(f"~${rev}M GPU assets — smaller but accessible (+12)")
    else:
        score += 8
        reasons.append("GPU asset value not estimated (+8)")
    
    # 4. Insurance Readiness
    maturity = company.get('maturity', '').lower()
    if 'public' in maturity or 'established' in maturity:
        score += 15
        reasons.append("Operationally mature — insurance-ready (+15)")
    elif 'growth' in maturity or 'scaling' in maturity:
        score += 12
        reasons.append("Growth stage — actively scaling infrastructure (+12)")
    elif 'early' in maturity:
        score += 8
        reasons.append("Early stage — building out (+8)")
    else:
        score += 10
        reasons.append("Maturity estimated (+10)")
    
    # 5. Contact Quality
    contacts = company.get('contacts', [])
    has_email = any(c.get('email') for c in contacts)
    has_senior = any(any(t in str(c.get('title', '')).lower() for t in 
                    ['ceo', 'cfo', 'coo', 'president', 'founder', 'head of', 'vp', 'svp', 'chief'])
                    for c in contacts)
    cq = 0
    if has_email: cq += 4
    if has_senior: cq += 4
    if len(contacts) >= 2: cq += 2
    cq = min(10, cq)
    score += cq
    reasons.append(f"{len(contacts)} contacts, {'senior + email' if has_email and has_senior else 'partial info'} (+{cq})")
    
    score = min(100, score)
    grade = 'A' if score >= 80 else 'B' if score >= 60 else 'C'
    
    nl = ". ".join(reasons) + f". Overall grade: {grade} ({score}/100)."
    return score, grade, nl

# ============================================================
# COMPANY DATA — researched from web
# ============================================================
gpu_companies = [
    {
        'company': 'Crusoe Energy Systems',
        'description': 'AI factory company building renewable-powered GPU data centers. Raised $1.375B Series E at $10B+ valuation. Over $10B in GPU-backed debt facilities. Largest clean-energy AI infrastructure provider. NVIDIA GB200, HGX B200, AMD MI355x deployments.',
        'desc_source': 'Company website + SEC filings + news (2025)',
        'industry': 'Neocloud / AI Infrastructure / Clean Energy',
        'website': 'crusoe.ai',
        'hq': 'San Francisco, CA',
        'founded': 2018,
        'employees': '~400',
        'emp_source': 'LinkedIn (2025)',
        'total_raised': '$2.8B+ equity + $10B+ debt',
        'last_round': 'Series E $1.375B (Oct 2025)',
        'gpu_fleet': 50000,
        'hardware': 'NVIDIA GB200, HGX B200, AMD MI355x, MI300x',
        'est_gpu_value_m': 2000,
        'financing': 'GPU-backed debt ($10B+) + venture equity',
        'maturity': 'Established / Pre-IPO',
        'contacts': [
            {'name': 'Chase Lochmiller', 'title': 'Co-Founder & CEO', 'email': None, 'linkedin': 'https://linkedin.com/in/lochmiller', 'phone': None},
            {'name': 'Cully Cavness', 'title': 'Co-Founder, President & CSO', 'email': None, 'linkedin': 'https://linkedin.com/in/cullycavness', 'phone': None},
            {'name': 'Nadav Eiron', 'title': 'SVP Cloud Engineering', 'email': None, 'linkedin': None, 'phone': None},
        ],
        'company_linkedin': 'https://linkedin.com/company/crusoe-energy-systems',
        'company_phone': None,
    },
    {
        'company': 'Lambda Labs',
        'description': 'The "Superintelligence Cloud" — GPU cloud and deep learning workstation company. Raised $480M Series D at $2.5B valuation. $1.5B GPU-backed facility via sale-leaseback with IDF/Macquarie. ~300 employees. Serves major AI labs.',
        'desc_source': 'Company website + Crunchbase + news (2025)',
        'industry': 'GPU Cloud / AI Infrastructure',
        'website': 'lambda.ai',
        'hq': 'San Francisco, CA',
        'founded': 2012,
        'employees': '~300',
        'emp_source': 'CBInsights (2025)',
        'total_raised': '$800M+ equity + $1.5B debt',
        'last_round': 'Series D $480M (Feb 2025)',
        'gpu_fleet': 20000,
        'hardware': 'NVIDIA H100, H200, B200',
        'est_gpu_value_m': 600,
        'financing': 'GPU-backed debt ($1.5B sale-leaseback) + venture equity',
        'maturity': 'Growth / Pre-IPO',
        'contacts': [
            {'name': 'Stephen Balaban', 'title': 'Co-Founder & CEO', 'email': None, 'linkedin': 'https://linkedin.com/in/sbalaban', 'phone': None},
            {'name': 'Michael Balaban', 'title': 'Co-Founder & CTO', 'email': None, 'linkedin': None, 'phone': None},
            {'name': 'Robert Brooks IV', 'title': 'VP Revenue', 'email': None, 'linkedin': None, 'phone': None},
            {'name': 'Peter Seibold', 'title': 'CFO', 'email': None, 'linkedin': None, 'phone': None},
        ],
        'company_linkedin': 'https://linkedin.com/company/lambda-labs',
        'company_phone': None,
    },
    {
        'company': 'Nscale',
        'description': 'European AI infrastructure giant. Raised $2B Series C (largest in European history) at $14.6B valuation. €1.1B GPU debt facility (Feb 2026). $14B contract with Microsoft for 104,000 NVIDIA GB300 chips. HQ London, data centers in Portugal, Texas.',
        'desc_source': 'Company website + EU-Startups + news (2026)',
        'industry': 'Neocloud / AI Infrastructure / Sovereign Cloud',
        'website': 'nscale.com',
        'hq': 'London, UK',
        'founded': 2021,
        'employees': '~200',
        'emp_source': 'ZoomInfo (2025)',
        'total_raised': '$3.3B+ equity + €1.1B debt',
        'last_round': 'Series C $2B (Mar 2026)',
        'gpu_fleet': 100000,
        'hardware': 'NVIDIA GB300, GB200',
        'est_gpu_value_m': 5000,
        'financing': 'GPU-backed debt (€1.1B DDTL) + venture equity',
        'maturity': 'Growth / Scaling rapidly',
        'contacts': [
            {'name': 'Josh Payne', 'title': 'Founder & CEO', 'email': None, 'linkedin': 'https://linkedin.com/in/josh-payne-nscale', 'phone': None},
        ],
        'company_linkedin': 'https://linkedin.com/company/nscale',
        'company_phone': None,
    },
    {
        'company': 'FluidStack',
        'description': 'AI cloud platform for training and inference. Raised $450M equity (Jan 2026), total $653M. Selected by Anthropic for $50B custom data center partnership in Texas and New York. 136 employees. London HQ.',
        'desc_source': 'Company website + Sacra + Crunchbase (2026)',
        'industry': 'Neocloud / GPU Cloud',
        'website': 'fluidstack.io',
        'hq': 'London, UK',
        'founded': 2018,
        'employees': 136,
        'emp_source': 'Tracxn (Jan 2026)',
        'total_raised': '$653M',
        'last_round': '$450M equity (Jan 2026)',
        'gpu_fleet': 30000,
        'hardware': 'NVIDIA H100, H200, B200',
        'est_gpu_value_m': 800,
        'financing': 'Venture equity + strategic (Anthropic partnership)',
        'maturity': 'Growth / Scaling',
        'contacts': [],
        'company_linkedin': 'https://linkedin.com/company/fluidstack',
        'company_phone': None,
    },
    {
        'company': 'TensorWave',
        'description': 'All-AMD GPU cloud — only neocloud using exclusively AMD Instinct accelerators. Raised $100M Series A led by Magnetar and AMD Ventures. 8,192-GPU cluster in Arizona on AMD MI325X. 111 employees.',
        'desc_source': 'TechCrunch + company website (2025)',
        'industry': 'GPU Cloud / AMD Specialist',
        'website': 'tensorwave.com',
        'hq': 'Las Vegas, NV',
        'founded': 2023,
        'employees': 111,
        'emp_source': 'Tracxn (Feb 2026)',
        'total_raised': '$143M',
        'last_round': 'Series A $100M (May 2025)',
        'gpu_fleet': 8192,
        'hardware': 'AMD MI325X, MI300X',
        'est_gpu_value_m': 150,
        'financing': 'Venture equity (Series A)',
        'maturity': 'Growth / Early scaling',
        'contacts': [
            {'name': 'Piotr Tomasik', 'title': 'CEO & Co-Founder', 'email': None, 'linkedin': 'https://linkedin.com/in/piotr-tomasik', 'phone': None},
        ],
        'company_linkedin': 'https://linkedin.com/company/tensorwave',
        'company_phone': None,
    },
    {
        'company': 'Applied Digital (APLD)',
        'description': 'Public (NASDAQ: APLD) next-gen digital infrastructure operator. HPC/AI data centers + GPU cloud services. $160M strategic financing from NVIDIA. Spinning out cloud business as ChronoScale. Dallas HQ.',
        'desc_source': 'SEC filings + company website (2025)',
        'industry': 'Data Center / GPU Cloud / Public',
        'website': 'applieddigital.com',
        'hq': 'Dallas, TX',
        'founded': 2015,
        'employees': '~200',
        'emp_source': 'SEC filings (2025)',
        'total_raised': '$5.8B (public + private)',
        'last_round': '$160M strategic (NVIDIA, Related Companies)',
        'gpu_fleet': 20000,
        'hardware': 'NVIDIA H100, B200',
        'est_gpu_value_m': 600,
        'financing': 'Public equity + debt + NVIDIA strategic',
        'maturity': 'Public / Established',
        'contacts': [
            {'name': 'Wes Cummins', 'title': 'Chairman & CEO', 'email': None, 'linkedin': 'https://linkedin.com/in/wescummins', 'phone': None},
        ],
        'company_linkedin': 'https://linkedin.com/company/applied-digital-corp',
        'company_phone': None,
    },
    {
        'company': 'Massed Compute',
        'description': 'GPU cloud provider with $300M investment from Digital Alpha + Cisco strategic collaboration. On-demand GPU/CPU instances, bare metal servers, inventory API. Small team (24) but heavily capitalized.',
        'desc_source': 'BusinessWire + Crunchbase (2025)',
        'industry': 'GPU Cloud / HPC',
        'website': 'massedcompute.com',
        'hq': 'Seattle, WA',
        'founded': 2021,
        'employees': 24,
        'emp_source': 'CBInsights (2025)',
        'total_raised': '$322M',
        'last_round': '$300M from Digital Alpha (Aug 2025)',
        'gpu_fleet': 5000,
        'hardware': 'NVIDIA H100, A100',
        'est_gpu_value_m': 120,
        'financing': 'Strategic investment (Digital Alpha)',
        'maturity': 'Growth / Scaling',
        'contacts': [],
        'company_linkedin': 'https://linkedin.com/company/massed-compute',
        'company_phone': None,
    },
    {
        'company': 'Voltage Park / Lightning AI',
        'description': 'Lightning AI merged with Voltage Park (Jan 2026) — "first cloud built for AI." Voltage Park had $500M initial funding from Navigation Fund (Jed McCaleb). 24,000 NVIDIA H100 GPUs across 6 US data centers.',
        'desc_source': 'BusinessWire + Sacra (2026)',
        'industry': 'GPU Cloud / AI Development Platform',
        'website': 'voltagepark.com',
        'hq': 'US (multi-site)',
        'founded': 2023,
        'employees': '~100',
        'emp_source': 'Estimated (2025)',
        'total_raised': '$500M+',
        'last_round': 'Merger with Lightning AI (Jan 2026)',
        'gpu_fleet': 24000,
        'hardware': 'NVIDIA H100 SXM5',
        'est_gpu_value_m': 700,
        'financing': 'Endowment capital (Navigation Fund)',
        'maturity': 'Established / Post-merger',
        'contacts': [
            {'name': 'Ozan Kaya', 'title': 'CEO', 'email': None, 'linkedin': 'https://linkedin.com/in/ozankaya', 'phone': None},
            {'name': 'Saurabh Giri', 'title': 'Chief Product & Technology Officer', 'email': None, 'linkedin': None, 'phone': None},
        ],
        'company_linkedin': 'https://linkedin.com/company/voltage-park',
        'company_phone': None,
    },
    {
        'company': 'NexGen Cloud / Hyperstack',
        'description': 'European GPU-as-a-service. $59M total funding ($45M Series A, Apr 2025). $1B AI Supercloud planned with 20,000 H100s. £72M revenue 2023-2024. Sovereign AI infrastructure focus in Europe. $354M valuation.',
        'desc_source': 'GlobeNewswire + DCD + Tracxn (2025)',
        'industry': 'GPU Cloud / Sovereign AI Infrastructure',
        'website': 'hyperstack.cloud',
        'hq': 'London, UK',
        'founded': 2020,
        'employees': '~50',
        'emp_source': 'Estimated (2025)',
        'total_raised': '$59M',
        'last_round': 'Series A $45M (Apr 2025)',
        'gpu_fleet': 10000,
        'hardware': 'NVIDIA H100, A100',
        'est_gpu_value_m': 250,
        'financing': 'Venture equity (Series A)',
        'maturity': 'Growth / Scaling',
        'contacts': [],
        'company_linkedin': 'https://linkedin.com/showcase/hyperstackcloud',
        'company_phone': None,
    },
    {
        'company': 'Verda (formerly DataCrunch)',
        'description': 'European AI cloud hyperscaler powered by renewable energy. $64M Series A (Sep 2025) with debt from Nordea/Armada. Total $77.5M. Helsinki HQ. NVIDIA A100, H200, GB200 GPUs.',
        'desc_source': 'TechCrunch + DCD + company website (2025)',
        'industry': 'GPU Cloud / Green AI Infrastructure',
        'website': 'verda.com',
        'hq': 'Helsinki, Finland',
        'founded': 2021,
        'employees': '~40',
        'emp_source': 'Estimated (2025)',
        'total_raised': '$77.5M',
        'last_round': 'Series A $64M (Sep 2025)',
        'gpu_fleet': 3000,
        'hardware': 'NVIDIA A100, H200, GB200',
        'est_gpu_value_m': 80,
        'financing': 'Venture equity + debt (Nordea, Danske Bank)',
        'maturity': 'Growth / Early scaling',
        'contacts': [
            {'name': 'Leena Kassinen', 'title': 'CEO', 'email': None, 'linkedin': None, 'phone': None},
        ],
        'company_linkedin': 'https://linkedin.com/company/verda-cloud',
        'company_phone': None,
    },
    {
        'company': 'GMI Cloud',
        'description': 'GPU cloud infrastructure. $82M Series A (Oct 2024), total $93M ($15M equity + $67M debt). Led by Headline Asia + Banpu Next + Wistron. Originally Bitcoin mining, pivoted to AI cloud. New US data center in Colorado.',
        'desc_source': 'TechCrunch + Crunchbase (2024)',
        'industry': 'GPU Cloud / AI Infrastructure',
        'website': 'gmicloud.ai',
        'hq': 'Taipei, Taiwan / US operations',
        'founded': 2022,
        'employees': '~50',
        'emp_source': 'Estimated (2024)',
        'total_raised': '$93M',
        'last_round': 'Series A $82M (Oct 2024)',
        'gpu_fleet': 5000,
        'hardware': 'NVIDIA H100, A100',
        'est_gpu_value_m': 130,
        'financing': 'Venture equity + debt ($67M)',
        'maturity': 'Growth',
        'contacts': [],
        'company_linkedin': 'https://linkedin.com/company/gmi-cloud',
        'company_phone': None,
    },
    {
        'company': 'Foundry / Mithril',
        'description': 'AI-optimized public cloud. $80M seed+A from Lightspeed, Sequoia, Microsoft Ventures. Resellable GPU instances. Founded by ex-Google/AWS engineers. $350M valuation. Rebranded to Mithril (Aug 2025).',
        'desc_source': 'SiliconAngle + Crunchbase (2025)',
        'industry': 'GPU Cloud / AI Platform',
        'website': 'mithril.ai',
        'hq': 'San Francisco, CA',
        'founded': 2023,
        'employees': '~50',
        'emp_source': 'Estimated (2024)',
        'total_raised': '$80M',
        'last_round': 'Seed + Series A $80M (Mar 2024)',
        'gpu_fleet': 5000,
        'hardware': 'NVIDIA H100, A100',
        'est_gpu_value_m': 120,
        'financing': 'Venture equity (Lightspeed, Sequoia)',
        'maturity': 'Early / Growth',
        'contacts': [],
        'company_linkedin': 'https://linkedin.com/company/foundryai',
        'company_phone': None,
    },
    {
        'company': 'Denvr Dataworks',
        'description': 'Canadian sovereign AI infrastructure. GPUs (H200, H100, A100, Intel Gaudi2) from data centers in Canada and USA. Small team (15) but growing. $10.8M raised. Calgary HQ.',
        'desc_source': 'Crunchbase + PitchBook (2024)',
        'industry': 'GPU Cloud / Sovereign Data Centers',
        'website': 'denvr.com',
        'hq': 'Calgary, Canada',
        'founded': 2017,
        'employees': 15,
        'emp_source': 'Crunchbase (Dec 2024)',
        'total_raised': '$10.8M',
        'last_round': 'Undisclosed',
        'gpu_fleet': 1000,
        'hardware': 'NVIDIA H200, H100, A100, Intel Gaudi2',
        'est_gpu_value_m': 30,
        'financing': 'Venture equity',
        'maturity': 'Early / Niche',
        'contacts': [
            {'name': 'David King', 'title': 'CEO & Co-Founder', 'email': None, 'linkedin': None, 'phone': None},
            {'name': 'Geoff Gordon', 'title': 'Co-Founder', 'email': None, 'linkedin': None, 'phone': None},
        ],
        'company_linkedin': 'https://linkedin.com/company/denvr-dataworks',
        'company_phone': None,
    },
    {
        'company': 'Ori Industries / Radiant',
        'description': 'AI cloud provider merged with Brookfield-owned Radiant (Feb 2026). Raised $178M total. $140M raise in Feb 2025. Strategic investment from Wa\'ed Ventures (Saudi Aramco). Distributed GPU cloud across edge locations.',
        'desc_source': 'Sifted + BeBeez + Tracxn (2026)',
        'industry': 'GPU Cloud / Edge AI Infrastructure',
        'website': 'ori.co',
        'hq': 'London, UK',
        'founded': 2018,
        'employees': '~80',
        'emp_source': 'Estimated (2025)',
        'total_raised': '$178M',
        'last_round': '$140M (Feb 2025) + Radiant merger (Feb 2026)',
        'gpu_fleet': 10000,
        'hardware': 'NVIDIA H100, DGX',
        'est_gpu_value_m': 250,
        'financing': 'Venture equity + Brookfield backing',
        'maturity': 'Growth / Post-merger scaling',
        'contacts': [],
        'company_linkedin': 'https://linkedin.com/company/ori-industries',
        'company_phone': None,
    },
    {
        'company': 'Salad Cloud',
        'description': 'Distributed GPU cloud — 60,000+ daily active GPUs from 450,000+ providers in 190+ countries. "Airbnb for GPUs." $27.5M raised. Unique model: consumer GPUs aggregated for inference workloads at $0.02/hr.',
        'desc_source': 'Company website + Crunchbase (2025)',
        'industry': 'Distributed GPU Cloud / Marketplace',
        'website': 'salad.com',
        'hq': 'Baltimore, MD',
        'founded': 2018,
        'employees': '~40',
        'emp_source': 'Estimated (2025)',
        'total_raised': '$27.5M',
        'last_round': 'Series A $17M (2022)',
        'gpu_fleet': 60000,
        'hardware': 'Consumer GPUs (distributed network)',
        'est_gpu_value_m': 50,
        'financing': 'Venture equity',
        'maturity': 'Growth / Established marketplace',
        'contacts': [],
        'company_linkedin': 'https://linkedin.com/company/salad-technologies',
        'company_phone': None,
    },
    {
        'company': 'Hosted.ai',
        'description': 'GPU infrastructure optimization platform — multi-tenant/federated GPU resource sharing. $19M seed (Mar 2026) from Creandum. Founded by ex-NVIDIA engineers. Makes GPUs profitable for operators, cheaper for developers.',
        'desc_source': 'Company website + FinsmeS (2026)',
        'industry': 'GPU Infrastructure Software / Optimization',
        'website': 'hosted.ai',
        'hq': 'US / EMEA / APAC',
        'founded': 2024,
        'employees': '~20',
        'emp_source': 'Estimated (2026)',
        'total_raised': '$19M',
        'last_round': 'Seed $19M (Mar 2026)',
        'gpu_fleet': 0,
        'hardware': 'Platform/software (manages others\' GPUs)',
        'est_gpu_value_m': 0,
        'financing': 'Venture equity (Creandum)',
        'maturity': 'Early / Seed stage',
        'contacts': [],
        'company_linkedin': 'https://linkedin.com/company/hosted-ai',
        'company_phone': None,
    },
    {
        'company': 'Shadeform',
        'description': 'GPU cloud marketplace — connects developers with compute across 20+ cloud providers. YC S23. Hit $1M revenue Jun 2024 with 5-person team. Aggregator model: finds cheapest GPUs across neoclouds.',
        'desc_source': 'Company website + GetLatka (2024)',
        'industry': 'GPU Cloud Marketplace / Aggregator',
        'website': 'shadeform.com',
        'hq': 'San Francisco, CA',
        'founded': 2023,
        'employees': '~10',
        'emp_source': 'GetLatka (2024)',
        'total_raised': 'Undisclosed (YC + angels)',
        'last_round': 'YC S23',
        'gpu_fleet': 0,
        'hardware': 'Marketplace (aggregates others)',
        'est_gpu_value_m': 0,
        'financing': 'Venture equity (YC)',
        'maturity': 'Early / Seed',
        'contacts': [],
        'company_linkedin': 'https://linkedin.com/company/shadeform',
        'company_phone': None,
    },
]

# ============================================================
# SCORE AND BUILD
# ============================================================
for co in gpu_companies:
    co['score'], co['grade'], co['breakdown'] = score_operator(co)

# Sort by score desc
gpu_companies.sort(key=lambda x: -x['score'])

# Filter to 75+ only
qualified = [c for c in gpu_companies if c['score'] >= 75]
others = [c for c in gpu_companies if c['score'] < 75]

print(f"Total companies: {len(gpu_companies)}")
print(f"Score 75+: {len(qualified)}")
print(f"Score <75: {len(others)}")

# Build enriched lead rows (one per contact, or one for company if no contacts)
enriched_rows = []
summary_rows = []

for co in gpu_companies:
    contacts = co.get('contacts', [])
    company_email = f"info@{co['website'].replace('https://','').replace('http://','')}" if co.get('website') else None
    
    summary = {
        'RVG Score': co['score'],
        'Grade': co['grade'],
        'Stage': co.get('maturity', ''),
        'Score Breakdown': co['breakdown'],
        'Company': co['company'],
        'Description': co['description'],
        'Description Source': co['desc_source'],
        'Industry': co['industry'],
        'Website': co['website'],
        'HQ': co['hq'],
        'Founded': co['founded'],
        'Employees': co['employees'],
        'Employee Source': co['emp_source'],
        'Total Raised': co['total_raised'],
        'Last Round': co['last_round'],
        'GPU Fleet': f"{co['gpu_fleet']:,}" if co['gpu_fleet'] else 'N/A',
        'Hardware': co['hardware'],
        'Est. GPU Value ($M)': co['est_gpu_value_m'] or 'N/A',
        'Financing Profile': co['financing'],
        'Company Phone': co.get('company_phone'),
        'Company LinkedIn': co.get('company_linkedin'),
        'Total Contacts': len(contacts),
    }
    summary_rows.append(summary)
    
    if contacts:
        for c in contacts:
            email = c.get('email')
            email_type = 'Personal' if email else 'Company' if company_email else 'None'
            if not email:
                email = company_email
            
            row = {**summary,
                'Contact Name': c.get('name'),
                'Contact Title': c.get('title'),
                'Email': email,
                'Email Type': email_type,
                'Email Source': 'Web research (2025)' if c.get('email') else 'Company general email',
                'LinkedIn URL': c.get('linkedin'),
                'Direct Phone': c.get('phone'),
            }
            enriched_rows.append(row)
    else:
        row = {**summary,
            'Contact Name': None,
            'Contact Title': None,
            'Email': company_email,
            'Email Type': 'Company' if company_email else 'None',
            'Email Source': 'Company general email',
            'LinkedIn URL': None,
            'Direct Phone': None,
        }
        enriched_rows.append(row)

# ============================================================
# WRITE EXCEL
# ============================================================
wb = openpyxl.Workbook()

# Styling
H_GOLD   = PatternFill('solid', fgColor='F57F17')
H_ORANGE = PatternFill('solid', fgColor='FF5C00')
H_PURPLE = PatternFill('solid', fgColor='6A1B9A')
H_BLUE   = PatternFill('solid', fgColor='1565C0')
H_TEAL   = PatternFill('solid', fgColor='00838F')
H_GREEN  = PatternFill('solid', fgColor='2E7D32')
H_RED    = PatternFill('solid', fgColor='C62828')

D_GREEN  = PatternFill('solid', fgColor='C8E6C9')
D_ORANGE = PatternFill('solid', fgColor='FFF3E0')
D_PURPLE = PatternFill('solid', fgColor='F3E5F5')
D_CONTACT= PatternFill('solid', fgColor='E8EAF6')
D_TEAL   = PatternFill('solid', fgColor='E0F2F1')
D_GOLD   = PatternFill('solid', fgColor='FFF8E1')
D_RED    = PatternFill('solid', fgColor='FFEBEE')

DA_GREEN  = PatternFill('solid', fgColor='A5D6A7')
DA_ORANGE = PatternFill('solid', fgColor='FFE0B2')
DA_PURPLE = PatternFill('solid', fgColor='E1BEE7')
DA_CONTACT= PatternFill('solid', fgColor='C5CAE9')
DA_TEAL   = PatternFill('solid', fgColor='B2DFDB')
DA_GOLD   = PatternFill('solid', fgColor='FFECB3')
DA_RED    = PatternFill('solid', fgColor='FFCDD2')

header_font = Font(color='FFFFFF', bold=True, size=11, name='Calibri')
data_font = Font(size=10, name='Calibri')
data_font_bold = Font(size=10, name='Calibri', bold=True)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
data_align = Alignment(vertical='top', wrap_text=True)
data_align_center = Alignment(horizontal='center', vertical='top', wrap_text=True)
thin_border = Border(left=Side('thin', 'D0D0D0'), right=Side('thin', 'D0D0D0'), top=Side('thin', 'D0D0D0'), bottom=Side('thin', 'D0D0D0'))

EL_COLS = [
    ('RVG Score',          'RVG Score',         13, H_GOLD,   D_GREEN,   DA_GREEN,   True, True),
    ('Grade',              'Grade',              8, H_GOLD,   D_GREEN,   DA_GREEN,   True, True),
    ('Stage',              'Stage',             18, H_GOLD,   D_GOLD,    DA_GOLD,    True, False),
    ('Score Breakdown',    'Score Breakdown',   65, H_GOLD,   D_GOLD,    DA_GOLD,    False,False),
    ('Company',            'Company',           30, H_ORANGE, D_ORANGE,  DA_ORANGE,  False,True),
    ('Description',        'Description',       55, H_ORANGE, D_ORANGE,  DA_ORANGE,  False,False),
    ('Description Source', 'Description Source', 22, H_PURPLE, D_PURPLE,  DA_PURPLE,  False,False),
    ('Industry',           'Industry',          28, H_ORANGE, D_ORANGE,  DA_ORANGE,  False,False),
    ('Website',            'Website',           25, H_TEAL,   D_TEAL,    DA_TEAL,    False,False),
    ('HQ',                 'HQ',                22, H_ORANGE, D_ORANGE,  DA_ORANGE,  False,False),
    ('Founded',            'Founded',           10, H_ORANGE, D_ORANGE,  DA_ORANGE,  True, False),
    ('Employees',          'Employees',         12, H_ORANGE, D_ORANGE,  DA_ORANGE,  True, False),
    ('Employee Source',    'Employee Source',    18, H_PURPLE, D_PURPLE,  DA_PURPLE,  False,False),
    ('Total Raised',       'Total Raised',      22, H_ORANGE, D_ORANGE,  DA_ORANGE,  False,False),
    ('Last Round',         'Last Round',        25, H_ORANGE, D_ORANGE,  DA_ORANGE,  False,False),
    ('GPU Fleet',          'GPU Fleet',         14, H_RED,    D_RED,     DA_RED,     True, True),
    ('Hardware',           'Hardware',          30, H_RED,    D_RED,     DA_RED,     False,False),
    ('Est. GPU Value ($M)','Est. GPU Value ($M)',16,H_RED,    D_RED,     DA_RED,     True, True),
    ('Financing Profile',  'Financing Profile', 35, H_GREEN,  D_GREEN,   DA_GREEN,   False,False),
    ('Contact Name',       'Contact Name',      22, H_BLUE,   D_CONTACT, DA_CONTACT, False,True),
    ('Contact Title',      'Contact Title',     28, H_BLUE,   D_CONTACT, DA_CONTACT, False,False),
    ('Email',              'Email',             28, H_BLUE,   D_CONTACT, DA_CONTACT, False,False),
    ('Email Type',         'Email Type',        12, H_BLUE,   D_CONTACT, DA_CONTACT, True, False),
    ('Email Source',       'Email Source',       22, H_PURPLE, D_PURPLE,  DA_PURPLE,  False,False),
    ('LinkedIn URL',       'LinkedIn URL',      32, H_BLUE,   D_CONTACT, DA_CONTACT, False,False),
    ('Direct Phone',       'Direct Phone',      16, H_BLUE,   D_CONTACT, DA_CONTACT, False,False),
    ('Company Phone',      'Company Phone',     16, H_TEAL,   D_TEAL,    DA_TEAL,    False,False),
    ('Company LinkedIn',   'Company LinkedIn',  32, H_TEAL,   D_TEAL,    DA_TEAL,    False,False),
]

# Enriched Leads sheet
ws = wb.active
ws.title = 'Enriched Leads'

for c, (header, key, width, h_fill, *_) in enumerate(EL_COLS, 1):
    cell = ws.cell(row=1, column=c, value=header)
    cell.font = header_font; cell.fill = h_fill; cell.alignment = header_align; cell.border = thin_border
    ws.column_dimensions[get_column_letter(c)].width = width

current_company = None
is_alt = False
for r_idx, row in enumerate(enriched_rows, 2):
    co = row.get('Company')
    if co != current_company: current_company = co; is_alt = not is_alt
    ws.row_dimensions[r_idx].height = 65
    for c_idx, (h, key, w, hf, df, daf, center, bold) in enumerate(EL_COLS, 1):
        cell = ws.cell(row=r_idx, column=c_idx, value=row.get(key))
        cell.fill = daf if is_alt else df
        cell.font = data_font_bold if bold else data_font
        cell.alignment = data_align_center if center else data_align
        cell.border = thin_border

ws.freeze_panes = 'A2'
ws.auto_filter.ref = f"A1:{get_column_letter(len(EL_COLS))}{len(enriched_rows)+1}"

# Company Summary sheet
CS_COLS = [
    ('RVG Score',          'RVG Score',         13, H_GOLD,   D_GREEN,   DA_GREEN,   True, True),
    ('Grade',              'Grade',              8, H_GOLD,   D_GREEN,   DA_GREEN,   True, True),
    ('Stage',              'Stage',             18, H_GOLD,   D_GOLD,    DA_GOLD,    True, False),
    ('Score Breakdown',    'Score Breakdown',   65, H_GOLD,   D_GOLD,    DA_GOLD,    False,False),
    ('Company',            'Company',           30, H_ORANGE, D_ORANGE,  DA_ORANGE,  False,True),
    ('Description',        'Description',       55, H_ORANGE, D_ORANGE,  DA_ORANGE,  False,False),
    ('Description Source', 'Description Source', 22, H_PURPLE, D_PURPLE,  DA_PURPLE,  False,False),
    ('Industry',           'Industry',          28, H_ORANGE, D_ORANGE,  DA_ORANGE,  False,False),
    ('Website',            'Website',           25, H_TEAL,   D_TEAL,    DA_TEAL,    False,False),
    ('HQ',                 'HQ',                22, H_ORANGE, D_ORANGE,  DA_ORANGE,  False,False),
    ('Founded',            'Founded',           10, H_ORANGE, D_ORANGE,  DA_ORANGE,  True, False),
    ('Employees',          'Employees',         12, H_ORANGE, D_ORANGE,  DA_ORANGE,  True, False),
    ('Employee Source',    'Employee Source',    18, H_PURPLE, D_PURPLE,  DA_PURPLE,  False,False),
    ('Total Raised',       'Total Raised',      22, H_ORANGE, D_ORANGE,  DA_ORANGE,  False,False),
    ('Last Round',         'Last Round',        25, H_ORANGE, D_ORANGE,  DA_ORANGE,  False,False),
    ('GPU Fleet',          'GPU Fleet',         14, H_RED,    D_RED,     DA_RED,     True, True),
    ('Hardware',           'Hardware',          30, H_RED,    D_RED,     DA_RED,     False,False),
    ('Est. GPU Value ($M)','Est. GPU Value ($M)',16,H_RED,    D_RED,     DA_RED,     True, True),
    ('Financing Profile',  'Financing Profile', 35, H_GREEN,  D_GREEN,   DA_GREEN,   False,False),
    ('Company Phone',      'Company Phone',     16, H_TEAL,   D_TEAL,    DA_TEAL,    False,False),
    ('Company LinkedIn',   'Company LinkedIn',  32, H_TEAL,   D_TEAL,    DA_TEAL,    False,False),
    ('Total Contacts',     'Total Contacts',    13, H_BLUE,   D_CONTACT, DA_CONTACT, True, False),
]

ws2 = wb.create_sheet('Company Summary')
for c, (header, key, width, h_fill, *_) in enumerate(CS_COLS, 1):
    cell = ws2.cell(row=1, column=c, value=header)
    cell.font = header_font; cell.fill = h_fill; cell.alignment = header_align; cell.border = thin_border
    ws2.column_dimensions[get_column_letter(c)].width = width

for r_idx, row in enumerate(summary_rows, 2):
    ws2.row_dimensions[r_idx].height = 65
    is_alt = r_idx % 2 == 0
    for c_idx, (h, key, w, hf, df, daf, center, bold) in enumerate(CS_COLS, 1):
        cell = ws2.cell(row=r_idx, column=c_idx, value=row.get(key))
        cell.fill = daf if is_alt else df
        cell.font = data_font_bold if bold else data_font
        cell.alignment = data_align_center if center else data_align
        cell.border = thin_border

ws2.freeze_panes = 'A2'
ws2.auto_filter.ref = f"A1:{get_column_letter(len(CS_COLS))}{len(summary_rows)+1}"

wb.save(OUTPUT)
print(f"\nSaved to {OUTPUT}")
print(f"Enriched Leads: {len(enriched_rows)} rows")
print(f"Company Summary: {len(summary_rows)} companies")

print(f"\n--- ALL COMPANIES BY SCORE ---")
for co in gpu_companies:
    flag = '✅' if co['score'] >= 75 else '⚠️'
    print(f"  {flag} {co['score']:3d} ({co['grade']}) — {co['company']} | GPU fleet: {co['gpu_fleet']:,} | Est value: ${co['est_gpu_value_m']}M | {co['financing'][:50]}")
