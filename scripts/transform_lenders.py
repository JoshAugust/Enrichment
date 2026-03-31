#!/usr/bin/env python3
"""Transform lenders file with GPU RVG-specific scoring and DealScope formatting."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import re

INPUT = '/Users/corgi12/.eragon-joshua_augustine/media/inbound/corgi-lenders---17b2d514-aae3-422b-8473-7b036a32e22d.xlsx'
OUTPUT = '/Users/corgi12/.eragon-joshua_augustine/joshua_augustine_workspace/DealScope_Lenders_v1.xlsx'

wb = openpyxl.load_workbook(INPUT)

# ============================================================
# READ DATA
# ============================================================
ws_co = wb['Lender Companies']
co_headers = [ws_co.cell(row=1, column=c).value for c in range(1, ws_co.max_column + 1)]
companies = {}
for r in range(2, ws_co.max_row + 1):
    row = {}
    for c in range(1, ws_co.max_column + 1):
        row[co_headers[c-1]] = ws_co.cell(row=r, column=c).value
    name = row.get('Company', '')
    companies[name] = row

ws_ct = wb['Lender Contacts']
ct_headers = [ws_ct.cell(row=1, column=c).value for c in range(1, ws_ct.max_column + 1)]
contacts = []
for r in range(2, ws_ct.max_row + 1):
    row = {}
    for c in range(1, ws_ct.max_column + 1):
        row[ct_headers[c-1]] = ws_ct.cell(row=r, column=c).value
    contacts.append(row)

print(f"Loaded {len(companies)} companies, {len(contacts)} contacts")

# ============================================================
# GPU RVG LENDER SCORING MODEL
# ============================================================
# Total: 100 points
# 1. GPU Lending Track Record (0-35 pts)
# 2. Lending Type Fit (0-20 pts)
# 3. Deal Size Alignment (0-20 pts)
# 4. Sector Focus (0-15 pts)
# 5. Contact Quality (0-10 pts)

def score_lender(company, company_contacts):
    """Score a lender for GPU RVG insurance relevance."""
    desc = str(company.get('Description', '') or '').lower()
    segment = str(company.get('Industry Segment', '') or '').lower()
    financing = str(company.get('Financing Status', '') or '').lower()
    gpu_scale = str(company.get('GPU Scale', '') or '').lower()
    name = str(company.get('Company', '') or '').lower()
    employees = str(company.get('Employees', '') or '')
    combined = f"{desc} {segment} {financing} {gpu_scale}"
    
    score = 0
    reasons = []
    
    # --- 1. GPU Lending Track Record (0-35 pts) ---
    gpu_pts = 0
    
    # Confirmed GPU lender / GPU-specific deals
    if any(kw in combined for kw in ['confirmed gpu', 'gpu-backed', 'gpu-collateralized', 
                                       'gpu financing', 'gpu lender', 'gpu hardware',
                                       'gpu infrastructure', 'gpu chips']):
        gpu_pts += 25
        reasons.append("Confirmed GPU lending activity (+25)")
    elif any(kw in combined for kw in ['gpu', 'nvidia']):
        gpu_pts += 15
        reasons.append("GPU/NVIDIA mentions in profile (+15)")
    
    # Named landmark deals (CoreWeave, Lambda, Crusoe etc.)
    landmark_deals = ['coreweave', 'lambda', 'crusoe', 'together ai', 'cerebras']
    deals_found = [d for d in landmark_deals if d in combined]
    if deals_found:
        bonus = min(10, len(deals_found) * 5)
        gpu_pts += bonus
        reasons.append(f"Landmark GPU deals ({', '.join(deals_found)}) (+{bonus})")
    
    gpu_pts = min(35, gpu_pts)
    score += gpu_pts
    
    # --- 2. Lending Type Fit (0-20 pts) ---
    type_pts = 0
    
    if any(kw in segment for kw in ['equipment finance', 'equipment leasing', 'equipment lending']):
        type_pts = 20
        reasons.append("Equipment finance specialist — natural RVG buyer (+20)")
    elif any(kw in segment for kw in ['direct lending', 'private credit']):
        type_pts = 15
        reasons.append("Direct lending / private credit — strong fit (+15)")
    elif any(kw in segment for kw in ['infrastructure', 'digital infrastructure']):
        type_pts = 15
        reasons.append("Infrastructure lending focus (+15)")
    elif any(kw in segment for kw in ['investment bank']):
        type_pts = 10
        reasons.append("Investment bank — can mandate RVG in syndicated deals (+10)")
    elif any(kw in segment for kw in ['venture', 'bdc']):
        type_pts = 8
        reasons.append("Venture/BDC lending (+8)")
    elif any(kw in segment for kw in ['revenue-based', 'alternative']):
        type_pts = 5
        reasons.append("Alternative lending (+5)")
    else:
        type_pts = 3
        reasons.append("General lender (+3)")
    
    score += type_pts
    
    # --- 3. Deal Size Alignment (0-20 pts) ---
    # Sweet spot: mid-market ($50M-$500M deals) — not mega-funds, not micro
    size_pts = 0
    
    # Check for deal size mentions
    deal_amounts = re.findall(r'\$(\d+(?:\.\d+)?)\s*(B|M|billion|million)', combined)
    max_deal = 0
    for amount, unit in deal_amounts:
        val = float(amount)
        if unit.lower() in ['b', 'billion']:
            val *= 1000
        max_deal = max(max_deal, val)
    
    # Employee count as proxy for firm size
    try:
        emp = int(str(employees).replace(',', '').replace('+', '').replace('~', '').split('-')[0].split(' ')[0])
    except:
        emp = 0
    
    if 50 <= max_deal <= 500:
        size_pts = 20
        reasons.append(f"Mid-market deal size (~${int(max_deal)}M) — sweet spot (+20)")
    elif max_deal > 500 and max_deal <= 2000:
        size_pts = 15
        reasons.append(f"Large deal size (~${int(max_deal)}M) — can still mandate RVG (+15)")
    elif max_deal > 2000:
        size_pts = 8
        reasons.append(f"Mega deals (${int(max_deal)}M+) — may be too large for Corgi's current capacity (+8)")
    elif max_deal > 0:
        size_pts = 12
        reasons.append(f"Smaller deal size (~${int(max_deal)}M) — accessible market (+12)")
    else:
        # Use employee count as proxy
        if 10 <= emp <= 500:
            size_pts = 15
            reasons.append("Mid-size firm (employee proxy) — likely mid-market deals (+15)")
        elif emp > 500:
            size_pts = 10
            reasons.append("Large firm — may have mid-market desks (+10)")
        else:
            size_pts = 8
            reasons.append("Small/boutique firm — deal size unknown (+8)")
    
    score += size_pts
    
    # --- 4. Sector Focus (0-15 pts) ---
    sector_pts = 0
    
    if any(kw in combined for kw in ['ai infrastructure', 'data center', 'digital infrastructure', 
                                       'tech infrastructure', 'cloud infrastructure']):
        sector_pts = 15
        reasons.append("AI/data center infrastructure focus (+15)")
    elif any(kw in combined for kw in ['technology', 'tech', 'innovation', 'tmt']):
        sector_pts = 10
        reasons.append("Technology sector focus (+10)")
    elif any(kw in combined for kw in ['asset-backed', 'abf', 'collateral']):
        sector_pts = 10
        reasons.append("Asset-backed lending — understands collateral protection (+10)")
    else:
        sector_pts = 3
        reasons.append("Generalist sector (+3)")
    
    score += sector_pts
    
    # --- 5. Contact Quality (0-10 pts) ---
    contact_pts = 0
    has_email = any(c.get('Email') for c in company_contacts)
    has_phone = any(c.get('Phone') for c in company_contacts)
    has_senior = any(any(t in str(c.get('Title', '')).lower() for t in 
                    ['ceo', 'cfo', 'managing director', 'partner', 'head', 'president', 'svp', 'evp', 'chief'])
                    for c in company_contacts)
    
    if has_email: contact_pts += 3
    if has_phone: contact_pts += 2
    if has_senior: contact_pts += 3
    if len(company_contacts) >= 2: contact_pts += 2
    contact_pts = min(10, contact_pts)
    
    if has_email and has_senior:
        reasons.append(f"Strong contacts: {len(company_contacts)} contacts, senior + email (+{contact_pts})")
    elif has_email:
        reasons.append(f"Has email contacts (+{contact_pts})")
    else:
        reasons.append(f"Limited contact info (+{contact_pts})")
    
    score += contact_pts
    
    # Grade
    if score >= 80:
        grade = 'A'
    elif score >= 60:
        grade = 'B'
    elif score >= 40:
        grade = 'C'
    else:
        grade = 'D'
    
    # Natural language breakdown
    nl = ". ".join(reasons) + f". Overall grade: {grade} ({score}/100)."
    
    return score, grade, nl

# ============================================================
# DETERMINE LENDER STAGE
# ============================================================
def get_lender_stage(company):
    desc = str(company.get('Description', '') or '').lower()
    last_round = str(company.get('Last Funding Round', '') or '').lower()
    total_raised = str(company.get('Total Raised', '') or '').lower()
    founded = company.get('Founded')
    employees = str(company.get('Employees', '') or '')
    combined = f"{desc} {last_round} {total_raised}"
    
    # Banks
    if any(kw in combined for kw in ['investment bank', 'commercial bank', 'bank holding']):
        return 'Established Bank'
    
    # Public
    if any(kw in combined for kw in ['nasdaq', 'nyse', 'publicly traded', 'public company', 'bdc']):
        return 'Public / BDC'
    
    # Large institutional
    try:
        emp = int(str(employees).replace(',', '').replace('+', '').replace('~', '').split('-')[0].split(' ')[0])
    except:
        emp = 0
    
    if emp > 500 or (founded and isinstance(founded, (int, float)) and founded < 2000):
        return 'Established Institution'
    
    # Fund stages
    if any(kw in combined for kw in ['fund iii', 'fund iv', 'fund v', 'fund 3', 'fund 4', 'fund 5']):
        return 'Mature Fund (III+)'
    if any(kw in combined for kw in ['fund ii', 'fund 2']):
        return 'Growth Fund (II)'
    if any(kw in combined for kw in ['fund i', 'fund 1', 'inaugural', 'first fund']):
        return 'Emerging Fund (I)'
    
    # Estimate from age
    if founded and isinstance(founded, (int, float)):
        age = 2026 - int(founded)
        if age >= 15:
            return 'Established Institution'
        elif age >= 7:
            return 'Mature Platform'
        elif age >= 3:
            return 'Growth Platform'
        else:
            return 'Emerging Platform'
    
    return 'Platform (est.)'

# ============================================================
# EMAIL FALLBACK
# ============================================================
def get_company_email(website):
    if not website:
        return None
    domain = str(website).replace('http://', '').replace('https://', '').replace('www.', '').split('/')[0]
    if domain:
        return f"info@{domain}"
    return None

# ============================================================
# PROCESS COMPANIES AND BUILD MERGED ROWS
# ============================================================
# Group contacts by company
contacts_by_company = {}
for c in contacts:
    name = c.get('Company', '')
    if name not in contacts_by_company:
        contacts_by_company[name] = []
    contacts_by_company[name].append(c)

# Process each company
enriched_rows = []
summary_rows = []

for company_name, company in companies.items():
    company_contacts = contacts_by_company.get(company_name, [])
    
    # Score
    new_score, grade, breakdown = score_lender(company, company_contacts)
    stage = get_lender_stage(company)
    
    # Company email for fallback
    company_email = get_company_email(company.get('Website'))
    
    # Description source
    desc = str(company.get('Description', '') or '')
    if 'sec' in desc.lower() or 'publicly' in desc.lower():
        desc_source = 'Public filings + web research'
    else:
        desc_source = 'Web research + industry intelligence'
    
    # Build summary row
    summary = {
        'RVG Score': new_score,
        'Grade': grade,
        'Lender Stage': stage,
        'Score Breakdown': breakdown,
        'Company': company_name,
        'Company Description': company.get('Description'),
        'Description Source': desc_source,
        'Industry Segment': company.get('Industry Segment'),
        'Website': company.get('Website'),
        'HQ': company.get('HQ'),
        'Founded': company.get('Founded'),
        'Employees': company.get('Employees'),
        'Employee Source': 'DealScope (2025)',
        'Total Raised': company.get('Total Raised'),
        'Last Funding Round': company.get('Last Funding Round'),
        'GPU Lending Activity': company.get('GPU Scale') or company.get('Financing Status'),
        'Financing Status': company.get('Financing Status'),
        'Company Phone': company.get('Phone'),
        'Company LinkedIn': company.get('LinkedIn'),
        'Total Contacts': len(company_contacts),
        'Contacts & Titles': '; '.join(
            f"{c.get('Contact Name', '')} — {c.get('Title', '')}"
            for c in company_contacts
        ) if company_contacts else None,
        'All Emails': '; '.join(filter(None, [c.get('Email') for c in company_contacts])) or None,
        'All Phones': '; '.join(filter(None, [str(c.get('Phone', '')) for c in company_contacts if c.get('Phone')])) or None,
        'All LinkedIn URLs': '; '.join(filter(None, [c.get('LinkedIn') for c in company_contacts])) or None,
    }
    summary_rows.append(summary)
    
    # Build enriched lead rows (one per contact)
    if company_contacts:
        for contact in company_contacts:
            email = contact.get('Email')
            email_type = 'Personal'
            email_source = 'DealScope (2025)'
            
            if not email:
                if company_email:
                    email = company_email
                    email_type = 'Company'
                    email_source = 'Company general email (no personal email found)'
                else:
                    email_type = 'None'
            
            row = {
                'RVG Score': new_score,
                'Grade': grade,
                'Lender Stage': stage,
                'Score Breakdown': breakdown,
                'Company': company_name,
                'Company Description': company.get('Description'),
                'Description Source': desc_source,
                'Industry Segment': company.get('Industry Segment'),
                'Website': company.get('Website'),
                'HQ': company.get('HQ'),
                'Founded': company.get('Founded'),
                'Employees': company.get('Employees'),
                'Employee Source': 'DealScope (2025)',
                'Total Raised': company.get('Total Raised'),
                'Last Funding Round': company.get('Last Funding Round'),
                'GPU Lending Activity': company.get('GPU Scale') or company.get('Financing Status'),
                'Financing Status': company.get('Financing Status'),
                'Contact Name': contact.get('Contact Name'),
                'Contact Title': contact.get('Title'),
                'Email Address': email,
                'Email Type': email_type,
                'Email Source': email_source,
                'LinkedIn URL': contact.get('LinkedIn'),
                'Direct Phone': contact.get('Phone'),
                'Phone Source': 'DealScope (2025)' if contact.get('Phone') else None,
                'Company Phone': company.get('Phone'),
                'Company LinkedIn': company.get('LinkedIn'),
            }
            enriched_rows.append(row)
    else:
        # Company with no contacts - still include
        row = {
            'RVG Score': new_score,
            'Grade': grade,
            'Lender Stage': stage,
            'Score Breakdown': breakdown,
            'Company': company_name,
            'Company Description': company.get('Description'),
            'Description Source': desc_source,
            'Industry Segment': company.get('Industry Segment'),
            'Website': company.get('Website'),
            'HQ': company.get('HQ'),
            'Founded': company.get('Founded'),
            'Employees': company.get('Employees'),
            'Employee Source': 'DealScope (2025)',
            'Total Raised': company.get('Total Raised'),
            'Last Funding Round': company.get('Last Funding Round'),
            'GPU Lending Activity': company.get('GPU Scale') or company.get('Financing Status'),
            'Financing Status': company.get('Financing Status'),
            'Contact Name': None,
            'Contact Title': None,
            'Email Address': company_email,
            'Email Type': 'Company' if company_email else 'None',
            'Email Source': 'Company general email' if company_email else None,
            'LinkedIn URL': None,
            'Direct Phone': None,
            'Phone Source': None,
            'Company Phone': company.get('Phone'),
            'Company LinkedIn': company.get('LinkedIn'),
        }
        enriched_rows.append(row)

# Sort by RVG Score descending, group by company
enriched_rows.sort(key=lambda r: (-r.get('RVG Score', 0), r.get('Company', '')))
summary_rows.sort(key=lambda r: -r.get('RVG Score', 0))

print(f"Processed: {len(enriched_rows)} lead rows, {len(summary_rows)} companies")

# ============================================================
# STYLING (matching DealScope format)
# ============================================================
H_GOLD     = PatternFill('solid', fgColor='F57F17')
H_ORANGE   = PatternFill('solid', fgColor='FF5C00')
H_PURPLE   = PatternFill('solid', fgColor='6A1B9A')
H_BLUE     = PatternFill('solid', fgColor='1565C0')
H_TEAL     = PatternFill('solid', fgColor='00838F')
H_GREEN    = PatternFill('solid', fgColor='2E7D32')
H_RED      = PatternFill('solid', fgColor='C62828')

D_GREEN   = PatternFill('solid', fgColor='C8E6C9')
D_ORANGE  = PatternFill('solid', fgColor='FFF3E0')
D_PURPLE  = PatternFill('solid', fgColor='F3E5F5')
D_CONTACT = PatternFill('solid', fgColor='E8EAF6')
D_TEAL    = PatternFill('solid', fgColor='E0F2F1')
D_VERIFY  = PatternFill('solid', fgColor='E8F5E9')
D_GOLD    = PatternFill('solid', fgColor='FFF8E1')
D_RED     = PatternFill('solid', fgColor='FFEBEE')

DA_GREEN   = PatternFill('solid', fgColor='A5D6A7')
DA_ORANGE  = PatternFill('solid', fgColor='FFE0B2')
DA_PURPLE  = PatternFill('solid', fgColor='E1BEE7')
DA_CONTACT = PatternFill('solid', fgColor='C5CAE9')
DA_TEAL    = PatternFill('solid', fgColor='B2DFDB')
DA_VERIFY  = PatternFill('solid', fgColor='C8E6C9')
DA_GOLD    = PatternFill('solid', fgColor='FFECB3')
DA_RED     = PatternFill('solid', fgColor='FFCDD2')

header_font = Font(color='FFFFFF', bold=True, size=11, name='Calibri')
data_font = Font(size=10, name='Calibri')
data_font_bold = Font(size=10, name='Calibri', bold=True)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
data_align = Alignment(vertical='top', wrap_text=True)
data_align_center = Alignment(horizontal='center', vertical='top', wrap_text=True)
thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

# ============================================================
# ENRICHED LEADS SHEET
# ============================================================
EL_COLS = [
    # (header, key, width, h_fill, d_fill, da_fill, center, bold)
    ('RVG Score',            'RVG Score',            13, H_GOLD,   D_GREEN,   DA_GREEN,   True,  True),
    ('Grade',                'Grade',                 8, H_GOLD,   D_GREEN,   DA_GREEN,   True,  True),
    ('Lender Stage',         'Lender Stage',         18, H_GOLD,   D_GOLD,    DA_GOLD,    True,  False),
    ('Score Breakdown',      'Score Breakdown',      60, H_GOLD,   D_GOLD,    DA_GOLD,    False, False),
    ('Company',              'Company',              35, H_ORANGE, D_ORANGE,  DA_ORANGE,  False, True),
    ('Company Description',  'Company Description',  55, H_ORANGE, D_ORANGE,  DA_ORANGE,  False, False),
    ('Description Source',   'Description Source',   20, H_PURPLE, D_PURPLE,  DA_PURPLE,  False, False),
    ('Industry Segment',     'Industry Segment',     28, H_ORANGE, D_ORANGE,  DA_ORANGE,  False, False),
    ('Website',              'Website',              30, H_TEAL,   D_TEAL,    DA_TEAL,    False, False),
    ('HQ',                   'HQ',                   22, H_ORANGE, D_ORANGE,  DA_ORANGE,  False, False),
    ('Founded',              'Founded',              10, H_ORANGE, D_ORANGE,  DA_ORANGE,  True,  False),
    ('Employees',            'Employees',            12, H_ORANGE, D_ORANGE,  DA_ORANGE,  True,  False),
    ('Employee Source',      'Employee Source',       18, H_PURPLE, D_PURPLE,  DA_PURPLE,  False, False),
    ('Total Raised',         'Total Raised',         20, H_ORANGE, D_ORANGE,  DA_ORANGE,  False, False),
    ('Last Funding Round',   'Last Funding Round',   25, H_ORANGE, D_ORANGE,  DA_ORANGE,  False, False),
    ('GPU Lending Activity', 'GPU Lending Activity',  35, H_RED,    D_RED,     DA_RED,     False, False),
    ('Financing Status',     'Financing Status',     18, H_GREEN,  D_VERIFY,  DA_VERIFY,  True,  False),
    ('Contact Name',         'Contact Name',         22, H_BLUE,   D_CONTACT, DA_CONTACT, False, True),
    ('Contact Title',        'Contact Title',        28, H_BLUE,   D_CONTACT, DA_CONTACT, False, False),
    ('Email Address',        'Email Address',        30, H_BLUE,   D_CONTACT, DA_CONTACT, False, False),
    ('Email Type',           'Email Type',           12, H_BLUE,   D_CONTACT, DA_CONTACT, True,  False),
    ('Email Source',         'Email Source',          22, H_PURPLE, D_PURPLE,  DA_PURPLE,  False, False),
    ('LinkedIn URL',         'LinkedIn URL',         32, H_BLUE,   D_CONTACT, DA_CONTACT, False, False),
    ('Direct Phone',         'Direct Phone',         16, H_BLUE,   D_CONTACT, DA_CONTACT, False, False),
    ('Phone Source',         'Phone Source',          18, H_PURPLE, D_PURPLE,  DA_PURPLE,  False, False),
    ('Company Phone',        'Company Phone',        16, H_TEAL,   D_TEAL,    DA_TEAL,    False, False),
    ('Company LinkedIn',     'Company LinkedIn',     32, H_TEAL,   D_TEAL,    DA_TEAL,    False, False),
]

# Remove old sheets and create new ones
for sn in wb.sheetnames:
    wb.remove(wb[sn])

ws_el = wb.create_sheet('Enriched Leads')

# Write headers
for c, (header, key, width, h_fill, *_) in enumerate(EL_COLS, 1):
    cell = ws_el.cell(row=1, column=c, value=header)
    cell.font = header_font
    cell.fill = h_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws_el.column_dimensions[get_column_letter(c)].width = width

# Track company groups for zebra striping
current_company = None
is_alt = False

for r_idx, row in enumerate(enriched_rows, 2):
    company = row.get('Company')
    if company != current_company:
        current_company = company
        is_alt = not is_alt
    
    ws_el.row_dimensions[r_idx].height = 65
    
    for c_idx, (header, key, width, h_fill, d_fill, da_fill, center, bold) in enumerate(EL_COLS, 1):
        cell = ws_el.cell(row=r_idx, column=c_idx, value=row.get(key))
        cell.fill = da_fill if is_alt else d_fill
        cell.font = data_font_bold if bold else data_font
        cell.alignment = data_align_center if center else data_align
        cell.border = thin_border

ws_el.freeze_panes = 'A2'
ws_el.auto_filter.ref = f"A1:{get_column_letter(len(EL_COLS))}{len(enriched_rows)+1}"

print(f"Wrote Enriched Leads: {len(enriched_rows)} rows")

# ============================================================
# COMPANY SUMMARY SHEET
# ============================================================
CS_COLS = [
    ('RVG Score',            'RVG Score',            13, H_GOLD,   D_GREEN,   DA_GREEN,   True,  True),
    ('Grade',                'Grade',                 8, H_GOLD,   D_GREEN,   DA_GREEN,   True,  True),
    ('Lender Stage',         'Lender Stage',         18, H_GOLD,   D_GOLD,    DA_GOLD,    True,  False),
    ('Score Breakdown',      'Score Breakdown',      60, H_GOLD,   D_GOLD,    DA_GOLD,    False, False),
    ('Company',              'Company',              35, H_ORANGE, D_ORANGE,  DA_ORANGE,  False, True),
    ('Company Description',  'Company Description',  55, H_ORANGE, D_ORANGE,  DA_ORANGE,  False, False),
    ('Description Source',   'Description Source',   20, H_PURPLE, D_PURPLE,  DA_PURPLE,  False, False),
    ('Industry Segment',     'Industry Segment',     28, H_ORANGE, D_ORANGE,  DA_ORANGE,  False, False),
    ('Website',              'Website',              30, H_TEAL,   D_TEAL,    DA_TEAL,    False, False),
    ('HQ',                   'HQ',                   22, H_ORANGE, D_ORANGE,  DA_ORANGE,  False, False),
    ('Founded',              'Founded',              10, H_ORANGE, D_ORANGE,  DA_ORANGE,  True,  False),
    ('Employees',            'Employees',            12, H_ORANGE, D_ORANGE,  DA_ORANGE,  True,  False),
    ('Employee Source',      'Employee Source',       18, H_PURPLE, D_PURPLE,  DA_PURPLE,  False, False),
    ('Total Raised',         'Total Raised',         20, H_ORANGE, D_ORANGE,  DA_ORANGE,  False, False),
    ('Last Funding Round',   'Last Funding Round',   25, H_ORANGE, D_ORANGE,  DA_ORANGE,  False, False),
    ('GPU Lending Activity', 'GPU Lending Activity',  35, H_RED,    D_RED,     DA_RED,     False, False),
    ('Financing Status',     'Financing Status',     18, H_GREEN,  D_VERIFY,  DA_VERIFY,  True,  False),
    ('Company Phone',        'Company Phone',        16, H_TEAL,   D_TEAL,    DA_TEAL,    False, False),
    ('Company LinkedIn',     'Company LinkedIn',     32, H_TEAL,   D_TEAL,    DA_TEAL,    False, False),
    ('Total Contacts',       'Total Contacts',       13, H_BLUE,   D_CONTACT, DA_CONTACT, True,  False),
    ('Contacts & Titles',    'Contacts & Titles',    45, H_BLUE,   D_CONTACT, DA_CONTACT, False, False),
    ('All Emails',           'All Emails',           35, H_BLUE,   D_CONTACT, DA_CONTACT, False, False),
    ('All Phones',           'All Phones',           22, H_BLUE,   D_CONTACT, DA_CONTACT, False, False),
    ('All LinkedIn URLs',    'All LinkedIn URLs',    35, H_BLUE,   D_CONTACT, DA_CONTACT, False, False),
]

ws_cs = wb.create_sheet('Company Summary')

for c, (header, key, width, h_fill, *_) in enumerate(CS_COLS, 1):
    cell = ws_cs.cell(row=1, column=c, value=header)
    cell.font = header_font
    cell.fill = h_fill
    cell.alignment = header_align
    cell.border = thin_border
    ws_cs.column_dimensions[get_column_letter(c)].width = width

for r_idx, row in enumerate(summary_rows, 2):
    ws_cs.row_dimensions[r_idx].height = 65
    is_alt = r_idx % 2 == 0
    
    for c_idx, (header, key, width, h_fill, d_fill, da_fill, center, bold) in enumerate(CS_COLS, 1):
        cell = ws_cs.cell(row=r_idx, column=c_idx, value=row.get(key))
        cell.fill = da_fill if is_alt else d_fill
        cell.font = data_font_bold if bold else data_font
        cell.alignment = data_align_center if center else data_align
        cell.border = thin_border

ws_cs.freeze_panes = 'A2'
ws_cs.auto_filter.ref = f"A1:{get_column_letter(len(CS_COLS))}{len(summary_rows)+1}"

print(f"Wrote Company Summary: {len(summary_rows)} rows")

# ============================================================
# SAVE
# ============================================================
wb.save(OUTPUT)
print(f"\nSaved to {OUTPUT}")

# Stats
grades = {}
for r in enriched_rows:
    g = r.get('Grade', '?')
    grades[g] = grades.get(g, 0) + 1

stages = {}
for r in summary_rows:
    s = r.get('Lender Stage', '?')
    stages[s] = stages.get(s, 0) + 1

personal_emails = sum(1 for r in enriched_rows if r.get('Email Type') == 'Personal')
company_emails = sum(1 for r in enriched_rows if r.get('Email Type') == 'Company')
no_emails = sum(1 for r in enriched_rows if r.get('Email Type') == 'None')
scores = [r.get('RVG Score', 0) for r in enriched_rows]

print(f"\n--- STATS ---")
print(f"Total lead rows: {len(enriched_rows)}")
print(f"Companies: {len(summary_rows)}")
print(f"Score range: {min(scores)}-{max(scores)}")
print(f"Grades: {dict(sorted(grades.items()))}")
print(f"Personal emails: {personal_emails}")
print(f"Company fallback emails: {company_emails}")
print(f"No email: {no_emails}")
print(f"\nLender stages:")
for stage, count in sorted(stages.items(), key=lambda x: -x[1]):
    print(f"  {stage}: {count}")

# Top 10
print(f"\n--- TOP 10 LENDERS ---")
for r in summary_rows[:10]:
    print(f"  {r['RVG Score']} ({r['Grade']}) - {r['Company']} [{r['Lender Stage']}]")
