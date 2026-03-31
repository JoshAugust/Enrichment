#!/usr/bin/env python3
"""Build enriched_leads_v5.xlsx from v4 with direct phone population and zero-coverage removal."""

import json
import re
import os
import copy
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

# ── helpers ──────────────────────────────────────────────────────────────────

REMOVE_WORDS = re.compile(
    r'\b(inc|llc|corp|ltd|lp|plc|mr|ms|dr|jr|sr|mrs|pc)\b', re.I
)

def normalize(s):
    if not s:
        return ''
    s = str(s).lower()
    s = re.sub(r'[^\w\s]', ' ', s)   # strip punctuation → space
    s = REMOVE_WORDS.sub('', s)
    s = re.sub(r'\s+', ' ', s).strip()
    return s


def is_blank(v):
    return v is None or str(v).strip() in ('', '—', '-', 'None')


# ── load input files ─────────────────────────────────────────────────────────

WS = '/Users/corgi12/.eragon-joshua_augustine/joshua_augustine_workspace'

with open(f'{WS}/valid_contact_phones.json') as f:
    valid_phones = json.load(f)

with open(f'{WS}/db_phones.json') as f:
    db_phones = json.load(f)

with open(f'{WS}/apollo_emails.json') as f:
    apollo_emails_data = json.load(f)

# Build contact phone lookup: (norm_company, norm_contact) → phone
phone_lookup = {}
for entry in valid_phones:
    key = (normalize(entry['company']), normalize(entry['contact']))
    phone_lookup[key] = entry['phone']

print(f"Phone lookup entries: {len(phone_lookup)}")

# ── load v4 ──────────────────────────────────────────────────────────────────

wb_src = load_workbook(f'{WS}/enriched_leads_v4.xlsx')
ws_src = wb_src['Enriched Leads']

# Map header → column index (0-based)
headers = [cell.value for cell in ws_src[1]]
print(f"Headers ({len(headers)}): {headers}")

# Column indices (0-based)
COL = {h: i for i, h in enumerate(headers) if h}
print(f"Column map sample: Company Phone={COL.get('Company Phone')}, Direct Phone={COL.get('Direct Phone / Extension')}")

IDX_COMPANY   = COL['Company Name']
IDX_CONTACT   = COL['Contact Name']
IDX_CO_PHONE  = COL['Company Phone']
IDX_PH_SRC    = COL['Phone Source']
IDX_EMAIL     = COL['Apollo Email']
IDX_DIR_PHONE = COL['Direct Phone / Extension']
IDX_DIR_SRC   = COL['Direct Phone Source']

# Read all data rows as lists (mutable)
data_rows = []
for row in ws_src.iter_rows(min_row=2, values_only=True):
    data_rows.append(list(row))

print(f"Data rows loaded: {len(data_rows)}")

# ── Step 2+3 — populate direct phone column ──────────────────────────────────

matched = 0
same_as_main = 0
unchanged = 0

for row in data_rows:
    company = row[IDX_COMPANY]
    contact = row[IDX_CONTACT]
    co_phone = row[IDX_CO_PHONE]

    key = (normalize(company), normalize(contact))
    found_phone = phone_lookup.get(key)

    if found_phone:
        # Compare to company phone
        co_phone_str = str(co_phone).strip() if not is_blank(co_phone) else ''
        # normalise phone digits for comparison
        def digits(p):
            return re.sub(r'\D', '', str(p) if p else '')

        if co_phone_str and digits(found_phone) == digits(co_phone_str):
            row[IDX_DIR_PHONE] = f"Same as main: {found_phone}"
            row[IDX_DIR_SRC]   = "Corgi Pipeline DB"
            same_as_main += 1
        else:
            row[IDX_DIR_PHONE] = found_phone
            row[IDX_DIR_SRC]   = "Corgi Pipeline DB"
            matched += 1
    else:
        # keep as —
        if is_blank(row[IDX_DIR_PHONE]):
            row[IDX_DIR_PHONE] = '—'
        row[IDX_DIR_SRC] = row[IDX_DIR_SRC]  # preserve existing
        unchanged += 1

print(f"Direct phone: {matched} real matches, {same_as_main} same-as-main, {unchanged} unchanged")

# ── Step 4 — identify zero-coverage companies ────────────────────────────────

# Group rows by company
company_rows = {}
for row in data_rows:
    cname = row[IDX_COMPANY]
    company_rows.setdefault(cname, []).append(row)

zero_coverage_companies = []
for cname, rows in company_rows.items():
    has_co_phone = any(not is_blank(r[IDX_CO_PHONE]) for r in rows)
    has_email    = any(not is_blank(r[IDX_EMAIL])    for r in rows)
    has_direct   = any(
        not is_blank(r[IDX_DIR_PHONE]) and not str(r[IDX_DIR_PHONE]).startswith('Same as main')
        for r in rows
    )
    if not has_co_phone and not has_email and not has_direct:
        zero_coverage_companies.append(cname)

print(f"\nZero-coverage companies to remove ({len(zero_coverage_companies)}):")
for c in zero_coverage_companies:
    print(f"  - {c}")

# Filter data rows
kept_rows  = [r for r in data_rows if r[IDX_COMPANY] not in zero_coverage_companies]
removed_rows_by_company = {c: company_rows[c] for c in zero_coverage_companies}

print(f"\nKept rows: {len(kept_rows)}  |  Removed rows: {len(data_rows) - len(kept_rows)}")

# ── Step 5 — build output XLSX ───────────────────────────────────────────────

wb = openpyxl.Workbook()

# ── colour palette ────────────────────────────────────────────────────────────
NAV_FILL  = PatternFill("solid", fgColor="1F4E79")
ALT_FILL  = PatternFill("solid", fgColor="D6E4F0")
WHITE_FILL = PatternFill("solid", fgColor="FFFFFF")

VERIFIED_FILL    = PatternFill("solid", fgColor="C6EFCE")
VERIFIED_FONT    = Font(color="276221", size=10)
UNVERIFIED_FILL  = PatternFill("solid", fgColor="FFEB9C")
UNVERIFIED_FONT  = Font(color="9C5700", size=10)
DISCREPANCY_FILL = PatternFill("solid", fgColor="FFC7CE")
DISCREPANCY_FONT = Font(color="9C0006", size=10)

QS_FILLS = {
    5: PatternFill("solid", fgColor="C6EFCE"),
    4: PatternFill("solid", fgColor="E2EFDA"),
    3: PatternFill("solid", fgColor="FFEB9C"),
    2: PatternFill("solid", fgColor="FCE4D6"),
    1: PatternFill("solid", fgColor="FFC7CE"),
}

BLUE_FONT  = Font(color="0563C1", size=10)
GREY_FONT  = Font(color="999999", size=10)
GREEN_FONT = Font(color="276221", size=10)
WHITE_FONT = Font(color="FFFFFF", bold=True, size=11)
BASE_FONT  = Font(size=10)


def apply_header_row(ws, row_num=1):
    """Bold white on navy for header row."""
    for cell in ws[row_num]:
        cell.fill = NAV_FILL
        cell.font = WHITE_FONT
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    ws.row_dimensions[row_num].height = 30


def apply_data_row(ws, row_num, alt=False):
    fill = ALT_FILL if alt else WHITE_FILL
    for cell in ws[row_num]:
        if cell.fill.patternType is None or cell.fill.patternType == 'none':
            cell.fill = fill
        if cell.font.name is None or cell.font == Font():
            cell.font = BASE_FONT
    ws.row_dimensions[row_num].height = 18


def write_cell(ws, row, col, value, font=None, fill=None, wrap=False):
    """col is 1-based."""
    c = ws.cell(row=row, column=col, value=value)
    if font:
        c.font = font
    if fill:
        c.fill = fill
    if wrap:
        c.alignment = Alignment(wrap_text=True, vertical='top')
    else:
        c.alignment = Alignment(vertical='top')
    return c


# ── Sheet 1: Enriched Leads ───────────────────────────────────────────────────

ws1 = wb.active
ws1.title = 'Enriched Leads'

# Write header
for col_idx, h in enumerate(headers, 1):
    write_cell(ws1, 1, col_idx, h)
apply_header_row(ws1, 1)
ws1.freeze_panes = 'A2'

# Column widths (match v4 roughly)
col_widths = {
    1: 35, 2: 8, 3: 22, 4: 10, 5: 12, 6: 15, 7: 50,
    8: 25, 9: 30, 10: 10, 11: 35, 12: 35, 13: 30, 14: 12,
    15: 14, 16: 50, 17: 8, 18: 40, 19: 50, 20: 35, 21: 50,
    22: 18, 23: 25, 24: 30, 25: 35, 26: 20, 27: 20,
}
for col_n, width in col_widths.items():
    ws1.column_dimensions[get_column_letter(col_n)].width = width

# Verification status col (O = col 15), Quality Score (Q = col 17)
# Apollo Email (X = col 24), Direct Phone (Z = col 26)
IDX_VSTATUS = COL.get('Verification Status', 14)   # 0-based
IDX_QSCORE  = COL.get('Data Quality Score', 16)
IDX_APOLLO  = IDX_EMAIL
IDX_DIRPH   = IDX_DIR_PHONE

for r_idx, row in enumerate(kept_rows, 2):
    alt = (r_idx % 2 == 1)  # alternate starting row 2
    bg_fill = ALT_FILL if alt else WHITE_FILL

    for col_idx, val in enumerate(row, 1):
        c = ws1.cell(row=r_idx, column=col_idx, value=val)
        c.fill = bg_fill
        c.font = BASE_FONT
        c.alignment = Alignment(vertical='top', wrap_text=(col_idx in (7, 16, 18, 19, 21)))

        zero_idx = col_idx - 1  # 0-based

        # Verification Status (col 15, 0-based 14)
        if zero_idx == IDX_VSTATUS:
            if val == 'Verified':
                c.fill = VERIFIED_FILL
                c.font = Font(color="276221", size=10, bold=True)
            elif val == 'Unverified':
                c.fill = UNVERIFIED_FILL
                c.font = Font(color="9C5700", size=10, bold=True)
            elif val and 'Discrepancy' in str(val):
                c.fill = DISCREPANCY_FILL
                c.font = Font(color="9C0006", size=10, bold=True)

        # Quality Score (col 17, 0-based 16)
        elif zero_idx == IDX_QSCORE:
            try:
                qs = int(val)
                if qs in QS_FILLS:
                    c.fill = QS_FILLS[qs]
            except (TypeError, ValueError):
                pass

        # Apollo Email (col 24, 0-based 23)
        elif zero_idx == IDX_APOLLO:
            if is_blank(val):
                c.value = '—'
                c.font = GREY_FONT
            else:
                c.font = BLUE_FONT

        # Direct Phone (col 26, 0-based 25)
        elif zero_idx == IDX_DIRPH:
            if is_blank(val) or val == '—':
                c.value = '—'
                c.font = GREY_FONT
            elif str(val).startswith('Same as main'):
                c.font = GREY_FONT
            else:
                # real direct line
                c.font = GREEN_FONT

        # Direct Phone Source (col 27, 0-based 26)
        elif zero_idx == IDX_DIR_SRC:
            if is_blank(val):
                c.value = '—'
                c.font = GREY_FONT

    ws1.row_dimensions[r_idx].height = 18

ws1.auto_filter.ref = f"A1:{get_column_letter(len(headers))}{len(kept_rows)+1}"

print(f"\nSheet 1 written: {len(kept_rows)} data rows")

# ── Sheet 2: Company Summary ──────────────────────────────────────────────────

ws_sum_src = wb_src['Company Summary']
ws2 = wb.create_sheet('Company Summary')

sum_headers = [cell.value for cell in ws_sum_src[1]]
removed_set = set(zero_coverage_companies)

# Write header
for col_idx, h in enumerate(sum_headers, 1):
    write_cell(ws2, 1, col_idx, h)
apply_header_row(ws2, 1)
ws2.freeze_panes = 'A2'

# Column widths for summary
sum_widths = {1:35,2:8,3:22,4:10,5:12,6:15,7:50,8:35,9:12,10:60,11:30,12:12,13:35,14:14,15:50,16:8,17:40,18:50,19:35,20:50,21:18,22:25,23:14,24:50}
for col_n, width in sum_widths.items():
    ws2.column_dimensions[get_column_letter(col_n)].width = width

# Map company summary column indices
sum_col = {h: i for i, h in enumerate(sum_headers) if h}
S_COMPANY = sum_col.get('Company Name', 0)
S_VSTATUS = sum_col.get('Verification Status', 13)
S_QSCORE  = sum_col.get('Data Quality Score', 15)

written_sum = 0
for r_idx_src, row_src in enumerate(ws_sum_src.iter_rows(min_row=2, values_only=True), 2):
    cname = row_src[S_COMPANY]
    if cname in removed_set:
        continue
    written_sum += 1
    r_out = written_sum + 1
    alt = (written_sum % 2 == 0)
    bg_fill = ALT_FILL if alt else WHITE_FILL

    for col_idx, val in enumerate(row_src, 1):
        c = ws2.cell(row=r_out, column=col_idx, value=val)
        c.fill = bg_fill
        c.font = BASE_FONT
        c.alignment = Alignment(vertical='top', wrap_text=(col_idx in (7, 10, 15, 18, 20)))

        zero_idx = col_idx - 1
        if zero_idx == S_VSTATUS:
            if val == 'Verified':
                c.fill = VERIFIED_FILL
                c.font = Font(color="276221", size=10, bold=True)
            elif val == 'Unverified':
                c.fill = UNVERIFIED_FILL
                c.font = Font(color="9C5700", size=10, bold=True)
            elif val and 'Discrepancy' in str(val):
                c.fill = DISCREPANCY_FILL
                c.font = Font(color="9C0006", size=10, bold=True)
        elif zero_idx == S_QSCORE:
            try:
                qs = int(val)
                if qs in QS_FILLS:
                    c.fill = QS_FILLS[qs]
            except (TypeError, ValueError):
                pass

    ws2.row_dimensions[r_out].height = 18

ws2.auto_filter.ref = f"A1:{get_column_letter(len(sum_headers))}{written_sum+1}"
print(f"Sheet 2 written: {written_sum} company rows")

# ── Sheet 3: Removed Leads ────────────────────────────────────────────────────

ws_rem_src = wb_src['Removed Leads']
ws3 = wb.create_sheet('Removed Leads')

rem_headers = [cell.value for cell in ws_rem_src[1]]
for col_idx, h in enumerate(rem_headers, 1):
    write_cell(ws3, 1, col_idx, h)
apply_header_row(ws3, 1)
ws3.freeze_panes = 'A2'

rem_widths = {1:35,2:8,3:55,4:20,5:60}
for col_n, width in rem_widths.items():
    ws3.column_dimensions[get_column_letter(col_n)].width = width

# Map removal cols
rem_col = {h: i for i, h in enumerate(rem_headers) if h}
R_VSTATUS = rem_col.get('Original Verification Status', 3)

# Copy existing removed entries
written_rem = 0
for row_src in ws_rem_src.iter_rows(min_row=2, values_only=True):
    if all(v is None for v in row_src):
        continue
    written_rem += 1
    r_out = written_rem + 1
    alt = (written_rem % 2 == 0)
    bg_fill = ALT_FILL if alt else WHITE_FILL
    for col_idx, val in enumerate(row_src, 1):
        c = ws3.cell(row=r_out, column=col_idx, value=val)
        c.fill = bg_fill
        c.font = BASE_FONT
        c.alignment = Alignment(vertical='top', wrap_text=(col_idx == 5))
        if col_idx - 1 == R_VSTATUS:
            if val == 'Verified':
                c.fill = VERIFIED_FILL
                c.font = Font(color="276221", size=10, bold=True)
            elif val == 'Unverified':
                c.fill = UNVERIFIED_FILL
                c.font = Font(color="9C5700", size=10, bold=True)
            elif val and 'Discrepancy' in str(val):
                c.fill = DISCREPANCY_FILL
                c.font = Font(color="9C0006", size=10, bold=True)
    ws3.row_dimensions[r_out].height = 18

# Append newly removed zero-coverage companies
# Get their verification status from the source enriched leads sheet
company_vstatus = {}
for row in data_rows:
    c = row[IDX_COMPANY]
    vs = row[COL.get('Verification Status', 14)]
    if c not in company_vstatus and vs:
        company_vstatus[c] = vs

new_removals = 0
for cname in zero_coverage_companies:
    written_rem += 1
    new_removals += 1
    r_out = written_rem + 1
    alt = (written_rem % 2 == 0)
    bg_fill = ALT_FILL if alt else WHITE_FILL

    # Get state from data
    state_val = None
    for row in data_rows:
        if row[IDX_COMPANY] == cname:
            state_val = row[COL.get('State', 1)]
            break

    vs = company_vstatus.get(cname, 'Unverified')
    removal_reason = "No contact information available (no phone, no email)"

    row_data = [cname, state_val, removal_reason, vs, None]
    for col_idx, val in enumerate(row_data, 1):
        c = ws3.cell(row=r_out, column=col_idx, value=val)
        c.fill = bg_fill
        c.font = BASE_FONT
        c.alignment = Alignment(vertical='top', wrap_text=(col_idx == 3))
        if col_idx - 1 == R_VSTATUS:
            if vs == 'Verified':
                c.fill = VERIFIED_FILL
                c.font = Font(color="276221", size=10, bold=True)
            elif vs == 'Unverified':
                c.fill = UNVERIFIED_FILL
                c.font = Font(color="9C5700", size=10, bold=True)
            elif vs and 'Discrepancy' in str(vs):
                c.fill = DISCREPANCY_FILL
                c.font = Font(color="9C0006", size=10, bold=True)
    ws3.row_dimensions[r_out].height = 18

ws3.auto_filter.ref = f"A1:{get_column_letter(len(rem_headers))}{written_rem+1}"
print(f"Sheet 3 written: {written_rem} removed entries ({new_removals} new)")

# ── Save ──────────────────────────────────────────────────────────────────────

out_path = f'{WS}/enriched_leads_v5.xlsx'
wb.save(out_path)
file_size = os.path.getsize(out_path)

# ── Summary ───────────────────────────────────────────────────────────────────

# Count unique companies remaining
remaining_companies = set(r[IDX_COMPANY] for r in kept_rows)
contacts_with_direct = sum(
    1 for r in kept_rows
    if not is_blank(r[IDX_DIR_PHONE])
    and r[IDX_DIR_PHONE] != '—'
    and not str(r[IDX_DIR_PHONE]).startswith('Same as main')
)
contacts_with_email = sum(
    1 for r in kept_rows
    if not is_blank(r[IDX_EMAIL])
)

print(f"""
╔══════════════════════════════════════════════════════╗
║          enriched_leads_v5.xlsx — SUMMARY           ║
╠══════════════════════════════════════════════════════╣
║  Original company count          : 53               ║
║  Companies removed (no coverage) : {len(zero_coverage_companies):<4}               ║
║  Final company count             : {len(remaining_companies):<4}               ║
║  Final contact count             : {len(kept_rows):<4}               ║
║  Contacts with direct phone      : {contacts_with_direct:<4}               ║
║  Contacts with Apollo email      : {contacts_with_email:<4}               ║
║  File size                       : {file_size:,} bytes     ║
╚══════════════════════════════════════════════════════╝
Output: {out_path}
""")
