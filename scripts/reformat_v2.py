#!/usr/bin/env python3
"""Re-format DealScope v2 with proper styling matching original."""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side, numbers
from openpyxl.utils import get_column_letter

INPUT = '/Users/corgi12/.eragon-joshua_augustine/joshua_augustine_workspace/DealScope_Enriched_v2.xlsx'
OUTPUT = '/Users/corgi12/.eragon-joshua_augustine/joshua_augustine_workspace/DealScope_Enriched_v2.xlsx'

wb = openpyxl.load_workbook(INPUT)

# ============================================================
# Original color scheme (from the file the TC sent)
# ============================================================
# Header fills (bold white text on colored bg)
H_GOLD     = PatternFill('solid', fgColor='F57F17')   # Score cols
H_ORANGE   = PatternFill('solid', fgColor='FF5C00')   # Company info
H_PURPLE   = PatternFill('solid', fgColor='6A1B9A')   # Source cols
H_BLUE     = PatternFill('solid', fgColor='1565C0')   # Contact cols
H_TEAL     = PatternFill('solid', fgColor='00838F')   # Intelligence/links
H_GREEN    = PatternFill('solid', fgColor='2E7D32')   # Verification/quality

# Data fills (light tints)
D_GREEN    = PatternFill('solid', fgColor='C8E6C9')   # Score cells
D_BLUE     = PatternFill('solid', fgColor='E3F2FD')   # Default data
D_ORANGE   = PatternFill('solid', fgColor='FFF3E0')   # Company info
D_PURPLE   = PatternFill('solid', fgColor='F3E5F5')   # Source cells
D_CONTACT  = PatternFill('solid', fgColor='E8EAF6')   # Contact cells (indigo tint)
D_TEAL     = PatternFill('solid', fgColor='E0F2F1')   # Intelligence
D_VERIFY   = PatternFill('solid', fgColor='E8F5E9')   # Verification
D_GOLD     = PatternFill('solid', fgColor='FFF8E1')   # Stage/score breakdown

# Alt-row data fills (slightly darker for zebra striping)
DA_GREEN   = PatternFill('solid', fgColor='A5D6A7')
DA_BLUE    = PatternFill('solid', fgColor='BBDEFB')
DA_ORANGE  = PatternFill('solid', fgColor='FFE0B2')
DA_PURPLE  = PatternFill('solid', fgColor='E1BEE7')
DA_CONTACT = PatternFill('solid', fgColor='C5CAE9')
DA_TEAL    = PatternFill('solid', fgColor='B2DFDB')
DA_VERIFY  = PatternFill('solid', fgColor='C8E6C9')
DA_GOLD    = PatternFill('solid', fgColor='FFECB3')

header_font = Font(color='FFFFFF', bold=True, size=11, name='Calibri')
data_font = Font(size=10, name='Calibri')
data_font_bold = Font(size=10, name='Calibri', bold=True)
header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
data_align = Alignment(vertical='top', wrap_text=True)
data_align_center = Alignment(horizontal='center', vertical='top', wrap_text=True)

thin_border = Border(
    left=Side(style='thin', color='D0D0D0'),
    right=Side(style='thin', color='D0D0D0'),
    top=Side(style='thin', color='D0D0D0'),
    bottom=Side(style='thin', color='D0D0D0')
)

# ============================================================
# Column definitions with color groups
# ============================================================
# Enriched Leads columns
EL_COLS = [
    # (header, width, header_fill, data_fill, alt_fill, center?)
    ('Blueprint Score',      14, H_GOLD,   D_GREEN,   DA_GREEN,   True),
    ('Startup Stage',        16, H_GOLD,   D_GOLD,    DA_GOLD,    True),
    ('Score Breakdown',      55, H_GOLD,   D_GOLD,    DA_GOLD,    False),
    ('Company Name',         30, H_ORANGE, D_ORANGE,  DA_ORANGE,  False),
    ('Company Description',  50, H_ORANGE, D_ORANGE,  DA_ORANGE,  False),
    ('Description Source',   20, H_PURPLE, D_PURPLE,  DA_PURPLE,  False),
    ('State',                 6, H_ORANGE, D_ORANGE,  DA_ORANGE,  True),
    ('Industry / Vertical',  25, H_ORANGE, D_ORANGE,  DA_ORANGE,  False),
    ('Website',              28, H_TEAL,   D_TEAL,    DA_TEAL,    False),
    ('Website Verified',     12, H_GREEN,  D_VERIFY,  DA_VERIFY,  True),
    ('Founded',               9, H_ORANGE, D_ORANGE,  DA_ORANGE,  True),
    ('Employees',            11, H_ORANGE, D_ORANGE,  DA_ORANGE,  True),
    ('Employee Source',      25, H_PURPLE, D_PURPLE,  DA_PURPLE,  False),
    ('Revenue (USD M)',      14, H_ORANGE, D_ORANGE,  DA_ORANGE,  True),
    ('Revenue Confidence',   14, H_ORANGE, D_ORANGE,  DA_ORANGE,  True),
    ('Revenue Source',       30, H_PURPLE, D_PURPLE,  DA_PURPLE,  False),
    ('Contact Name',         22, H_BLUE,   D_CONTACT, DA_CONTACT, False),
    ('Contact Title',        24, H_BLUE,   D_CONTACT, DA_CONTACT, False),
    ('Email Address',        28, H_BLUE,   D_CONTACT, DA_CONTACT, False),
    ('Email Type',           12, H_BLUE,   D_CONTACT, DA_CONTACT, True),
    ('Email Source',         20, H_PURPLE, D_PURPLE,  DA_PURPLE,  False),
    ('LinkedIn URL',         32, H_BLUE,   D_CONTACT, DA_CONTACT, False),
    ('Direct Phone',         16, H_BLUE,   D_CONTACT, DA_CONTACT, False),
    ('Phone Source',         22, H_PURPLE, D_PURPLE,  DA_PURPLE,  False),
    ('Company LinkedIn',     32, H_TEAL,   D_TEAL,    DA_TEAL,    False),
    ('Verification Status',  14, H_GREEN,  D_VERIFY,  DA_VERIFY,  True),
    ('Data Quality Score',   13, H_GREEN,  D_VERIFY,  DA_VERIFY,  True),
    ('Hiring Signals',       30, H_TEAL,   D_TEAL,    DA_TEAL,    False),
    ('Recent News',          40, H_TEAL,   D_TEAL,    DA_TEAL,    False),
    ('News Source',          30, H_PURPLE, D_PURPLE,  DA_PURPLE,  False),
    ('Enrichment Notes',     40, H_TEAL,   D_TEAL,    DA_TEAL,    False),
]

# Company Summary columns
CS_COLS = [
    ('Blueprint Score',      14, H_GOLD,   D_GREEN,   DA_GREEN,   True),
    ('Startup Stage',        16, H_GOLD,   D_GOLD,    DA_GOLD,    True),
    ('Score Breakdown',      55, H_GOLD,   D_GOLD,    DA_GOLD,    False),
    ('Company Name',         30, H_ORANGE, D_ORANGE,  DA_ORANGE,  False),
    ('Company Description',  50, H_ORANGE, D_ORANGE,  DA_ORANGE,  False),
    ('Description Source',   20, H_PURPLE, D_PURPLE,  DA_PURPLE,  False),
    ('State',                 6, H_ORANGE, D_ORANGE,  DA_ORANGE,  True),
    ('Industry / Vertical',  25, H_ORANGE, D_ORANGE,  DA_ORANGE,  False),
    ('Website',              28, H_TEAL,   D_TEAL,    DA_TEAL,    False),
    ('Website Verified',     12, H_GREEN,  D_VERIFY,  DA_VERIFY,  True),
    ('Founded',               9, H_ORANGE, D_ORANGE,  DA_ORANGE,  True),
    ('Employees',            11, H_ORANGE, D_ORANGE,  DA_ORANGE,  True),
    ('Employee Source',      25, H_PURPLE, D_PURPLE,  DA_PURPLE,  False),
    ('Revenue (USD M)',      14, H_ORANGE, D_ORANGE,  DA_ORANGE,  True),
    ('Revenue Confidence',   14, H_ORANGE, D_ORANGE,  DA_ORANGE,  True),
    ('Revenue Source',       30, H_PURPLE, D_PURPLE,  DA_PURPLE,  False),
    ('Total Contacts',       13, H_BLUE,   D_CONTACT, DA_CONTACT, True),
    ('Contacts & Titles',    40, H_BLUE,   D_CONTACT, DA_CONTACT, False),
    ('Company LinkedIn',     32, H_TEAL,   D_TEAL,    DA_TEAL,    False),
    ('Verification Status',  14, H_GREEN,  D_VERIFY,  DA_VERIFY,  True),
    ('Data Quality Score',   13, H_GREEN,  D_VERIFY,  DA_VERIFY,  True),
    ('Hiring Signals',       30, H_TEAL,   D_TEAL,    DA_TEAL,    False),
    ('Recent News',          40, H_TEAL,   D_TEAL,    DA_TEAL,    False),
    ('News Source',          30, H_PURPLE, D_PURPLE,  DA_PURPLE,  False),
    ('Total Emails',         12, H_BLUE,   D_CONTACT, DA_CONTACT, True),
    ('All Emails',           30, H_BLUE,   D_CONTACT, DA_CONTACT, False),
    ('All Phones',           20, H_BLUE,   D_CONTACT, DA_CONTACT, False),
    ('All LinkedIn URLs',    35, H_BLUE,   D_CONTACT, DA_CONTACT, False),
]

def style_sheet(ws, col_defs):
    """Apply full styling to a sheet."""
    max_row = ws.max_row
    max_col = len(col_defs)
    
    # Track company names for zebra striping by company group
    company_col_idx = None
    for i, (name, *_) in enumerate(col_defs):
        if name == 'Company Name':
            company_col_idx = i + 1
            break
    
    # Build company group alternation map
    alt_map = {}
    if company_col_idx:
        current_company = None
        is_alt = False
        for r in range(2, max_row + 1):
            company = ws.cell(row=r, column=company_col_idx).value
            if company != current_company:
                current_company = company
                is_alt = not is_alt
            alt_map[r] = is_alt
    
    # Style headers
    for c, (name, width, h_fill, d_fill, da_fill, center) in enumerate(col_defs, 1):
        cell = ws.cell(row=1, column=c)
        cell.font = header_font
        cell.fill = h_fill
        cell.alignment = header_align
        cell.border = thin_border
        ws.column_dimensions[get_column_letter(c)].width = width
    
    # Style data rows
    for r in range(2, max_row + 1):
        is_alt = alt_map.get(r, r % 2 == 0)
        # Row height
        ws.row_dimensions[r].height = 65
        
        for c, (name, width, h_fill, d_fill, da_fill, center) in enumerate(col_defs, 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = da_fill if is_alt else d_fill
            cell.font = data_font_bold if name in ('Company Name', 'Contact Name', 'Blueprint Score') else data_font
            cell.alignment = data_align_center if center else data_align
            cell.border = thin_border
    
    # Freeze + filter
    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(max_col)}{max_row}"
    
    print(f"  Styled {max_row - 1} data rows, {max_col} columns")

# ============================================================
# Apply styling
# ============================================================
print("Styling Enriched Leads...")
style_sheet(wb['Enriched Leads'], EL_COLS)

print("Styling Company Summary...")
style_sheet(wb['Company Summary'], CS_COLS)

wb.save(OUTPUT)
print(f"\nSaved to {OUTPUT}")
