#!/usr/bin/env python3
"""
XLSX Builder v2 — Final enriched leads output
"""

import json
import os
from collections import defaultdict
import openpyxl
from openpyxl.styles import (
    PatternFill, Font, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter

WORKSPACE = "/Users/corgi12/.eragon-joshua_augustine/joshua_augustine_workspace"
OUTPUT_PATH = os.path.join(WORKSPACE, "enriched_leads_v2.xlsx")

# ── Colours ──────────────────────────────────────────────────────────────────
C_HEADER_FILL  = "1F4E79"
C_HEADER_FONT  = "FFFFFF"
C_ALT_ROW      = "D6E4F0"
C_WHITE        = "FFFFFF"

C_VERIFIED_FILL   = "C6EFCE"; C_VERIFIED_FONT   = "276221"
C_UNVERIFIED_FILL = "FFEB9C"; C_UNVERIFIED_FONT = "9C5700"
C_DISCREPANCY_FILL= "FFC7CE"; C_DISCREPANCY_FONT= "9C0006"

C_SCORE = {5:"C6EFCE", 4:"E2EFDA", 3:"FFEB9C", 2:"FCE4D6", 1:"FFC7CE"}

TAB_BLUE  = "2E75B6"
TAB_GREEN = "70AD47"
TAB_RED   = "FF0000"

# ── Helpers ───────────────────────────────────────────────────────────────────

def load_json(fname):
    with open(os.path.join(WORKSPACE, fname)) as f:
        return json.load(f)

def norm(name):
    return name.strip().lower() if name else ""

def best(a, b):
    """Return whichever is non-empty; prefer a."""
    if a and str(a).strip(): return a
    return b

def make_fill(hex_color):
    return PatternFill(start_color=hex_color, end_color=hex_color, fill_type="solid")

def make_font(hex_color, bold=False, size=10):
    return Font(color=hex_color, bold=bold, size=size)

def header_style():
    return (
        make_fill(C_HEADER_FILL),
        Font(color=C_HEADER_FONT, bold=True, size=10),
        Alignment(horizontal="center", vertical="center", wrap_text=True),
    )

def row_alignment(wrap=False):
    return Alignment(vertical="top", wrap_text=wrap)

def style_sheet(ws, col_specs, freeze_row=2):
    """Apply header + column widths to a sheet."""
    fill, font, align = header_style()
    for col_idx, (header, width, wrap) in enumerate(col_specs, start=1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.fill = fill
        cell.font = font
        cell.alignment = align
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    ws.row_dimensions[1].height = 30
    ws.freeze_panes = ws.cell(row=freeze_row, column=1)
    ws.auto_filter.ref = ws.dimensions

def apply_data_row(ws, row_idx, values, wrap_cols=None):
    """Write a data row with alternating fill and optional wrap."""
    wrap_cols = wrap_cols or set()
    fill_color = C_ALT_ROW if row_idx % 2 == 0 else C_WHITE
    row_fill = make_fill(fill_color)
    ws.row_dimensions[row_idx].height = 18
    for col_idx, val in enumerate(values, start=1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = val if val is not None else ""
        cell.fill = row_fill
        cell.font = Font(size=10)
        cell.alignment = Alignment(
            vertical="top",
            wrap_text=(col_idx in wrap_cols)
        )

def color_status_cell(cell, status):
    if not status:
        return
    s = str(status).strip()
    if s == "Verified":
        cell.fill = make_fill(C_VERIFIED_FILL)
        cell.font = Font(color=C_VERIFIED_FONT, size=10)
    elif s == "Unverified":
        cell.fill = make_fill(C_UNVERIFIED_FILL)
        cell.font = Font(color=C_UNVERIFIED_FONT, size=10)
    elif "Discrepancy" in s:
        cell.fill = make_fill(C_DISCREPANCY_FILL)
        cell.font = Font(color=C_DISCREPANCY_FONT, size=10)

def color_score_cell(cell, score):
    try:
        s = int(score)
        if s in C_SCORE:
            cell.fill = make_fill(C_SCORE[s])
            cell.font = Font(size=10)
    except (TypeError, ValueError):
        pass


# ─────────────────────────────────────────────────────────────────────────────
# STEP 1 — Build master dataset
# ─────────────────────────────────────────────────────────────────────────────

print("Loading source files...")
keep_data     = load_json("cleanup_keep.json")          # 42 verified
research_data = load_json("research_results.json")      # 16 companies
removed_data  = load_json("cleanup_removed.json")       # 35 removed
research_rem  = load_json("research_removes.json")      # 1 removed
pipeline_data = load_json("pipeline_enriched_results.json")  # 100 pipeline

# Build pipeline lookup by normalised name
pipeline_by_name = {}
for p in pipeline_data:
    pipeline_by_name[norm(p.get("name",""))] = p

# Master dict keyed by norm name (to handle dedup)
master = {}

def add_to_master(record):
    key = norm(record["company_name"])
    if key in master:
        # Merge contacts
        existing = master[key]
        existing_names = {norm(c.get("name","")) for c in existing.get("contacts",[])}
        for c in record.get("contacts", []):
            if norm(c.get("name","")) not in existing_names:
                existing.setdefault("contacts", []).append(c)
                existing_names.add(norm(c.get("name","")))
        # Keep best scalar fields
        for field in ["description","website","verification_notes","enrichment_notes",
                       "recent_news","news_source_url","data_quality_score",
                       "hiring_signals","linkedin_url","sec_ticker"]:
            existing[field] = best(existing.get(field), record.get(field))
    else:
        master[key] = dict(record)

# 1. All 42 keep records
for rec in keep_data:
    add_to_master(rec)

# 2. Research results
removal_log = list(removed_data)  # start with cleanup_removed
for rec in research_data:
    action = rec.get("action","").upper()
    if action in ("UPGRADE","KEEP_IMPROVED"):
        add_to_master(rec)
    else:  # REMOVE
        removal_log.append({
            "company_name": rec.get("company_name",""),
            "reason": rec.get("verification_notes") or rec.get("enrichment_notes","Research: marked for removal"),
            "original_status": rec.get("verification_status",""),
            "state": rec.get("state",""),
            "notes": rec.get("enrichment_notes",""),
        })

# Add research_removes
for rr in research_rem:
    removal_log.append({
        "company_name": rr.get("company_name",""),
        "reason": rr.get("reason",""),
        "original_status": rr.get("original_status",""),
        "state": rr.get("state",""),
        "notes": "",
    })

# 3. Merge pipeline enrichment
merged_count = 0
for key, rec in master.items():
    p = pipeline_by_name.get(key)
    if not p:
        # Try partial match (company name might have suffix differences)
        for pk in pipeline_by_name:
            if pk and (key.startswith(pk[:15]) or pk.startswith(key[:15])):
                p = pipeline_by_name[pk]
                break
    if p:
        merged_count += 1
        # linkedin_url
        if not rec.get("linkedin_url"):
            rec["linkedin_url"] = p.get("linkedin_url")
        # recent_news
        if not rec.get("recent_news"):
            rec["recent_news"] = p.get("recent_news")
        # hiring_signals
        if not rec.get("hiring_signals"):
            rec["hiring_signals"] = p.get("hiring_signals")
        # sec_ticker
        if not rec.get("sec_ticker"):
            rec["sec_ticker"] = p.get("sec_ticker")
        # Merge pipeline contacts if they have names not in master
        existing_contact_names = {norm(c.get("name","")) for c in rec.get("contacts",[])}
        for pc in p.get("contacts", []):
            pname = norm(pc.get("name",""))
            if pname and pname not in existing_contact_names:
                rec.setdefault("contacts",[]).append({
                    "name": pc.get("name",""),
                    "title": pc.get("title",""),
                    "source_url": pc.get("source_url",""),
                    "linkedin_url": pc.get("linkedin_url",""),
                    "is_original": False,
                })
                existing_contact_names.add(pname)

master_list = list(master.values())
print(f"  → {len(master_list)} companies in master (merged {merged_count} pipeline records)")
print(f"  → {len(removal_log)} companies in removal log")

# ─────────────────────────────────────────────────────────────────────────────
# STEP 2 — Build XLSX
# ─────────────────────────────────────────────────────────────────────────────

wb = openpyxl.Workbook()
wb.remove(wb.active)  # remove default sheet

# ── SHEET 1: Enriched Leads ───────────────────────────────────────────────────
ws1 = wb.create_sheet("Enriched Leads")
ws1.sheet_properties.tabColor = TAB_BLUE

LEADS_COLS = [
    # (header, width, wrap)
    ("Company Name",          35, False),
    ("State",                  8, False),
    ("Industry",              22, False),
    ("Founded",               10, False),
    ("Employees",             12, False),
    ("Revenue (USD 000s)",    16, False),
    ("Company Description",   52, True),
    ("Contact Name",          28, False),
    ("Contact Title",         28, False),
    ("Is Original Contact",   18, False),
    ("Contact Source URL",    42, False),
    ("LinkedIn URL",          42, False),
    ("Website",               42, False),
    ("Website Verified",      16, False),
    ("Verification Status",   20, False),
    ("Verification Notes",    48, True),
    ("Data Quality Score",    10, False),
    ("Hiring Signals",        30, False),
    ("Recent News Summary",   48, True),
    ("News Source URL",       42, False),
    ("Enrichment Notes",      48, True),
]

style_sheet(ws1, LEADS_COLS)

# Wrap cols (1-indexed)
LEADS_WRAP = {7, 16, 19, 21}  # Description, Verification Notes, Recent News, Enrichment Notes

# Status col = 15, Score col = 17
row_idx = 2
total_contacts = 0
for rec in master_list:
    contacts = rec.get("contacts", [])
    # Filter out contacts with no name
    contacts = [c for c in contacts if c.get("name","").strip()]
    
    if not contacts:
        # Still write one row with empty contact fields
        contacts = [{"name":"","title":"","source_url":"","linkedin_url":"","is_original":False}]

    company_linkedin = rec.get("linkedin_url","")
    
    for ct in contacts:
        ct_name = ct.get("name","") or ""
        if not ct_name.strip():
            continue  # skip nameless contacts
        
        total_contacts += 1
        ct_linkedin = ct.get("linkedin_url","") or company_linkedin
        is_orig = ct.get("is_original", False)
        is_orig_str = "Yes" if is_orig else "No"

        values = [
            rec.get("company_name",""),
            rec.get("state",""),
            rec.get("industry",""),
            rec.get("founded",""),
            rec.get("employees",""),
            rec.get("revenue_usd_000s",""),
            rec.get("description",""),
            ct_name,
            ct.get("title","") or "",
            is_orig_str,
            ct.get("source_url","") or "",
            ct_linkedin or "",
            rec.get("website","") or "",
            "Yes" if rec.get("website_verified") else "No",
            rec.get("verification_status","") or "",
            rec.get("verification_notes","") or "",
            rec.get("data_quality_score","") or "",
            rec.get("hiring_signals","") or "",
            rec.get("recent_news","") or "",
            rec.get("news_source_url","") or "",
            rec.get("enrichment_notes","") or "",
        ]

        apply_data_row(ws1, row_idx, values, wrap_cols=LEADS_WRAP)

        # Color status cell (col 15)
        color_status_cell(ws1.cell(row=row_idx, column=15), rec.get("verification_status",""))
        # Color score cell (col 17)
        color_score_cell(ws1.cell(row=row_idx, column=17), rec.get("data_quality_score"))

        row_idx += 1

# Update auto-filter range to actual data
ws1.auto_filter.ref = f"A1:{get_column_letter(len(LEADS_COLS))}{row_idx-1}"
print(f"  Sheet 1 'Enriched Leads': {row_idx-2} contact rows written")

# ── SHEET 2: Company Summary ──────────────────────────────────────────────────
ws2 = wb.create_sheet("Company Summary")
ws2.sheet_properties.tabColor = TAB_GREEN

SUMMARY_COLS = [
    ("Company Name",          35, False),
    ("State",                  8, False),
    ("Industry",              22, False),
    ("Founded",               10, False),
    ("Employees",             12, False),
    ("Revenue (USD 000s)",    16, False),
    ("Company Description",   52, True),
    ("Description Source URL",42, False),
    ("Total Contacts Found",  18, False),
    ("Contact Names & Titles",52, True),
    ("Website",               42, False),
    ("Website Verified",      16, False),
    ("Company LinkedIn URL",  42, False),
    ("Verification Status",   20, False),
    ("Verification Notes",    48, True),
    ("Data Quality Score",    10, False),
    ("Hiring Signals",        30, False),
    ("Recent News",           48, True),
    ("News Source URL",       42, False),
    ("Enrichment Notes",      48, True),
]

style_sheet(ws2, SUMMARY_COLS)
SUMMARY_WRAP = {7, 10, 15, 18, 20}  # Description, Contact Names, Verification Notes, News, Enrichment

row_idx = 2
for rec in master_list:
    contacts = [c for c in rec.get("contacts",[]) if c.get("name","").strip()]
    contact_str = "; ".join(
        f"{c['name']} — {c.get('title','')}" for c in contacts
    ) if contacts else ""

    values = [
        rec.get("company_name",""),
        rec.get("state",""),
        rec.get("industry",""),
        rec.get("founded",""),
        rec.get("employees",""),
        rec.get("revenue_usd_000s",""),
        rec.get("description","") or "",
        rec.get("description_source","") or "",
        len(contacts),
        contact_str,
        rec.get("website","") or "",
        "Yes" if rec.get("website_verified") else "No",
        rec.get("linkedin_url","") or "",
        rec.get("verification_status","") or "",
        rec.get("verification_notes","") or "",
        rec.get("data_quality_score","") or "",
        rec.get("hiring_signals","") or "",
        rec.get("recent_news","") or "",
        rec.get("news_source_url","") or "",
        rec.get("enrichment_notes","") or "",
    ]

    apply_data_row(ws2, row_idx, values, wrap_cols=SUMMARY_WRAP)
    color_status_cell(ws2.cell(row=row_idx, column=14), rec.get("verification_status",""))
    color_score_cell(ws2.cell(row=row_idx, column=16), rec.get("data_quality_score"))
    row_idx += 1

ws2.auto_filter.ref = f"A1:{get_column_letter(len(SUMMARY_COLS))}{row_idx-1}"
print(f"  Sheet 2 'Company Summary': {row_idx-2} company rows written")

# ── SHEET 3: Removed Leads ────────────────────────────────────────────────────
ws3 = wb.create_sheet("Removed Leads")
ws3.sheet_properties.tabColor = TAB_RED

REMOVED_COLS = [
    ("Company Name",          35, False),
    ("State",                  8, False),
    ("Removal Reason",        52, True),
    ("Original Verification Status", 26, False),
    ("Notes",                 48, True),
]

style_sheet(ws3, REMOVED_COLS)
REMOVED_WRAP = {3, 5}

row_idx = 2
for rem in removal_log:
    values = [
        rem.get("company_name",""),
        rem.get("state","") or "",
        rem.get("reason","") or "",
        rem.get("original_status","") or "",
        rem.get("notes","") or "",
    ]
    apply_data_row(ws3, row_idx, values, wrap_cols=REMOVED_WRAP)
    row_idx += 1

ws3.auto_filter.ref = f"A1:{get_column_letter(len(REMOVED_COLS))}{row_idx-1}"
print(f"  Sheet 3 'Removed Leads': {row_idx-2} removed rows written")

# ── Save ──────────────────────────────────────────────────────────────────────
wb.save(OUTPUT_PATH)
file_size_kb = os.path.getsize(OUTPUT_PATH) / 1024

# ─────────────────────────────────────────────────────────────────────────────
# STEP 4 — Summary
# ─────────────────────────────────────────────────────────────────────────────

print("\n" + "="*60)
print("ENRICHED LEADS v2 — BUILD SUMMARY")
print("="*60)
print(f"Total companies in v2  : {len(master_list)}")
print(f"Total contacts         : {total_contacts}")

# Verification breakdown
v_counts = defaultdict(int)
for rec in master_list:
    v_counts[rec.get("verification_status","Unknown")] += 1
print("\nVerification breakdown:")
for status, cnt in sorted(v_counts.items(), key=lambda x:-x[1]):
    print(f"  {status:25s}: {cnt}")

# Quality score distribution
q_counts = defaultdict(int)
for rec in master_list:
    q = rec.get("data_quality_score")
    q_counts[str(q) if q is not None else "N/A"] += 1
print("\nData Quality Score distribution:")
for score in sorted(q_counts.keys(), reverse=True):
    print(f"  Score {score:5s}: {q_counts[score]}")

# Removed
print(f"\nTotal removed          : {len(removal_log)}")
reason_counts = defaultdict(int)
for r in removal_log:
    reason = r.get("reason","") or ""
    short = reason[:60].split(".")[0].strip() if reason else "Unknown"
    reason_counts[short] += 1
print("Top removal reasons:")
for reason, cnt in sorted(reason_counts.items(), key=lambda x:-x[1])[:8]:
    print(f"  ({cnt}x) {reason}")

print(f"\nFile size              : {file_size_kb:.1f} KB")
print(f"Output path            : {OUTPUT_PATH}")
print("\n✅ enriched_leads_v2.xlsx written successfully.")
