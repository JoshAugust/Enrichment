#!/usr/bin/env python3
"""Sweep SEC EDGAR for public company officers and filings data."""

import json
import re
import time
import urllib.request
import urllib.parse
import ssl
import openpyxl

# SEC EDGAR requires a User-Agent with contact info
HEADERS = {
    'User-Agent': 'DealScope Research josh@corgi.insure',
    'Accept': 'application/json'
}

def sec_request(url, retries=3):
    """Make a request to SEC EDGAR with rate limiting."""
    ctx = ssl.create_default_context()
    req = urllib.request.Request(url, headers=HEADERS)
    for attempt in range(retries):
        try:
            with urllib.request.urlopen(req, timeout=15, context=ctx) as resp:
                return json.loads(resp.read().decode('utf-8'))
        except Exception as e:
            if attempt < retries - 1:
                time.sleep(1)
            else:
                print(f"  Failed: {url} - {e}")
                return None
    return None

def search_company(company_name):
    """Search EDGAR full-text search for a company."""
    # Clean name for search
    clean = re.sub(r'\s*\(.*?\)\s*', ' ', company_name).strip()
    clean = re.sub(r'\s+(INC\.?|LLC|CORP\.?|LTD\.?|CO\.?)$', '', clean, flags=re.IGNORECASE).strip()
    
    url = f"https://efts.sec.gov/LATEST/search-index?q=%22{urllib.parse.quote(clean)}%22&dateRange=custom&startdt=2024-01-01&enddt=2026-03-30&forms=10-K,DEF%2014A"
    # Use company search endpoint instead
    url = f"https://efts.sec.gov/LATEST/search-index?q={urllib.parse.quote(clean)}"
    
    # Better: use company tickers endpoint
    return None

def lookup_by_ticker(ticker):
    """Look up CIK by ticker symbol."""
    url = "https://www.sec.gov/files/company_tickers.json"
    data = sec_request(url)
    if not data:
        return None
    
    for entry in data.values():
        if entry.get('ticker', '').upper() == ticker.upper():
            cik = str(entry['cik_str']).zfill(10)
            return cik, entry.get('title', '')
    return None

def lookup_by_name(name):
    """Search company by name via EDGAR company search."""
    clean = re.sub(r'\s*\(.*?\)\s*', ' ', name).strip()
    clean = re.sub(r'\s+(INC\.?|LLC|CORP\.?|LTD\.?|CO\.?)$', '', clean, flags=re.IGNORECASE).strip()
    
    url = f"https://efts.sec.gov/LATEST/search-index?q=%22{urllib.parse.quote(clean)}%22&forms=10-K"
    # Use the company search API
    url = f"https://efts.sec.gov/LATEST/search-index?company={urllib.parse.quote(clean)}&forms=10-K"
    return None

def get_company_info(cik):
    """Get company submissions/filings info from EDGAR."""
    url = f"https://data.sec.gov/submissions/CIK{cik}.json"
    return sec_request(url)

def get_officers_from_submission(cik):
    """Extract officer info from company filings."""
    info = get_company_info(cik)
    if not info:
        return None
    
    result = {
        'name': info.get('name'),
        'cik': cik,
        'sic': info.get('sic'),
        'sicDescription': info.get('sicDescription'),
        'stateOfIncorporation': info.get('stateOfIncorporation'),
        'ein': info.get('ein'),
        'phone': None,
        'address': None,
        'officers': [],
        'recent_filings': []
    }
    
    # Get business address and phone
    addresses = info.get('addresses', {})
    biz = addresses.get('business', {})
    if biz:
        result['phone'] = biz.get('phone')
        addr_parts = [biz.get('street1', ''), biz.get('street2', ''), 
                      biz.get('city', ''), biz.get('stateOrCountry', ''), 
                      biz.get('zipCode', '')]
        result['address'] = ', '.join(p for p in addr_parts if p)
    
    # Get officers from insider trading (Forms 3, 4, 5)
    # These are in the "officers" field if available
    # More reliable: check recent filings
    recent = info.get('filings', {}).get('recent', {})
    if recent:
        forms = recent.get('form', [])
        dates = recent.get('filingDate', [])
        accessions = recent.get('accessionNumber', [])
        primary_docs = recent.get('primaryDocument', [])
        
        # Get last few filings
        for i in range(min(5, len(forms))):
            result['recent_filings'].append({
                'form': forms[i],
                'date': dates[i],
                'accession': accessions[i]
            })
    
    # Check for insider ownership filings to get officer names
    # The ownership endpoint has the actual officers
    own_url = f"https://data.sec.gov/submissions/CIK{cik}.json"
    # Officers are often in the main JSON under 'insiders' or we need DEF 14A
    
    return result

def get_insider_owners(cik):
    """Get officers from insider ownership filings (Forms 3/4)."""
    url = f"https://data.sec.gov/submissions/CIK{cik}.json"
    data = sec_request(url)
    if not data:
        return []
    
    # Check for officers in recent Form 4 filings
    recent = data.get('filings', {}).get('recent', {})
    if not recent:
        return []
    
    forms = recent.get('form', [])
    # Look for ownership forms
    officers = set()
    
    # The actual officer data needs to come from the XML of Form 4 filings
    # But we can get company contact info from the main JSON
    return data

# ============================================================
# MAIN: Load tickers, sweep EDGAR
# ============================================================

# First, download the full ticker-to-CIK mapping
print("Downloading SEC ticker mapping...")
ticker_map = sec_request("https://www.sec.gov/files/company_tickers.json")
if not ticker_map:
    print("Failed to get ticker mapping!")
    exit(1)

# Build reverse lookup
ticker_to_cik = {}
name_to_cik = {}
for entry in ticker_map.values():
    ticker_to_cik[entry['ticker'].upper()] = (str(entry['cik_str']).zfill(10), entry.get('title', ''))
    # Also index by name for fuzzy matching
    name_to_cik[entry.get('title', '').upper()] = (str(entry['cik_str']).zfill(10), entry['ticker'])

print(f"Loaded {len(ticker_to_cik)} tickers")

# Load our public companies from both files
public_companies = []

# DealScope software file
wb1 = openpyxl.load_workbook('DealScope_Enriched_v2.xlsx')
ws1 = wb1['Company Summary']
headers1 = [ws1.cell(row=1, column=c).value for c in range(1, ws1.max_column+1)]

for r in range(2, ws1.max_row+1):
    stage = ws1.cell(row=r, column=headers1.index('Startup Stage')+1).value
    if stage == 'Public':
        name = ws1.cell(row=r, column=headers1.index('Company Name')+1).value
        desc = str(ws1.cell(row=r, column=headers1.index('Company Description')+1).value or '')
        
        tickers = re.findall(r'\((?:Nasdaq|NYSE|NASDAQ):\s*([A-Z]+)\)', desc)
        ticker = tickers[0] if tickers else None
        
        if name not in [c['name'] for c in public_companies]:
            public_companies.append({
                'name': name,
                'ticker': ticker,
                'source': 'software',
                'row': r
            })

# Lender file  
wb2 = openpyxl.load_workbook('DealScope_Lenders_v1.xlsx')
ws2 = wb2['Company Summary']
headers2 = [ws2.cell(row=1, column=c).value for c in range(1, ws2.max_column+1)]

for r in range(2, ws2.max_row+1):
    stage = ws2.cell(row=r, column=headers2.index('Lender Stage')+1).value
    if stage and 'public' in str(stage).lower():
        name = ws2.cell(row=r, column=headers2.index('Company')+1).value
        desc = str(ws2.cell(row=r, column=headers2.index('Company Description')+1).value or '')
        
        tickers = re.findall(r'\((?:Nasdaq|NYSE|NASDAQ):\s*([A-Z]+)\)', desc)
        ticker = tickers[0] if tickers else None
        
        if name not in [c['name'] for c in public_companies]:
            public_companies.append({
                'name': name,
                'ticker': ticker,
                'source': 'lender',
                'row': r
            })

print(f"\nFound {len(public_companies)} public companies to look up")

# Try to match companies without tickers by name
for co in public_companies:
    if not co['ticker']:
        # Try fuzzy name matching
        clean_name = re.sub(r'\s*\(.*?\)\s*', ' ', co['name']).strip().upper()
        clean_name = re.sub(r'\s+(INC\.?|LLC|CORP\.?|CORPORATION|LTD\.?|CO\.?)$', '', clean_name).strip()
        
        # Exact match
        if clean_name in name_to_cik:
            cik, ticker = name_to_cik[clean_name]
            co['ticker'] = ticker
            co['cik'] = cik
            continue
        
        # Partial match
        for sec_name, (cik, ticker) in name_to_cik.items():
            if clean_name in sec_name or sec_name in clean_name:
                co['ticker'] = ticker
                co['cik'] = cik
                break

# Look up CIKs for those with tickers
for co in public_companies:
    if co.get('ticker') and not co.get('cik'):
        if co['ticker'].upper() in ticker_to_cik:
            co['cik'], _ = ticker_to_cik[co['ticker'].upper()]

with_cik = sum(1 for c in public_companies if c.get('cik'))
print(f"Matched {with_cik}/{len(public_companies)} to SEC CIK numbers")

# Now fetch EDGAR data for each matched company
results = []
for i, co in enumerate(public_companies):
    if not co.get('cik'):
        continue
    
    print(f"[{i+1}/{with_cik}] Fetching {co['name']} (CIK: {co['cik']}, ticker: {co.get('ticker')})...")
    info = get_company_info(co['cik'])
    
    if info:
        addresses = info.get('addresses', {})
        biz = addresses.get('business', {})
        mail = addresses.get('mailing', {})
        
        phone = biz.get('phone') or mail.get('phone')
        
        addr_parts = [biz.get('street1', ''), biz.get('street2', ''),
                      biz.get('city', ''), biz.get('stateOrCountry', ''),
                      biz.get('zipCode', '')]
        address = ', '.join(p for p in addr_parts if p)
        
        # Recent filings
        recent = info.get('filings', {}).get('recent', {})
        recent_forms = []
        if recent:
            forms = recent.get('form', [])
            dates = recent.get('filingDate', [])
            for j in range(min(3, len(forms))):
                recent_forms.append(f"{forms[j]} ({dates[j]})")
        
        result = {
            'company': co['name'],
            'source_file': co['source'],
            'ticker': co.get('ticker'),
            'cik': co['cik'],
            'sec_name': info.get('name'),
            'sic': info.get('sic'),
            'sic_desc': info.get('sicDescription'),
            'state_of_inc': info.get('stateOfIncorporation'),
            'ein': info.get('ein'),
            'phone': phone,
            'address': address,
            'website': info.get('website', ''),
            'recent_filings': '; '.join(recent_forms),
            'category': info.get('category', ''),
            'fiscal_year_end': info.get('fiscalYearEnd', ''),
        }
        results.append(result)
        print(f"  ✓ Phone: {phone} | Address: {address[:50]}... | SIC: {info.get('sicDescription', 'N/A')}")
    
    # SEC rate limit: 10 requests/second
    time.sleep(0.15)

print(f"\n{'='*60}")
print(f"EDGAR SWEEP COMPLETE")
print(f"{'='*60}")
print(f"Companies queried: {with_cik}")
print(f"Results with phone: {sum(1 for r in results if r.get('phone'))}")
print(f"Results with address: {sum(1 for r in results if r.get('address'))}")

# Save results
with open('sec_edgar_results.json', 'w') as f:
    json.dump(results, f, indent=2)

print(f"\nSaved to sec_edgar_results.json")

# Print summary
print(f"\n--- RESULTS ---")
for r in results:
    phone_status = f"📞 {r['phone']}" if r['phone'] else "❌ No phone"
    print(f"  {r['ticker'] or '???':6s} | {r['company'][:40]:40s} | {phone_status}")
