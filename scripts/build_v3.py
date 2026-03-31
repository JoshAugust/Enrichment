#!/usr/bin/env python3
"""Build enriched_leads_v3.xlsx — v2 + phone numbers from db_phones.json"""

import json
import re
import os
import openpyxl
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter
from copy import copy

# ── Constants ────────────────────────────────────────────────────────────────
WS_DIR = "/Users/corgi12/.eragon-joshua_augustine/joshua_augustine_workspace"
SRC = os.path.join(WS_DIR, "enriched_leads_v2.xlsx")
DB  = os.path.join(WS_DIR, "db_phones.json")
OUT = os.path.join(WS_DIR, "enriched_leads_v3.xlsx")

HDR_FILL   = PatternFill("solid", fgColor="1F4E79")
HDR_FONT   = Font(bold=True, color="FFFFFF", size=10)
ALT_FILL_A = PatternFill("solid", fgColor="FFFFFF")
ALT_FILL_B = PatternFill("solid", fgColor="D6E4F0")
GREY_FONT  = Font(color="999999", size=10)
NORM_FONT  = Font(size=10)

STATUS_COLORS = {
    "Verified":   "C6EFCE",
    "Partial":    "FFEB9C",
    "Unverified": "FFC7CE",
    "Removed":    "FFC7CE",
}
SCORE_COLORS = {
    4: "C6EFCE",
    3: "FFEB9C",
    2: "FFC7CE",
    1: "FFC7CE",
}

WRAP_COLS_EL  = {"Company Description", "Verification Notes", "Recent News Summary", "Enrichment Notes"}
WRAP_COLS_CS  = {"Company Description", "Verification Notes", "Recent News", "Enrichment Notes", "Contact Names & Titles"}


def normalize(name: str) -> str:
    """Lowercase, strip Inc/LLC/Corp/Ltd and punctuation for fuzzy match."""
    if not name:
        return ""
    n = name.lower()
    n = re.sub(r'\b(inc|llc|corp|ltd|co|company|companies|group|holdings|services|solutions)\b', '', n)
    n = re.sub(r'[^a-z0-9 ]', '', n)
    n = re.sub(r'\s+', ' ', n).strip()
    return n


def build_phone_lookups(db_path: str):
    with open(db_path) as f:
        data = json.load(f)

    company_phones  = {}   # norm_name → {phone, phone_source}
    contact_phones  = {}   # (norm_company, norm_contact) → direct_phone
    raw_name_map    = {}   # norm_name → original name (for debugging)

    for entry in data:
        cname = entry.get("company_name", "") or ""
        norm  = normalize(cname)
        if entry.get("phone"):
            company_phones[norm] = {
                "phone":        entry["phone"],
                "phone_source": entry.get("phone_source", ""),
            }
        raw_name_map[norm] = cname

        for contact in entry.get("contacts", []):
            cname2 = contact.get("name", "") or ""
            dp     = contact.get("direct_phone")
            if dp:
                contact_phones[(norm, normalize(cname2))] = dp

    return company_phones, contact_phones


def lookup_company_phone(company_name, company_phones):
    norm = normalize(company_name or "")
    return company_phones.get(norm)


def lookup_contact_phone(company_name, contact_name, contact_phones):
    nc = normalize(company_name or "")
    np = normalize(contact_name or "")
    return contact_phones.get((nc, np))


def copy_cell_style(src_cell, dst_cell):
    """Copy all style properties from src to dst."""
    if src_cell.has_style:
        dst_cell.font        = copy(src_cell.font)
        dst_cell.fill        = copy(src_cell.fill)
        dst_cell.border      = copy(src_cell.border)
        dst_cell.alignment   = copy(src_cell.alignment)
        dst_cell.number_format = src_cell.number_format


def make_header_cell(ws, row, col, value):
    cell = ws.cell(row=row, column=col, value=value)
    cell.font      = HDR_FONT
    cell.fill      = HDR_FILL
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=False)
    return cell


def row_fill(row_idx):
    """Return fill for data rows (1-indexed data rows, row 2 = first data row)."""
    return ALT_FILL_A if (row_idx % 2 == 0) else ALT_FILL_B


def apply_data_cell(cell, value, fill, font=None, wrap=False, grey=False):
    cell.value = value
    cell.fill  = fill
    cell.font  = GREY_FONT if grey else (font or NORM_FONT)
    cell.alignment = Alignment(
        vertical="top",
        wrap_text=wrap,
        horizontal="left",
    )


# ── Main ─────────────────────────────────────────────────────────────────────
def main():
    print("Loading source files...")
    wb_src = openpyxl.load_workbook(SRC)
    company_phones, contact_phones = build_phone_lookups(DB)
    print(f"  Company phone entries in DB: {len(company_phones)}")
    print(f"  Contact direct phone entries in DB: {len(contact_phones)}")

    wb_out = openpyxl.Workbook()
    wb_out.remove(wb_out.active)  # remove default blank sheet

    # ── Sheet 1: Enriched Leads ───────────────────────────────────────────────
    ws_src = wb_src["Enriched Leads"]
    ws_out = wb_out.create_sheet("Enriched Leads")

    # Read all rows
    src_rows = list(ws_src.iter_rows(values_only=False))
    src_hdr  = [c.value for c in src_rows[0]]

    # We insert 3 new cols after "Contact Title" (0-indexed col 8, 1-indexed col 9)
    # New layout: cols 1..9 same, then col 10="Company Phone", 11="Company Phone Source",
    #             12="Contact Direct Phone", then original cols 10..21 → 13..24
    INSERT_AFTER = 9   # 1-indexed; insert after this column

    # New headers
    new_hdrs = src_hdr[:INSERT_AFTER] + [
        "Company Phone", "Company Phone Source", "Contact Direct Phone"
    ] + src_hdr[INSERT_AFTER:]

    # Copy column widths from source (shifted), set new col widths
    src_col_dims = {
        ws_src.column_dimensions[get_column_letter(i+1)].width
        for i in range(len(src_hdr))
    }
    # Copy individual widths
    for i, w in enumerate(
        [ws_src.column_dimensions[get_column_letter(i+1)].width for i in range(len(src_hdr))]
    ):
        col_idx = i + 1
        if col_idx <= INSERT_AFTER:
            out_col = col_idx
        else:
            out_col = col_idx + 3
        ws_out.column_dimensions[get_column_letter(out_col)].width = w or 12

    # New phone columns width
    for c in [INSERT_AFTER+1, INSERT_AFTER+2, INSERT_AFTER+3]:
        ws_out.column_dimensions[get_column_letter(c)].width = 20

    # Identify wrap columns (by index in new headers, 1-indexed)
    wrap_col_indices = set()
    for i, h in enumerate(new_hdrs):
        if h in WRAP_COLS_EL:
            wrap_col_indices.add(i + 1)

    # Find verification status col and data quality score col (1-indexed in new layout)
    vs_col  = new_hdrs.index("Verification Status") + 1  if "Verification Status"  in new_hdrs else None
    dq_col  = new_hdrs.index("Data Quality Score")  + 1  if "Data Quality Score"   in new_hdrs else None
    co_col  = new_hdrs.index("Company Name")         + 1  if "Company Name"         in new_hdrs else 1
    cn_col  = new_hdrs.index("Contact Name")         + 1  if "Contact Name"         in new_hdrs else 8
    ph_col  = new_hdrs.index("Company Phone")        + 1
    phs_col = new_hdrs.index("Company Phone Source") + 1
    dp_col  = new_hdrs.index("Contact Direct Phone") + 1

    # Write header row
    ws_out.row_dimensions[1].height = 30
    for col_idx, hdr in enumerate(new_hdrs, 1):
        make_header_cell(ws_out, 1, col_idx, hdr)

    # Freeze below header
    ws_out.freeze_panes = "A2"
    ws_out.auto_filter.ref = f"A1:{get_column_letter(len(new_hdrs))}1"

    # Stats
    companies_with_phone   = set()
    companies_without      = set()
    contacts_with_dp       = 0

    # Write data rows
    for row_idx, src_row in enumerate(src_rows[1:], 2):
        ws_out.row_dimensions[row_idx].height = 18
        fill = row_fill(row_idx)

        src_vals = [c.value for c in src_row]
        company_name = src_vals[0]
        contact_name = src_vals[7]  # "Contact Name" is index 7 in src (0-indexed)

        # Phone lookups
        cp  = lookup_company_phone(company_name, company_phones)
        cdp = lookup_contact_phone(company_name, contact_name, contact_phones)

        phone_val    = cp["phone"]        if cp else None
        phone_src    = cp["phone_source"] if cp else None
        contact_dp   = cdp if cdp else None

        if company_name:
            if phone_val:
                companies_with_phone.add(normalize(company_name))
            else:
                companies_without.add(normalize(company_name))
        if contact_dp:
            contacts_with_dp += 1

        # Build new row values (Python list, 0-indexed)
        new_vals = src_vals[:INSERT_AFTER] + [phone_val, phone_src, contact_dp] + src_vals[INSERT_AFTER:]

        for col_idx, val in enumerate(new_vals, 1):
            cell = ws_out.cell(row=row_idx, column=col_idx)
            wrap = col_idx in wrap_col_indices

            # Special coloring for verification status
            if col_idx == vs_col and val in STATUS_COLORS:
                cell.value = val
                cell.fill  = PatternFill("solid", fgColor=STATUS_COLORS[val])
                cell.font  = NORM_FONT
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx == dq_col and val in SCORE_COLORS:
                cell.value = val
                cell.fill  = PatternFill("solid", fgColor=SCORE_COLORS[val])
                cell.font  = NORM_FONT
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx in (ph_col, phs_col, dp_col):
                # Phone columns
                grey = not val
                apply_data_cell(cell, val if val else "—", fill, wrap=False, grey=grey)
            else:
                apply_data_cell(cell, val, fill, wrap=wrap)

    print(f"\n[Sheet 1] Enriched Leads written.")
    print(f"  Companies with phone    : {len(companies_with_phone)}")
    print(f"  Companies without phone : {len(companies_without)}")
    print(f"  Contacts with direct ph : {contacts_with_dp}")

    # ── Sheet 2: Company Summary ──────────────────────────────────────────────
    ws2_src = wb_src["Company Summary"]
    ws2_out = wb_out.create_sheet("Company Summary")

    src2_rows = list(ws2_src.iter_rows(values_only=False))
    src2_hdr  = [c.value for c in src2_rows[0]]

    # Insert after "Website Verified" (0-indexed 11, 1-indexed 12)
    INSERT2_AFTER = src2_hdr.index("Website Verified") + 1  # 1-indexed

    new2_hdrs = src2_hdr[:INSERT2_AFTER] + [
        "Company Phone", "Phone Source"
    ] + src2_hdr[INSERT2_AFTER:]

    # Copy column widths
    for i in range(len(src2_hdr)):
        col_idx = i + 1
        w = ws2_src.column_dimensions[get_column_letter(col_idx)].width or 12
        if col_idx <= INSERT2_AFTER:
            out_col = col_idx
        else:
            out_col = col_idx + 2
        ws2_out.column_dimensions[get_column_letter(out_col)].width = w

    for c in [INSERT2_AFTER+1, INSERT2_AFTER+2]:
        ws2_out.column_dimensions[get_column_letter(c)].width = 20

    # Identify wrap columns
    wrap2_col_indices = set()
    for i, h in enumerate(new2_hdrs):
        if h in WRAP_COLS_CS:
            wrap2_col_indices.add(i + 1)

    vs2_col  = new2_hdrs.index("Verification Status") + 1  if "Verification Status"  in new2_hdrs else None
    dq2_col  = new2_hdrs.index("Data Quality Score")  + 1  if "Data Quality Score"   in new2_hdrs else None
    ph2_col  = new2_hdrs.index("Company Phone")       + 1
    phs2_col = new2_hdrs.index("Phone Source")        + 1

    # Write header row
    ws2_out.row_dimensions[1].height = 30
    for col_idx, hdr in enumerate(new2_hdrs, 1):
        make_header_cell(ws2_out, 1, col_idx, hdr)

    ws2_out.freeze_panes = "A2"
    ws2_out.auto_filter.ref = f"A1:{get_column_letter(len(new2_hdrs))}1"

    # Data rows
    for row_idx, src_row in enumerate(src2_rows[1:], 2):
        ws2_out.row_dimensions[row_idx].height = 18
        fill = row_fill(row_idx)

        src_vals = [c.value for c in src_row]
        company_name = src_vals[0]

        cp        = lookup_company_phone(company_name, company_phones)
        phone_val = cp["phone"]        if cp else None
        phone_src = cp["phone_source"] if cp else None

        new_vals = src_vals[:INSERT2_AFTER] + [phone_val, phone_src] + src_vals[INSERT2_AFTER:]

        for col_idx, val in enumerate(new_vals, 1):
            cell = ws2_out.cell(row=row_idx, column=col_idx)
            wrap = col_idx in wrap2_col_indices

            if col_idx == vs2_col and val in STATUS_COLORS:
                cell.value = val
                cell.fill  = PatternFill("solid", fgColor=STATUS_COLORS[val])
                cell.font  = NORM_FONT
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx == dq2_col and val in SCORE_COLORS:
                cell.value = val
                cell.fill  = PatternFill("solid", fgColor=SCORE_COLORS[val])
                cell.font  = NORM_FONT
                cell.alignment = Alignment(horizontal="center", vertical="top")
            elif col_idx in (ph2_col, phs2_col):
                grey = not val
                apply_data_cell(cell, val if val else "—", fill, wrap=False, grey=grey)
            else:
                apply_data_cell(cell, val, fill, wrap=wrap)

    print(f"\n[Sheet 2] Company Summary written.")

    # ── Sheet 3: Removed Leads ────────────────────────────────────────────────
    ws3_src = wb_src["Removed Leads"]
    ws3_out = wb_out.create_sheet("Removed Leads")

    src3_rows = list(ws3_src.iter_rows(values_only=False))
    src3_hdr  = [c.value for c in src3_rows[0]]

    # Copy column widths
    for i in range(len(src3_hdr)):
        col_idx = i + 1
        w = ws3_src.column_dimensions[get_column_letter(col_idx)].width or 12
        ws3_out.column_dimensions[get_column_letter(col_idx)].width = w

    wrap3_col_indices = set()
    for i, h in enumerate(src3_hdr):
        if h in {"Removal Reason", "Notes", "Verification Notes"}:
            wrap3_col_indices.add(i + 1)

    # Write header
    ws3_out.row_dimensions[1].height = 30
    for col_idx, hdr in enumerate(src3_hdr, 1):
        make_header_cell(ws3_out, 1, col_idx, hdr)

    ws3_out.freeze_panes = "A2"
    ws3_out.auto_filter.ref = f"A1:{get_column_letter(len(src3_hdr))}1"

    # Data rows
    for row_idx, src_row in enumerate(src3_rows[1:], 2):
        ws3_out.row_dimensions[row_idx].height = 18
        fill = row_fill(row_idx)
        src_vals = [c.value for c in src_row]
        for col_idx, val in enumerate(src_vals, 1):
            cell = ws3_out.cell(row=row_idx, column=col_idx)
            wrap = col_idx in wrap3_col_indices
            apply_data_cell(cell, val, fill, wrap=wrap)

    print(f"\n[Sheet 3] Removed Leads written (copied unchanged).")

    # ── Save ──────────────────────────────────────────────────────────────────
    wb_out.save(OUT)
    size = os.path.getsize(OUT)
    print(f"\n✅ Written: {OUT}")
    print(f"   File size: {size:,} bytes ({size/1024:.1f} KB)")
    print(f"\n📊 Summary:")
    print(f"   Companies WITH phone number  : {len(companies_with_phone)}")
    print(f"   Companies WITHOUT phone      : {len(companies_without - companies_with_phone)}")
    print(f"   Contacts with direct phone   : {contacts_with_dp}")


if __name__ == "__main__":
    main()
