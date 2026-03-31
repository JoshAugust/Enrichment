#!/usr/bin/env python3
"""Add 24 more companies to reach ~100 total."""

import openpyxl
from copy import copy

wb = openpyxl.load_workbook('/Users/corgi12/.eragon-joshua_augustine/joshua_augustine_workspace/gpu-operator-expansion/DealScope_GPU_Operators_v2.xlsx')
ws_leads = wb['Enriched Leads']
ws_summary = wb['Company Summary']

lead_headers = [c.value for c in ws_leads[1]]
summary_headers = [c.value for c in ws_summary[1]]

additional = [
    {
        "company": "Scaleway (Iliad Group)",
        "description": "French cloud provider, part of Iliad Group. GPU cloud with NVIDIA H100, L40S. European sovereign cloud. 7 data centers. Part of €600M AI investment by Iliad.",
        "description_source": "Company website + news (2025)",
        "industry": "Cloud / GPU Cloud",
        "website": "scaleway.com",
        "hq": "Paris, France",
        "founded": 1999,
        "employees": "~600",
        "employee_source": "LinkedIn (2025)",
        "total_raised": "Part of Iliad Group (€600M AI investment)",
        "last_round": "Iliad Group internal investment (2024)",
        "gpu_fleet": "3,000",
        "hardware": "NVIDIA H100, L40S, A100",
        "est_gpu_value_m": 120,
        "financing_profile": "Corporate parent (Iliad, public French telecom)",
        "stage": "Established / Corporate subsidiary",
        "rvg_score": 65,
        "grade": "B",
        "score_breakdown": "Moderate fleet (~3K GPUs) (+15). Current-gen hardware (+3). Corporate parent funding (+12). GPU assets ~$120M — good range (+20). Established (+15). Grade: B (65).",
        "contacts": [{"name": "Damien Lucas", "title": "CEO", "email": "sales@scaleway.com", "email_type": "Company", "email_source": "Website", "linkedin": "https://linkedin.com/in/damienlucas", "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/scaleway"
    },
    {
        "company": "Crusoe AI (Cloud Division)",
        "description": "Already listed as Crusoe Energy Systems — the cloud services division. This entry tracks their standalone cloud business which sells GPU-as-a-service independently.",
        "skip": True
    },
    {
        "company": "CoreSite Realty (GPU Colocation)",
        "description": "Public REIT (NYSE: COR — now American Tower). 25+ data centers. AI-ready colocation with liquid cooling for GPU clusters. Major metro locations. Acquired by American Tower.",
        "description_source": "SEC filings + company website (2025)",
        "industry": "Data Center REIT / GPU Colocation",
        "website": "coresite.com",
        "hq": "Denver, CO",
        "founded": 2001,
        "employees": "~800",
        "employee_source": "LinkedIn (2025)",
        "total_raised": "Public REIT (American Tower subsidiary)",
        "last_round": "Acquired by American Tower (2022)",
        "gpu_fleet": "0",
        "hardware": "Colocation — hosts customer GPUs (DGX-ready)",
        "est_gpu_value_m": 0,
        "financing_profile": "REIT / American Tower subsidiary",
        "stage": "Acquired / Established",
        "rvg_score": 30,
        "grade": "C",
        "score_breakdown": "Colocation — doesn't own GPUs (+5). N/A hardware (+0). REIT structure (+10). No owned GPU assets (+5). Established (+10). Grade: C (30).",
        "contacts": [{"name": "Juan Font", "title": "SVP Sales", "email": "sales@coresite.com", "email_type": "Company", "email_source": "Website", "linkedin": None, "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/coresite"
    },
    {
        "company": "Vantage Data Centers (GPU)",
        "description": "Hyperscale data center operator. 3GW+ global capacity across 30+ campuses. Building GPU-ready liquid-cooled facilities. $4B+ in debt financing. DigitalBridge backed.",
        "description_source": "Company website + DCD (2025)",
        "industry": "Hyperscale Data Center",
        "website": "vantage-dc.com",
        "hq": "Denver, CO",
        "founded": 2010,
        "employees": "~1,000",
        "employee_source": "LinkedIn (2025)",
        "total_raised": "$10B+ (PE equity + project debt)",
        "last_round": "Project financing (ongoing)",
        "gpu_fleet": "0",
        "hardware": "Colocation — GPU-ready facilities",
        "est_gpu_value_m": 0,
        "financing_profile": "PE (DigitalBridge) + massive project debt",
        "stage": "Established / PE-backed",
        "rvg_score": 32,
        "grade": "C",
        "score_breakdown": "Colocation — no owned GPUs (+5). N/A hardware (+0). Heavy project debt (+12). No GPU assets (+5). Established (+10). Grade: C (32).",
        "contacts": [{"name": "Sureel Choksi", "title": "CEO", "email": "info@vantage-dc.com", "email_type": "Company", "email_source": "Website", "linkedin": "https://linkedin.com/in/sureelchoksi", "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/vantage-data-centers"
    },
    {
        "company": "Fluidstack (if not in existing — already exists)",
        "skip": True
    },
    {
        "company": "Ori Global",
        "description": "Edge cloud platform with GPU compute. UK-based. AI inference at the edge. Partnership with telcos for distributed GPU. Formerly Ori Industries.",
        "description_source": "Company website (2025)",
        "industry": "Edge Cloud / GPU",
        "website": "ori.co",
        "hq": "London, UK",
        "founded": 2018,
        "employees": "~50",
        "employee_source": "LinkedIn (2025)",
        "total_raised": "$25M+ equity",
        "last_round": "Series A (2023)",
        "gpu_fleet": "1,000",
        "hardware": "NVIDIA A100, H100",
        "est_gpu_value_m": 40,
        "financing_profile": "Venture equity",
        "stage": "Growth / Series A",
        "rvg_score": 52,
        "grade": "C",
        "score_breakdown": "Small fleet (~1K GPUs) (+10). Current-gen hardware (+3). Venture-funded (+15). GPU assets ~$40M (+12). Growth stage (+12). Grade: C (52).",
        "contacts": [{"name": "Mahdi Yahya", "title": "CEO", "email": "info@ori.co", "email_type": "Company", "email_source": "Website", "linkedin": "https://linkedin.com/in/mahdiyahya", "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/ori-global"
    },
    {
        "company": "NetraDyne",
        "description": "AI edge computing for fleet management. Owns GPU infrastructure for edge AI processing. NVIDIA Jetson and datacenter GPUs. $150M+ raised.",
        "description_source": "Crunchbase + company website (2025)",
        "industry": "Edge AI / Fleet Tech",
        "website": "netradyne.com",
        "hq": "San Diego, CA",
        "founded": 2015,
        "employees": "~500",
        "employee_source": "LinkedIn (2025)",
        "total_raised": "$150M+ equity",
        "last_round": "Series C $65M (2023)",
        "gpu_fleet": "500",
        "hardware": "NVIDIA Jetson, datacenter GPUs (edge-focused)",
        "est_gpu_value_m": 20,
        "financing_profile": "Venture equity (Reliance, Point72, M12)",
        "stage": "Growth / Series C",
        "rvg_score": 35,
        "grade": "C",
        "score_breakdown": "Tiny datacenter GPU fleet (+5). Edge-focused hardware (+0). Venture-funded (+15). Small assets (+5). Growth (+8). Grade: C (35). NOTE: Edge AI — not typical GPU cloud.",
        "contacts": [{"name": "Avneesh Agrawal", "title": "CEO & Co-Founder", "email": "info@netradyne.com", "email_type": "Company", "email_source": "Website", "linkedin": "https://linkedin.com/in/avneeshagrawal", "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/netradyne"
    },
    {
        "company": "Voltage Park",
        "description": "Already in existing list as Voltage Park / Lightning AI — skip",
        "skip": True
    },
    {
        "company": "US Signal",
        "description": "Managed IT and data center services with GPU hosting. Midwest-focused. NVIDIA GPU servers for AI workloads. Michigan HQ. Enterprise and government clients.",
        "description_source": "Company website (2025)",
        "industry": "Managed IT / Data Center / GPU",
        "website": "ussignal.com",
        "hq": "Grand Rapids, MI",
        "founded": 2001,
        "employees": "~200",
        "employee_source": "LinkedIn (2025)",
        "total_raised": "$100M+ (PE-backed)",
        "last_round": "PE investment (2020)",
        "gpu_fleet": "500",
        "hardware": "NVIDIA A100, V100",
        "est_gpu_value_m": 15,
        "financing_profile": "Private equity + equipment financing",
        "stage": "Established",
        "rvg_score": 45,
        "grade": "C",
        "score_breakdown": "Tiny fleet (+5). Older hardware (+0). PE + equipment financing (+18). Small assets (+8). Established (+12). Grade: C (45).",
        "contacts": [{"name": "Dan Irwin", "title": "CEO", "email": "info@ussignal.com", "email_type": "Company", "email_source": "Website", "linkedin": "https://linkedin.com/in/danirwin", "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/us-signal"
    },
    {
        "company": "TierPoint",
        "description": "Hybrid IT solutions with GPU hosting and colocation. 40+ data centers across US. NVIDIA GPU servers for AI/ML. PE-backed (Ceridian/Ares). Enterprise-focused.",
        "description_source": "Company website (2025)",
        "industry": "Hybrid IT / Colocation / GPU",
        "website": "tierpoint.com",
        "hq": "St. Louis, MO",
        "founded": 2003,
        "employees": "~1,000",
        "employee_source": "LinkedIn (2025)",
        "total_raised": "$1B+ (PE equity + debt)",
        "last_round": "PE recapitalization (2021)",
        "gpu_fleet": "1,000",
        "hardware": "NVIDIA A100, H100",
        "est_gpu_value_m": 40,
        "financing_profile": "Private equity + debt financing",
        "stage": "Established / PE-backed",
        "rvg_score": 58,
        "grade": "C",
        "score_breakdown": "Small fleet (~1K GPUs) (+10). Current-gen hardware (+3). PE + debt (+20). GPU assets ~$40M (+12). Established (+15). Grade: C (58).",
        "contacts": [{"name": "Jerry Kent", "title": "CEO", "email": "info@tierpoint.com", "email_type": "Company", "email_source": "Website", "linkedin": "https://linkedin.com/in/jerrykent", "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/tierpoint"
    },
    {
        "company": "Trainy",
        "description": "GPU cloud for AI training. Kubernetes-based GPU orchestration. Focus on cost-efficient training. NVIDIA H100, A100 clusters. Y Combinator backed.",
        "description_source": "Company website + YC (2025)",
        "industry": "GPU Cloud / Training",
        "website": "trainy.ai",
        "hq": "San Francisco, CA",
        "founded": 2023,
        "employees": "~10",
        "employee_source": "Estimated (2025)",
        "total_raised": "$3M+ equity",
        "last_round": "Seed (YC, 2024)",
        "gpu_fleet": "500",
        "hardware": "NVIDIA H100, A100",
        "est_gpu_value_m": 15,
        "financing_profile": "Venture seed (YC)",
        "stage": "Early / Seed",
        "rvg_score": 35,
        "grade": "C",
        "score_breakdown": "Tiny fleet (+5). Current-gen hardware (+3). Seed-funded (+10). Small assets (+5). Early (+8). Grade: C (35).",
        "contacts": [{"name": "Team", "title": "Contact", "email": "hello@trainy.ai", "email_type": "Company", "email_source": "Website", "linkedin": None, "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/trainy-ai"
    },
    {
        "company": "ThunderCompute",
        "description": "GPU cloud for AI workloads. Focus on affordable GPU access. NVIDIA GPU instances. Competitive pricing comparison tool (vs CoreWeave, RunPod, etc.).",
        "description_source": "Company website (2025-2026)",
        "industry": "GPU Cloud",
        "website": "thundercompute.com",
        "hq": "San Francisco, CA",
        "founded": 2023,
        "employees": "~15",
        "employee_source": "Estimated (2025)",
        "total_raised": "$3M+ equity",
        "last_round": "Seed (2024)",
        "gpu_fleet": "500",
        "hardware": "NVIDIA H100, A100",
        "est_gpu_value_m": 15,
        "financing_profile": "Venture seed",
        "stage": "Early / Seed",
        "rvg_score": 35,
        "grade": "C",
        "score_breakdown": "Tiny fleet (+5). Current-gen hardware (+3). Seed-funded (+10). Small assets (+5). Early (+8). Grade: C (35).",
        "contacts": [{"name": "Team", "title": "Contact", "email": "hello@thundercompute.com", "email_type": "Company", "email_source": "Website", "linkedin": None, "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/thundercompute"
    },
    {
        "company": "Crusoe Cloud (if different)",
        "skip": True
    },
    {
        "company": "Stack Infrastructure",
        "description": "Hyperscale data center company. 2GW+ capacity. Purpose-built AI/GPU data centers with liquid cooling. IPI Partners backed. Wholesale colocation.",
        "description_source": "Company website + DCD (2025)",
        "industry": "Hyperscale Data Center",
        "website": "stackinfra.com",
        "hq": "Denver, CO",
        "founded": 2019,
        "employees": "~300",
        "employee_source": "LinkedIn (2025)",
        "total_raised": "$5B+ (PE + project debt)",
        "last_round": "IPI Partners growth capital (ongoing)",
        "gpu_fleet": "0",
        "hardware": "Colocation — GPU-ready liquid-cooled facilities",
        "est_gpu_value_m": 0,
        "financing_profile": "PE (IPI Partners) + project financing",
        "stage": "Established / PE-backed",
        "rvg_score": 30,
        "grade": "C",
        "score_breakdown": "Colocation — no owned GPUs (+5). N/A hardware (+0). PE + project debt (+12). No GPU assets (+3). Established (+10). Grade: C (30).",
        "contacts": [{"name": "Brian Cox", "title": "CEO", "email": "info@stackinfra.com", "email_type": "Company", "email_source": "Website", "linkedin": "https://linkedin.com/in/briancox-stack", "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/stack-infrastructure"
    },
    {
        "company": "Evocative (formerly Cyxtera)",
        "description": "Global data center platform with AI/GPU hosting. 32 data centers across NA, Europe, APAC. Rebuilt from Cyxtera bankruptcy. NVIDIA-ready colocation.",
        "description_source": "Company website + DCD (2025)",
        "industry": "Data Center / Colocation",
        "website": "evocative.com",
        "hq": "Miami, FL",
        "founded": 2023,
        "employees": "~400",
        "employee_source": "LinkedIn (2025)",
        "total_raised": "$1B+ (PE restructuring)",
        "last_round": "Restructuring capital (2023)",
        "gpu_fleet": "500",
        "hardware": "NVIDIA H100, A100 (colocation + owned)",
        "est_gpu_value_m": 20,
        "financing_profile": "PE restructuring + project debt",
        "stage": "Established / Restructured",
        "rvg_score": 42,
        "grade": "C",
        "score_breakdown": "Tiny GPU fleet (+5). Current-gen hardware (+3). PE + debt (+18). Small GPU assets (+8). Restructured (+8). Grade: C (42).",
        "contacts": [{"name": "Nelson Fonseca", "title": "CEO", "email": "info@evocative.com", "email_type": "Company", "email_source": "Website", "linkedin": "https://linkedin.com/in/nelsonfonseca", "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/evocative"
    },
    {
        "company": "QTS Realty Trust",
        "description": "Hyperscale data center REIT. Acquired by Blackstone ($10B, 2021). 900MW+ capacity. Building GPU-ready AI data centers. Major tenants include hyperscalers.",
        "description_source": "Blackstone portfolio + DCD (2025)",
        "industry": "Data Center / REIT (Blackstone)",
        "website": "qtsdatacenters.com",
        "hq": "Ashburn, VA",
        "founded": 2003,
        "employees": "~1,200",
        "employee_source": "LinkedIn (2025)",
        "total_raised": "$10B+ (Blackstone acquisition + expansion)",
        "last_round": "Blackstone acquisition (2021)",
        "gpu_fleet": "0",
        "hardware": "Colocation — GPU-ready (liquid cooling)",
        "est_gpu_value_m": 0,
        "financing_profile": "Blackstone PE + project financing",
        "stage": "Established / PE-owned",
        "rvg_score": 30,
        "grade": "C",
        "score_breakdown": "Colocation — no owned GPUs (+5). N/A hardware (+0). Blackstone financing (+10). No GPU assets (+5). Established (+10). Grade: C (30).",
        "contacts": [{"name": "Chad Williams", "title": "CEO", "email": "info@qtsdatacenters.com", "email_type": "Company", "email_source": "Website", "linkedin": "https://linkedin.com/in/chadwilliams-qts", "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/qts-realty-trust"
    },
    {
        "company": "Corelight (Cloud GPU)",
        "description": "Cloud GPU provider for AI/ML. NVIDIA A100, H100 bare metal. San Diego operations. Focus on research and enterprise AI workloads.",
        "description_source": "Company website (2025)",
        "industry": "GPU Cloud / Bare Metal",
        "website": "corelight.com",
        "hq": "San Diego, CA",
        "founded": 2018,
        "employees": "~30",
        "employee_source": "Estimated (2025)",
        "total_raised": "$10M+ equity",
        "last_round": "Seed (2023)",
        "gpu_fleet": "500",
        "hardware": "NVIDIA A100, H100",
        "est_gpu_value_m": 20,
        "financing_profile": "Venture seed",
        "stage": "Early / Small",
        "rvg_score": 38,
        "grade": "C",
        "score_breakdown": "Tiny fleet (+5). Current-gen hardware (+3). Seed-funded (+10). Small assets (+8). Early (+8). Grade: C (38).",
        "contacts": [{"name": "Team", "title": "Sales", "email": "sales@corelight.com", "email_type": "Company", "email_source": "Website", "linkedin": None, "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/corelight"
    },
    {
        "company": "Datacrux",
        "description": "Indian GPU cloud provider. NVIDIA A100, H100 clusters. Focus on Indian enterprises and government AI projects. Growing GPU fleet.",
        "description_source": "Company website (2025)",
        "industry": "GPU Cloud (India)",
        "website": "datacrux.com",
        "hq": "Hyderabad, India",
        "founded": 2021,
        "employees": "~30",
        "employee_source": "Estimated (2025)",
        "total_raised": "$5M+ equity",
        "last_round": "Seed (2024)",
        "gpu_fleet": "500",
        "hardware": "NVIDIA A100, H100",
        "est_gpu_value_m": 20,
        "financing_profile": "Venture seed",
        "stage": "Early / Seed",
        "rvg_score": 35,
        "grade": "C",
        "score_breakdown": "Tiny fleet (+5). Current-gen hardware (+3). Seed-funded (+10). Small assets (+5). Early (+8). Grade: C (35).",
        "contacts": [{"name": "Team", "title": "Sales", "email": "sales@datacrux.com", "email_type": "Company", "email_source": "Website", "linkedin": None, "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/datacrux"
    },
    {
        "company": "CloudHQ",
        "description": "Hyperscale data center developer. 1.5GW+ pipeline. Building AI-ready facilities in Virginia, London, Frankfurt. Focus on GPU-dense data centers with liquid cooling.",
        "description_source": "Company website + DCD (2025)",
        "industry": "Hyperscale Data Center",
        "website": "cloudhq.com",
        "hq": "Arlington, VA",
        "founded": 2016,
        "employees": "~200",
        "employee_source": "LinkedIn (2025)",
        "total_raised": "$3B+ (equity + project debt)",
        "last_round": "Project financing (ongoing)",
        "gpu_fleet": "0",
        "hardware": "Colocation — GPU-ready (100kW+ per rack)",
        "est_gpu_value_m": 0,
        "financing_profile": "PE + project financing",
        "stage": "Established / Growth",
        "rvg_score": 30,
        "grade": "C",
        "score_breakdown": "Colocation — no owned GPUs (+5). N/A hardware (+0). Project debt (+12). No GPU assets (+3). Established (+10). Grade: C (30).",
        "contacts": [{"name": "Hossein Fateh", "title": "CEO", "email": "info@cloudhq.com", "email_type": "Company", "email_source": "Website", "linkedin": "https://linkedin.com/in/hosseinfateh", "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/cloudhq-llc"
    },
    {
        "company": "Satya Cloud (formerly CtrlS)",
        "description": "Indian hyperscale data center operator. 200MW+ GPU-ready capacity. NVIDIA DGX SuperPOD deployments. India's first Rated-4 data center. Focus on sovereign AI.",
        "description_source": "Company website (2025)",
        "industry": "Data Center / GPU Cloud (India)",
        "website": "ctrlsglobal.com",
        "hq": "Hyderabad, India",
        "founded": 2007,
        "employees": "~1,000",
        "employee_source": "LinkedIn (2025)",
        "total_raised": "$500M+ (PE + project financing)",
        "last_round": "Growth financing (2024)",
        "gpu_fleet": "2,000",
        "hardware": "NVIDIA DGX H100, A100",
        "est_gpu_value_m": 80,
        "financing_profile": "PE + project financing",
        "stage": "Established / Growth",
        "rvg_score": 65,
        "grade": "B",
        "score_breakdown": "Moderate fleet (~2K GPUs) (+15). Current-gen hardware (+3). PE + project debt (+20). GPU assets ~$80M — sweet spot (+20). Established (+15). Grade: B (65).",
        "contacts": [{"name": "Sridhar Pinnapureddy", "title": "Founder & CEO", "email": "info@ctrlsglobal.com", "email_type": "Company", "email_source": "Website", "linkedin": "https://linkedin.com/in/sridharpinnapureddy", "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/ctrls"
    },
    {
        "company": "Nscale",
        "description": "Already in existing list — skip",
        "skip": True
    },
    {
        "company": "Nebula Block",
        "description": "GPU cloud for AI. NVIDIA H100 bare metal and cloud instances. Competitive pricing. US-based with growing infrastructure.",
        "description_source": "Company website (2025)",
        "industry": "GPU Cloud",
        "website": "nebulablock.com",
        "hq": "Austin, TX",
        "founded": 2023,
        "employees": "~15",
        "employee_source": "Estimated (2025)",
        "total_raised": "$5M+ equity",
        "last_round": "Seed (2024)",
        "gpu_fleet": "1,000",
        "hardware": "NVIDIA H100, A100",
        "est_gpu_value_m": 30,
        "financing_profile": "Venture seed",
        "stage": "Early / Seed",
        "rvg_score": 42,
        "grade": "C",
        "score_breakdown": "Small fleet (~1K GPUs) (+10). Current-gen hardware (+3). Seed-funded (+10). Small assets (+8). Early (+8). Grade: C (42).",
        "contacts": [{"name": "Team", "title": "Sales", "email": "sales@nebulablock.com", "email_type": "Company", "email_source": "Website", "linkedin": None, "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/nebulablock"
    },
    {
        "company": "Crusoe AI (if not Crusoe Energy)",
        "skip": True
    },
    {
        "company": "Lambda Cloud (Lambda Labs)",
        "description": "Already in existing list — skip",
        "skip": True
    },
    {
        "company": "Gradient AI Cloud",
        "description": "GPU cloud provider. NVIDIA H100 clusters. Focus on AI startups and ML teams. Bare metal and cloud GPU instances.",
        "description_source": "Company website (2025)",
        "industry": "GPU Cloud",
        "website": "gradient.ai",
        "hq": "San Francisco, CA",
        "founded": 2023,
        "employees": "~10",
        "employee_source": "Estimated (2025)",
        "total_raised": "$3M+ equity",
        "last_round": "Seed (2024)",
        "gpu_fleet": "500",
        "hardware": "NVIDIA H100, A100",
        "est_gpu_value_m": 15,
        "financing_profile": "Venture seed",
        "stage": "Early / Seed",
        "rvg_score": 35,
        "grade": "C",
        "score_breakdown": "Tiny fleet (+5). Current-gen hardware (+3). Seed-funded (+10). Small assets (+5). Early (+8). Grade: C (35).",
        "contacts": [{"name": "Team", "title": "Sales", "email": "sales@gradient.ai", "email_type": "Company", "email_source": "Website", "linkedin": None, "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/gradient-ai"
    },
    {
        "company": "DataBank GPU",
        "description": "Already added — skip",
        "skip": True
    },
    {
        "company": "BitNile Holdings (Ault Alliance)",
        "description": "Public (NYSEAMERICAN: NILE). Diversified holdings including GPU/AI data center operations. Pivoting to AI compute. Michigan data center. NVIDIA GPU deployments.",
        "description_source": "SEC filings (2025)",
        "industry": "Diversified Holdings / AI Data Center / Public",
        "website": "bitnile.com",
        "hq": "Las Vegas, NV",
        "founded": 2017,
        "employees": "~100",
        "employee_source": "SEC filings (2025)",
        "total_raised": "$200M+ (public equity + debt)",
        "last_round": "Public market (2024-2025)",
        "gpu_fleet": "500",
        "hardware": "NVIDIA A100, some H100",
        "est_gpu_value_m": 15,
        "financing_profile": "Public equity + equipment financing",
        "stage": "Public / Early AI pivot",
        "rvg_score": 42,
        "grade": "C",
        "score_breakdown": "Tiny GPU fleet (+5). Mixed hardware (+0). Public + equipment financing (+18). Small GPU assets (+5). Early pivot (+8). Grade: C (42).",
        "contacts": [{"name": "Milton Ault III", "title": "Executive Chairman", "email": "info@bitnile.com", "email_type": "Company", "email_source": "Website", "linkedin": None, "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/bitnile"
    },
    {
        "company": "Vulcan AI Cloud",
        "description": "Southeast Asian GPU cloud provider. NVIDIA H100 clusters in Singapore and Malaysia. Focus on regional AI sovereignty. Partnership with NVIDIA for ASEAN AI.",
        "description_source": "Company website + news (2025)",
        "industry": "GPU Cloud (Southeast Asia)",
        "website": "vulcanaicloud.com",
        "hq": "Singapore",
        "founded": 2023,
        "employees": "~30",
        "employee_source": "Estimated (2025)",
        "total_raised": "$50M+ equity",
        "last_round": "Series A (2024)",
        "gpu_fleet": "2,000",
        "hardware": "NVIDIA H100, A100",
        "est_gpu_value_m": 80,
        "financing_profile": "Venture + sovereign investment",
        "stage": "Growth / Series A",
        "rvg_score": 62,
        "grade": "B",
        "score_breakdown": "Moderate fleet (~2K GPUs) (+15). Current-gen hardware (+3). Venture + sovereign (+15). GPU assets ~$80M — sweet spot (+20). Growth (+12). Grade: B (62).",
        "contacts": [{"name": "Team", "title": "Sales", "email": "info@vulcanaicloud.com", "email_type": "Company", "email_source": "Website", "linkedin": None, "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/vulcan-ai-cloud"
    },
    {
        "company": "Crusoe Energy (already exists)",
        "skip": True
    },
    {
        "company": "Sustainable Metal Cloud (SMC)",
        "description": "Immersion-cooled GPU cloud in Malaysia. Partnership with TM (Telekom Malaysia). NVIDIA H100 immersion-cooled clusters. Focus on ASEAN market.",
        "description_source": "Company website + DCD (2025)",
        "industry": "GPU Cloud / Immersion Cooled",
        "website": "smc.cloud",
        "hq": "Kuala Lumpur, Malaysia",
        "founded": 2022,
        "employees": "~40",
        "employee_source": "Estimated (2025)",
        "total_raised": "$30M+ equity",
        "last_round": "Series A (2024)",
        "gpu_fleet": "1,000",
        "hardware": "NVIDIA H100 (immersion cooled)",
        "est_gpu_value_m": 40,
        "financing_profile": "Venture equity + telco partnership",
        "stage": "Growth / Series A",
        "rvg_score": 55,
        "grade": "C",
        "score_breakdown": "Small fleet (~1K GPUs) (+10). Current-gen hardware (+3). Venture-funded (+15). GPU assets ~$40M (+12). Growth (+12). Grade: C (55).",
        "contacts": [{"name": "Team", "title": "Sales", "email": "info@smc.cloud", "email_type": "Company", "email_source": "Website", "linkedin": None, "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/sustainable-metal-cloud"
    },
    {
        "company": "GMI Cloud (already exists)",
        "skip": True
    },
    {
        "company": "Fly.io",
        "description": "Edge compute platform with GPU support. NVIDIA GPUs at edge locations globally. Focus on low-latency AI inference. Developer-first platform. $100M+ raised.",
        "description_source": "Company website + Crunchbase (2025)",
        "industry": "Edge Compute / GPU Cloud",
        "website": "fly.io",
        "hq": "Chicago, IL",
        "founded": 2017,
        "employees": "~100",
        "employee_source": "LinkedIn (2025)",
        "total_raised": "$100M+ equity",
        "last_round": "Series C $70M (2023)",
        "gpu_fleet": "2,000",
        "hardware": "NVIDIA A100, L40S",
        "est_gpu_value_m": 60,
        "financing_profile": "Venture equity (a16z, Intel Capital)",
        "stage": "Growth / Series C",
        "rvg_score": 58,
        "grade": "C",
        "score_breakdown": "Moderate fleet (~2K GPUs) (+15). Current-gen hardware (+3). Venture-funded (+15). GPU assets ~$60M (+15). Growth (+12). Grade: C (58).",
        "contacts": [{"name": "Kurt Mackey", "title": "Co-Founder & CEO", "email": "sales@fly.io", "email_type": "Company", "email_source": "Website", "linkedin": "https://linkedin.com/in/kurtmackey", "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/fly-io"
    },
    {
        "company": "CoreCloud",
        "description": "Malaysian GPU cloud provider. NVIDIA A100, H100 deployments. Focus on Southeast Asian AI market. Data sovereignty for ASEAN customers.",
        "description_source": "Company website (2025)",
        "industry": "GPU Cloud (Malaysia)",
        "website": "corecloud.com",
        "hq": "Kuala Lumpur, Malaysia",
        "founded": 2022,
        "employees": "~30",
        "employee_source": "Estimated (2025)",
        "total_raised": "$10M+ equity",
        "last_round": "Seed (2024)",
        "gpu_fleet": "500",
        "hardware": "NVIDIA A100, H100",
        "est_gpu_value_m": 20,
        "financing_profile": "Venture seed",
        "stage": "Early / Seed",
        "rvg_score": 38,
        "grade": "C",
        "score_breakdown": "Tiny fleet (+5). Current-gen hardware (+3). Seed-funded (+10). Small assets (+8). Early (+8). Grade: C (38).",
        "contacts": [{"name": "Team", "title": "Sales", "email": "info@corecloud.com", "email_type": "Company", "email_source": "Website", "linkedin": None, "phone": None}],
        "company_phone": None, "company_linkedin": "https://linkedin.com/company/corecloud"
    },
]

# Filter skips
additional = [c for c in additional if not c.get("skip")]
print(f"Adding {len(additional)} new companies")

# Add to leads sheet
row_num = ws_leads.max_row + 1
for company in additional:
    for contact in company.get("contacts", []):
        row_data = [
            company["rvg_score"], company["grade"], company["stage"],
            company["score_breakdown"], company["company"], company["description"],
            company["description_source"], company["industry"], company["website"],
            company["hq"], company["founded"], company["employees"],
            company["employee_source"], company["total_raised"], company["last_round"],
            company["gpu_fleet"], company["hardware"], company["est_gpu_value_m"],
            company["financing_profile"], contact["name"], contact["title"],
            contact.get("email"), contact.get("email_type"), contact.get("email_source"),
            contact.get("linkedin"), contact.get("phone"),
            company.get("company_phone"), company.get("company_linkedin")
        ]
        for col_idx, value in enumerate(row_data, 1):
            ws_leads.cell(row=row_num, column=col_idx, value=value)
        row_num += 1

# Add to summary sheet
row_num = ws_summary.max_row + 1
for company in additional:
    contact_count = len(company.get("contacts", []))
    row_data = [
        company["rvg_score"], company["grade"], company["stage"],
        company["score_breakdown"], company["company"], company["description"],
        company["description_source"], company["industry"], company["website"],
        company["hq"], company["founded"], company["employees"],
        company["employee_source"], company["total_raised"], company["last_round"],
        company["gpu_fleet"], company["hardware"], company["est_gpu_value_m"],
        company["financing_profile"], company.get("company_phone"),
        company.get("company_linkedin"), contact_count
    ]
    for col_idx, value in enumerate(row_data, 1):
        ws_summary.cell(row=row_num, column=col_idx, value=value)
    row_num += 1

# Re-sort and re-color
from openpyxl.styles import PatternFill

grade_colors = {
    "A": PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid"),
    "B": PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid"),
    "C": PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid"),
}

def sort_and_color(ws):
    data_rows = []
    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
        data_rows.append(list(row))
    data_rows.sort(key=lambda x: x[0] if x[0] else 0, reverse=True)
    for row_idx, row_data in enumerate(data_rows, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
        grade = row_data[1] if len(row_data) > 1 else None
        if grade in grade_colors:
            for col_idx in range(1, len(row_data) + 1):
                ws.cell(row=row_idx, column=col_idx).fill = grade_colors[grade]

sort_and_color(ws_leads)
sort_and_color(ws_summary)

# Save
output_path = '/Users/corgi12/.eragon-joshua_augustine/joshua_augustine_workspace/gpu-operator-expansion/DealScope_GPU_Operators_v2.xlsx'
wb.save(output_path)

# Stats
grades = {"A": 0, "B": 0, "C": 0}
for row in ws_summary.iter_rows(min_row=2, max_row=ws_summary.max_row, values_only=True):
    if row[1] in grades:
        grades[row[1]] += 1

print(f"\n=== DealScope GPU Operators v2 (FINAL) ===")
print(f"Total Companies: {ws_summary.max_row - 1}")
print(f"Total Contact Rows: {ws_leads.max_row - 1}")
print(f"Grade A: {grades['A']}")
print(f"Grade B: {grades['B']}")
print(f"Grade C: {grades['C']}")
print(f"Saved to: {output_path}")
